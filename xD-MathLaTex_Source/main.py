import os

import datetime

import win32com.client as win32com

from win32com.client import client as win32

import win32api

import re

import sys

import shutil

import pythoncom

import tempfile

import wmi

import docx

import random

import qrcode

import bisect

from spire.doc import Section, Paragraph, Table, TextRange, Shape, FileFormat, HorizontalAlignment, Regex

from spire.doc import UnderlineStyle

from spire.doc import Document as SpireDocument

from spire.doc import Color

from docx import Document

from docx.shared import Pt

from docx.enum.text import WD_PARAGRAPH_ALIGNMENT

from docx.shared import RGBColor

from openpyxl import Workbook

from openpyxl.styles import Font

from openpyxl.styles import Alignment

from openpyxl.styles import Border, Side

from openpyxl.styles import PatternFill

from openpyxl.styles import GradientFill

from openpyxl.utils import get_column_letter

from openpyxl import load_workbook

import tkinter as tk

from tkinter import ttk, Label, filedialog, Entry, Button, Frame, Listbox, Scrollbar, messagebox

from tkinter import simpledialog

import webbrowser

from core import functions as vbf

from answers import dapan_2025 as DA

from answers import dapan_danh_dau as DA2

from formatting import chuan_hoa as chd

from conversion import pdf as pdf

from tools import rename as vbrn

from registration import dangki_pc as dangki

from registration import dangki_usb as dangkiUSB

from conversion import word_to_pp as wtp

from mixing import mixeasy as vme

from mixing import mixeasy_file_mau as fmau

from tools import qr_code as Qr

from tools import tool_by_docx as docxtool

from tools import tool_by_spr as sprtool

from mixing import mix_docx_check_data as docxcheck

from mixing import mix_spr as mspr

from mixing import mix_spr_en as mspr_EN

from html_generators import trac_nghiem_online as tn_online

from html_generators import trac_nghiem_offline as tn_offline

from html_generators import vong_quay as quay

from html_generators import bang_diem as bangdiem

from html_generators import dong_ho_dem_nguoc as clockcd

from conversion import omml_to_latex as OMML_latex

from conversion import latex_to_omml as latex_OMML

from conversion import pdf_to_word as w2pdf

phien_ban = 'V2026.04.25'

addin_name = 'MixEasy_Addin_V20.0.61'

ngay_dung_thu = datetime.date(2026, 5, 30)

Quang_cao_1 = 'Tác giải: Nguyễn Nhật Huy \n Điện thoại, Zalo: 0914282232'

Quang_cao_2 = 'Tác giải: Nguyễn Nhật Huy\nĐiện thoại, Zalo: 0914282232\nGiá phần mềm: \n   + Chạy trực tiếp trên máy tính: 50 nghìn/1 máy tính/1 năm\n   + Chạy trực tiếp trên USB: 100 nghìn/1 USB/1 năm\nSố tài khoản: 0161 000 376 724\nNgân hàng: Vietcombank'

check_phien_ban = phien_ban[-3:]

if check_phien_ban == 'hoc':

    Quang_cao = Quang_cao_1

else:

    Quang_cao = Quang_cao_2

ngay_het_han1 = dangki.read_key_use_ngay_het_han(ngay_dung_thu)

ngay_het_han2 = dangkiUSB.read_key_use_ngay_het_han(ngay_dung_thu)

ngay_het_han = max(ngay_het_han1, ngay_het_han2)

formatted_date_het_han = ngay_het_han.strftime('%d-%m-%Y')

thongtin_dk_full = f'''Thời hạn sử dụng phần mềm của bạn đến\n {formatted_date_het_han}'''

word = vbf.khoi_tao_word()

root = tk.Tk()

root.withdraw()



def cap_nhat_time():

    entry_lantron.delete(0, tk.END)

    entry_lantron.insert(0, datetime.datetime.now().strftime('%Hh%M'))





def cap_nhat_time_EN():

    entry_lantron_EN.delete(0, tk.END)

    entry_lantron_EN.insert(0, datetime.datetime.now().strftime('%Hh%M'))





def creat_Qr_TNmaker(root):

    Qr.creat_Qr_TNmaker(root)





def save_as_de_to_output_me(doc, word, output_path):

    doc.SaveAs2(FileName = output_path)

    messagebox.showinfo('Thông báo', 'Tài liệu đã được lưu với tên MỚI')

    return None

# WARNING: Decompyle incomplete





def save_as_de_from_entry_me(input_entry1_tab0, entry_input):

    input_de = input_entry1_tab0.get()

# WARNING: Decompyle incomplete





def save_as_Me():

    save_as_de_from_entry_me(input_entry1_tab0, entry_input)





def save_as_Me_EN():

    save_as_de_from_entry_me(input_entry1_tab0_EN, entry_input_EN)





def browse_file(entry_input):

    input_file = filedialog.askopenfilename(filetypes = [

        ('Word files', '*.docx')])

    if not input_file:

        return None

    input_file = os.path.normpath(input_file)

    entry_input.delete(0, tk.END)

    entry_input.insert(tk.END, input_file)





def browse_file_Me():

    browse_file(entry_input)





def browse_file_Me_EN():

    browse_file(entry_input_EN)





def generate_numbers(entry_num, entry_made_start, entry_newname, entry_khoang_cach_made):

    count = int(entry_num.get())

    if count <= 0:

        raise ValueError

    made_first = int(entry_made_start.get())

    kcach = int(entry_khoang_cach_made.get())

    numbers = []

# WARNING: Decompyle incomplete





def tao_made_tudong():

    entry_khoang_cach_made

    generate_numbers(entry_num, entry_made_start, entry_newname, entry_khoang_cach_made)





def tao_made_tudong_EN():

    generate_numbers(entry_num_EN, entry_made_start_EN, entry_newname_EN, entry_khoang_cach_made_EN)





def tron_and_xuat_de_Me():

    mspr.tron_and_xuat_de(check1_1_var, check1_2_var, check2_1_var, check2_2_var, check3_1_var, check4_1_var, check5_1_var, check6_1_var, check7_1_var, entry_input, entry_lantron, entry_newname)





def tron_and_xuat_de_Me_NN():

    mspr_EN.tron_and_xuat_de_NN(entry_input_EN, entry_lantron_EN, entry_newname_EN, check6_1_EN_var, check7_1_EN_var, check8_1_EN_var)





def copy_add_dotm_for365(addin_name):

    vbf.copy_add_dotm(addin_name)





def open_link():

    link = 'https://drive.google.com/drive/folders/1y7BCM9uyfPEpDOLcODAHQ-e6XNe1MPWT?usp=sharing'

    webbrowser.open_new(link)





def open_link_face_book():

    link = 'https://www.facebook.com/MixEasyTool'

    webbrowser.open_new(link)





def open_link_url(url):

    webbrowser.open_new(url)





def check_sudung():

    if dangki.check_het_han_dangki(ngay_het_han):

        return True

    return False





def kiem_tra_cho_su_dung(root):

    kt = check_sudung()

    if kt:

        open_giao_dien(phien_ban)

        return None

    opendangki(phien_ban)





def opendangki(phien_ban):

    root.deiconify()

    root.title(f'''MixEasy {phien_ban}''')

    root.resizable(False, False)

    label_quahan = tk.Label(root, text = f'''Hạn dùng của bạn là: {formatted_date_het_han},\n nếu muốn dùng thêm bạn vui lòng đăng kí lại''', foreground = 'blue', background = 'yellowgreen')

    label_quahan.grid(row = 0, column = 0, columnspan = 2, padx = 25, pady = 5)

    thongtin_label = tk.Label(root, text = Quang_cao, fg = 'black', bg = 'burlywood', anchor = 'w', justify = 'left')

    thongtin_label.grid(row = 1, column = 0, columnspan = 2, padx = 5, pady = 5, sticky = 'ew')

    button_dk1 = tk.Button(root, text = 'Đăng kí cho máy tính', command = (lambda : dangki.giaodien_dangki(Quang_cao, formatted_date_het_han, root)), bg = 'cyan', fg = 'black', width = 20, height = 2)

    button_dk1.grid(row = 7, column = 0, padx = 5, pady = 5)

    button_dk2 = tk.Button(root, text = 'Đăng kí cho USB', command = (lambda : dangkiUSB.giaodien_dangki(Quang_cao, formatted_date_het_han, root)), bg = 'cyan', fg = 'black', width = 20, height = 2)

    button_dk2.grid(row = 7, column = 1, padx = 5, pady = 5)

    root.wm_attributes('-topmost', True)

    root.mainloop()





def open_giao_dien(phien_ban):

    global listbox_pdf, entry_input, entry_lantron, entry_num, entry_made_start, entry_khoang_cach_made, entry_newname, check1_2_var, check1_1_var, check2_2_var, check2_1_var, check3_1_var, check4_1_var, check5_1_var, check6_1_var, check7_1_var, entry_input_EN, entry_lantron_EN, entry_num_EN, entry_made_start_EN, entry_khoang_cach_made_EN, entry_newname_EN, check6_1_EN_var, check7_1_EN_var, check8_1_EN_var

    root.deiconify()

    root.title(f'''MixEasy {phien_ban}''')

    root.resizable(False, False)

    notebook = tk.ttk.Notebook(root)

    tabB = ttk.Frame(notebook)

    tabA = ttk.Frame(notebook)

    tab_EN = ttk.Frame(notebook)

    tabD = ttk.Frame(notebook)

    style = ttk.Style()

    style.configure('TNotebook.Tab', background = 'whitesmoke', foreground = 'blue', font = ('Arial', 8))

    style.map('TNotebook.Tab', background = [

        ('selected', 'blue')], foreground = [

        ('selected', 'crimson')])

    if check_phien_ban == 'hoc':

        notebook.add(tabA, text = '    TRỘN ĐỀ     ')

        notebook.add(tab_EN, text = '    NGOẠI NGỮ    ')

        notebook.add(tabB, text = '     TIỆN ÍCH      ')

    else:

        notebook.add(tabB, text = '     TIỆN ÍCH      ')

        notebook.add(tabA, text = '    TRỘN ĐỀ     ')

        notebook.add(tab_EN, text = '    NGOẠI NGỮ    ')

    notebook.add(tabD, text = '   ĐĂNG KÍ   ')

    notebook.pack(fill = 'both', expand = True)

    button_reset = tk.Button(tabB, text = 'Tắt chương trình bằng nút này nếu không hoạt động', command = (lambda : vbf.Exit_my_program()), foreground = 'blue', background = 'yellow', width = 40)

    button_reset.grid(row = 1, column = 0, columnspan = 3, padx = 5, pady = 5)

    notebook_tienich = tk.ttk.Notebook(tabB)

    notebook_tienich.grid(row = 2, column = 0, columnspan = 3, padx = 10, pady = 2)

    tab3 = ttk.Frame(notebook_tienich)

    tab4 = ttk.Frame(notebook_tienich)

    tab7 = ttk.Frame(notebook_tienich)

    tab5 = ttk.Frame(notebook_tienich)

    notebook_tienich.add(tab3, text = '   Tools 1   ')

    notebook_tienich.add(tab4, text = '   Tools 2   ')

    notebook_tienich.add(tab7, text = '    Powerpoint    ')

    notebook_tienich.add(tab5, text = '    Convert pdf   ')

    button_0_0 = tk.Button(tab3, text = 'Tạo Bảng ĐA', command = (lambda : DA.DapAn2025(word)), bg = 'aquamarine', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_0_0.grid(row = 0, column = 0, padx = 2, pady = (10, 2))

    button_0_1 = tk.Button(tab3, text = 'Tạo Bảng ĐA\nTN Maker', command = (lambda : DA.DapAn2025_TNmaker(word)), bg = 'aquamarine', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_0_1.grid(row = 0, column = 1, padx = 2, pady = (10, 2))

    button_0_2 = tk.Button(tab3, text = 'Đánh dấu\n Đáp án', command = (lambda : DA2.danh_dau_dap_an_main(root, word)), bg = 'aquamarine', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_0_2.grid(row = 0, column = 2, padx = 2, pady = (10, 2))

    button_0_3 = tk.Button(tab3, text = 'Qr code', command = (lambda : Qr.Qr_tool(root)), bg = 'aquamarine', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_0_3.grid(row = 0, column = 3, padx = 2, pady = (10, 2))

    button_1_0 = tk.Button(tab3, text = 'Fix Câu lỗi\ncấu trúc', command = (lambda : vbf.fix_cau_for_tool(word)), bg = 'whitesmoke', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_1_0.grid(row = 1, column = 0, padx = 2, pady = 2)

    button_1_1 = tk.Button(tab3, text = 'Đậm,Blue Câu\nV.dụ,P.Án', command = (lambda : vbf.Dam_cau_pa(word)), bg = 'azure', fg = 'blue', width = 11, height = 2, font = ('Arial', 8))

    button_1_1.grid(row = 1, column = 1, padx = 2, pady = 2)

    button_1_2 = tk.Button(tab3, text = 'Chuẩn hóa\n Lời giải', command = (lambda : vbf.chuanhoa_loigiai(word)), bg = 'palegoldenrod', fg = 'blue', width = 11, height = 2, font = ('Arial', 8))

    button_1_2.grid(row = 1, column = 2, padx = 2, pady = 2)

    button_1_3 = tk.Button(tab3, text = 'Chuẩn ĐA\nGạch chân, đỏ', command = (lambda : vbf.Gach_chan_red_dap_an(word)), bg = 'azure', fg = 'crimson', width = 11, height = 2, font = ('Arial', 8))

    button_1_3.grid(row = 1, column = 3, padx = 2, pady = 2)

    button_2_0 = tk.Button(tab3, text = 'STT Câu, Bài\n Ví dụ,Question', command = (lambda : vbf.STT_cau_new(root, word)), bg = 'chartreuse', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_2_0.grid(row = 2, column = 0, padx = 2, pady = 2)

    button_2_1 = tk.Button(tab3, text = 'Đánh STT\n Đề 2025', command = (lambda : vbf.STT_2025_new(word)), bg = 'chartreuse', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_2_1.grid(row = 2, column = 1, padx = 2, pady = 2)

    button_2_2 = tk.Button(tab3, text = 'Đánh STT\n Tùy chọn', bg = 'chartreuse', command = (lambda : vbf.STT_tuy_chon_khac(root, word)), fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_2_2.grid(row = 2, column = 2, padx = 2, pady = 2)

    button_2_3 = tk.Button(tab3, text = 'Chuyển đổi\n STT', command = (lambda : vbf.chuyen_doi_STT(root, word)), bg = 'chartreuse', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_2_3.grid(row = 2, column = 3, padx = 2, pady = 2)

    button_3_0 = tk.Button(tab3, text = 'Lệch Math', command = (lambda : vbf.lech_pos_pic_math(word)), bg = 'yellowgreen', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_3_0.grid(row = 3, column = 0, padx = 2, pady = 2)

    button_3_1 = tk.Button(tab3, text = 'Picture Inline\nCanh giữa', command = (lambda : vbf.pic_inline_center_tool(root, word)), bg = 'yellowgreen', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_3_1.grid(row = 3, column = 1, padx = 2, pady = 2)

    button_3_2 = tk.Button(tab3, text = 'Fix Equation\nSelect', command = (lambda : vbf.fix_equation_select(word)), bg = 'green', fg = 'white', width = 11, height = 2, font = ('Arial', 8))

    button_3_2.grid(row = 3, column = 2, padx = 2, pady = 2)

    button_3_3 = tk.Button(tab3, text = 'Fix Bảng\n số liệu', command = (lambda : vbf.gui_autoFit_table(root, word)), bg = 'darkseagreen', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_3_3.grid(row = 3, column = 3, padx = 2, pady = 2)

    button_4_0 = tk.Button(tab3, text = 'convert Table\n to text', command = (lambda : vbf.Table_to_text_tool(root, word)), bg = 'azure', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_4_0.grid(row = 4, column = 0, padx = 2, pady = 2)

    button_4_1 = tk.Button(tab3, text = 'Câu Auto\n to text', command = (lambda : vbf.Convert_Auto_To_Text(word)), bg = 'azure', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_4_1.grid(row = 4, column = 1, padx = 2, pady = 2)

    button_4_2 = tk.Button(tab3, text = 'Xử lí \nfile gốc', command = (lambda : chd.Fix_du_lieu_xuong_dong(root, word)), bg = 'yellow', fg = 'red', width = 11, height = 2, font = ('Arial', 8))

    button_4_2.grid(row = 4, column = 2, padx = 2, pady = 2)

    button_4_3 = tk.Button(tab3, text = 'Chuẩn hóa\nđáp số', command = (lambda : vbf.chuanhoa_dap_so_py32(word)), fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_4_3.grid(row = 4, column = 3, padx = 2, pady = 2)

    button_5_0 = tk.Button(tab3, text = 'Màu [ ] ( )\nChú thích', command = (lambda : vbf.dam_mau_chu_thich(word)), bg = 'whitesmoke', fg = 'green', width = 11, height = 2, font = ('Arial', 8))

    button_5_0.grid(row = 5, column = 0, padx = 2, pady = 2)

    button_5_1 = tk.Button(tab3, text = 'xóa [ ] ( )\nChú thích', command = (lambda : vbf.xoa_chu_thich_all(word)), bg = 'lightpink', fg = 'green', width = 11, height = 2, font = ('Arial', 8))

    button_5_1.grid(row = 5, column = 1, padx = 2, pady = 2)

    button_5_2 = tk.Button(tab3, text = 'Bỏ shading\ncứng đầu', command = (lambda : vbf.bo_mau_tung_kitu(word)), bg = 'silver', fg = 'green', width = 11, height = 2, font = ('Arial', 8))

    button_5_2.grid(row = 5, column = 2, padx = 2, pady = 2)

    button_5_3 = tk.Button(tab3, text = 'Bỏ U.d.line\n H.light, Shading', command = (lambda : vbf.bo_Unline_Hightlight_Red(word)), bg = 'chocolate', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_5_3.grid(row = 5, column = 3, padx = 2, pady = 2)

    button_6_0 = tk.Button(tab3, text = 'Xóa Lời giải\nCâu, ví dụ', command = (lambda : vbf.xoa_loi_giai_tool(root, word)), bg = 'plum', fg = 'blue', width = 11, height = 2, font = ('Arial', 8))

    button_6_0.grid(row = 6, column = 0, padx = 2, pady = 2)

    button_6_1 = tk.Button(tab3, text = 'xóa header', command = (lambda : vbf.xoa_headers_footers(word)), bg = 'lightpink', fg = 'mediumvioletred', width = 11, height = 2, font = ('Arial', 8))

    button_6_1.grid(row = 6, column = 1, padx = 2, pady = 2)

    button_6_2 = tk.Button(tab3, text = 'xóa tác giả', command = (lambda : vbf.xoa_tac_gia(word)), bg = 'lightpink', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_6_2.grid(row = 6, column = 2, padx = 2, pady = 2)

    button_7_3 = tk.Button(tab3, text = 'bỏ gạch chân\n ở tab', command = (lambda : vbf.thay_the_replace_no_format(word, '(^9)', '\\1')), bg = 'orange', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_7_3.grid(row = 6, column = 3, padx = 2, pady = 2)

    button_7_0 = tk.Button(tab3, text = 'xóa \ndòng trắng', command = (lambda : vbf.xoa_dong_trang(word)), bg = 'mistyrose', fg = 'mediumvioletred', width = 11, height = 2, font = ('Arial', 8))

    button_7_0.grid(row = 7, column = 0, padx = 2, pady = 2)

    button_7_1 = tk.Button(tab3, text = 'xóa more...', command = (lambda : vbf.xoa_more_tool(root, word)), bg = 'mistyrose', fg = 'mediumvioletred', width = 11, height = 2, font = ('Arial', 8))

    button_7_1.grid(row = 7, column = 1, padx = 2, pady = 2)

    button_9_2 = tk.Button(tab3, text = 'Times New\n Roman-12', command = (lambda : vbf.font_time_12(word)), bg = 'khaki', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_9_2.grid(row = 7, column = 2, padx = 2, pady = 2)

    button_9_3 = tk.Button(tab3, text = 'Save As\nNew name', command = (lambda : vbf.save_as_to_name(root, word)), bg = 'khaki', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_9_3.grid(row = 7, column = 3, padx = 2, pady = 2)

    button_6_3 = tk.Button(tab3, text = 'xóa dòng \n key nhập', command = (lambda : vbf.xoa_key_nhap(root, word)), bg = 'plum', fg = 'mediumvioletred', width = 11, height = 2, font = ('Arial', 8))

    button_6_3.grid(row = 8, column = 0, padx = 2, pady = 2)

    button_STT_xdong = tk.Button(tab3, text = 'Canh lề\nCanh tab', command = (lambda : chd.canh_le_tab_tool(root, word)), bg = 'whitesmoke', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_STT_xdong.grid(row = 8, column = 1, padx = 2, pady = 2)

    button_8_1 = tk.Button(tab3, text = 'Dòng chấm', command = (lambda : vbf.dong_cham_tool(root, word)), bg = 'lightgreen', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_8_1.grid(row = 8, column = 2, padx = 2, pady = 2)

    button_8_3 = tk.Button(tab3, text = 'Color', command = (lambda : vbf.color_tool(root, word)), bg = 'mediumpurple', fg = 'white', width = 11, height = 2, font = ('Arial', 8))

    button_8_3.grid(row = 8, column = 3, padx = 2, pady = 2)

    button_9_1 = tk.Button(tab3, text = 'Chuẩn hóa\n đề A4', command = (lambda : chd.Chuanhoade(word)), bg = 'yellowgreen', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_9_1.grid(row = 9, column = 0, padx = 2, pady = 2)

    button_7_2 = tk.Button(tab3, text = 'Chuẩn hóa\n đề 2 cột', command = (lambda : chd.Chuanhoade_2cot(word)), bg = 'yellowgreen', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_7_2.grid(row = 9, column = 1, padx = 2, pady = 2)

    button_8_2 = tk.Button(tab3, text = 'Chuẩn hóa\nNhanh A4', command = (lambda : chd.chuan_hoa_nhanh_A4_by_docx(word)), bg = 'green', fg = 'white', width = 11, height = 2, font = ('Arial', 8))

    button_8_2.grid(row = 9, column = 2, padx = 2, pady = 2)

    button_9_0 = tk.Button(tab3, text = 'Chuẩn hóa\nNhanh 2C', command = (lambda : chd.chuanhoa_bo_DA_docx(doc, messages)), bg = 'green', fg = 'white', width = 11, height = 2, font = ('Arial', 8))

    button_9_0.grid(row = 9, column = 3, padx = 2, pady = 2)

    button_4_1 = tk.Button(tab4, text = 'Xử lí SGK', command = (lambda : vbf.xu_li_SGK(root, word)), bg = 'aquamarine', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_4_1.grid(row = 0, column = 0, padx = 2, pady = (10, 2))

    button_4_3 = tk.Button(tab4, text = 'Xóa chữ viết\nDrawing tools', command = (lambda : vbf.delete_ink_drawings(word)), bg = 'mistyrose', fg = 'mediumvioletred', width = 11, height = 2, font = ('Arial', 8))

    button_4_3.grid(row = 0, column = 1, padx = 2, pady = (10, 2))

    button_5_0 = tk.Button(tab4, text = 'Đếm số câu', command = (lambda : vbf.dem_so_cau(word)), bg = 'aliceblue', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_5_0.grid(row = 0, column = 2, padx = 2, pady = (10, 2))

    button_5_1 = tk.Button(tab4, text = 'Lỗi shift enter', command = (lambda : vbf.xoa_shift_enter(word)), fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_5_1.grid(row = 0, column = 3, padx = 2, pady = (10, 2))

    button_5_2 = tk.Button(tab4, text = 'fix Đáp Án', command = (lambda : vbf.gui_fix_Dap_an(root, word)), bg = 'darkseagreen', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_5_2.grid(row = 1, column = 0, padx = 2, pady = 2)

    button_table_DS = tk.Button(tab4, text = 'Convert bảng\nĐúng/Sai', command = (lambda : vbf.gui_table_dung_sai(root, word)), fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_table_DS.grid(row = 1, column = 1, padx = 2, pady = 2)

    button_8_2 = tk.Button(tab4, text = 'Chuẩn hóa\nGhi chú phải', command = (lambda : chd.chuan_hoa_note(word)), bg = 'whitesmoke', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_8_2.grid(row = 1, column = 2, padx = 2, pady = 2)

    button_4_0 = tk.Button(tab4, text = 'Mathtype to\nLatex', command = (lambda : vbf.gui_mathtype2Latex(root, word)), bg = 'yellow', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_4_0.grid(row = 2, column = 0, padx = 2, pady = 2)

    button_5_1 = tk.Button(tab4, text = 'Latex to\nMathtype', command = (lambda : vbf.gui_latex2Mathtype(root, word)), bg = 'yellowgreen', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_5_1.grid(row = 2, column = 1, padx = 2, pady = 2)

    button_outline = tk.Button(tab4, text = 'Bỏ OutlineLevel\n và Headings', command = (lambda : vbf.gui_remove_collapsible(root, word)), bg = 'cornflowerblue', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_outline.grid(row = 2, column = 2, padx = 2, pady = 2)

    button_5_1 = tk.Button(tab4, text = 'Chem Draw\nlệch dòng', command = (lambda : vbf.lowered_chemdraw_in_word(word)), bg = '#33FF55', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_5_1.grid(row = 2, column = 3, padx = 2, pady = 2)

    button_5_3 = tk.Button(tab4, text = 'Replace \nnhiều cụm từ', command = (lambda : vbf.gui_replace(root, word)), bg = 'darksalmon', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_5_3.grid(row = 3, column = 0, padx = 2, pady = 2)

    button_save_pdf = tk.Button(tab4, text = 'Print pdf', command = (lambda : pdf.print_microsoft_pdf(word)), bg = 'moccasin', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_save_pdf.grid(row = 3, column = 1, padx = 2, pady = 2)

    rename_button = tk.Button(tab4, text = 'Đổi tên file', command = (lambda : vbrn.gui_rename(root)), bg = 'cornsilk', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    rename_button.grid(row = 3, column = 2, padx = 2, pady = 2)

    button_word_pdf = tk.Button(tab4, text = 'pdf to word', command = (lambda : w2pdf.gui_convert2pdf_olmocr(root)), bg = '#FF3300', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_word_pdf.grid(row = 3, column = 3, padx = 2, pady = 2)

    Label_nhanh = tk.Label(tab4, text = '-----CÔNG CỤ XỬ LÍ NHANH, CẦN ĐÓNG FILE-----', foreground = 'black', background = 'lightpink')

    Label_nhanh.grid(row = 4, column = 0, columnspan = 4, padx = 5, pady = (15, 5))

    button_9_0 = tk.Button(tab4, text = 'Tách đề -CH\nCực nhanh', command = (lambda : docxtool.Tach_de_chuan_hoa(root)), bg = 'yellowgreen', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_9_0.grid(row = 5, column = 0, padx = 2, pady = 2)

    button_9_1 = tk.Button(tab4, text = 'Tách file\n nhiều phần', command = (lambda : docxtool.Tach_file_large(root)), bg = 'olive', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_9_1.grid(row = 5, column = 1, padx = 2, pady = 2)

    button_9_2 = tk.Button(tab4, text = 'Tìm kiếm \ncâu  hỏi', command = (lambda : sprtool.Tim_kiem_cau_hoi(root)), bg = 'khaki', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_9_2.grid(row = 5, column = 2, padx = 2, pady = 2)

    button_9_3 = tk.Button(tab4, text = 'Thay thế\nnhiều file', command = (lambda : sprtool.Thay_the_spr(root)), bg = 'darkorchid', fg = 'white', width = 11, height = 2, font = ('Arial', 8))

    button_9_3.grid(row = 5, column = 3, padx = 2, pady = 2)

    button_eq2lt = tk.Button(tab4, text = 'Equation\n to latex', command = (lambda : OMML_latex.gui_convert_OMML_tolatex(root)), bg = 'yellow', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_eq2lt.grid(row = 6, column = 0, padx = 2, pady = 2)

    button_lt2eq = tk.Button(tab4, text = 'Latex to\nEquation', command = (lambda : latex_OMML.gui_convert_latex_to_OMML(root)), bg = 'darkseagreen', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_lt2eq.grid(row = 6, column = 1, padx = 2, pady = 2)

    button_5_1 = tk.Button(tab4, text = '-------', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_5_1.grid(row = 6, column = 2, padx = 2, pady = 2)

    button_5_1 = tk.Button(tab4, text = '-------', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_5_1.grid(row = 6, column = 3, padx = 2, pady = 2)

    Label_nhanh = tk.Label(tab4, text = '-----CÔNG CỤ TẠO FILE HTML-----', foreground = 'black', background = 'lightpink')

    Label_nhanh.grid(row = 7, column = 0, columnspan = 4, padx = 5, pady = (15, 5))

    button_9_0 = tk.Button(tab4, text = 'Trắc nghiệm\nTrình chiếu', command = (lambda : tn_offline.gui_on(root, word)), bg = '#FF9900', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_9_0.grid(row = 8, column = 0, padx = 2, pady = 2)

    button_9_1 = tk.Button(tab4, text = 'Trắc nghiệm\nOnline', command = (lambda : tn_online.gui_on(root, word)), bg = '#FF99CC', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_9_1.grid(row = 8, column = 1, padx = 2, pady = 2)

    button_9_2 = tk.Button(tab4, text = 'Vòng quay', command = (lambda : quay.gui_vongquay(root)), bg = '#33FF66', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_9_2.grid(row = 8, column = 2, padx = 2, pady = 2)

    button_9_3 = tk.Button(tab4, text = 'Bảng điểm', command = (lambda : bangdiem.gui_bangdiem(root)), bg = '#FFFF00', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_9_3.grid(row = 8, column = 3, padx = 2, pady = 2)

    button_1_0 = tk.Button(tab7, text = 'Xuống dòng \n phương án', command = (lambda : chd.xuongdong_phuongan_Tool(word)), bg = 'green', fg = 'white', width = 11, height = 2, font = ('Arial', 8))

    button_1_0.grid(row = 1, column = 0, padx = 2, pady = (10, 2))

    button_1_1 = tk.Button(tab7, text = 'word to PP All\n (Equation)', command = (lambda : wtp.w2p_all(word)), bg = 'palegoldenrod', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_1_1.grid(row = 1, column = 1, padx = 2, pady = (10, 2))

    button_3_1 = tk.Button(tab7, text = 'Chọn size Font\nAll slide', command = wtp.Fix_size_Font_All_silde, bg = 'indian red', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_3_1.grid(row = 1, column = 2, padx = 2, pady = (10, 2))

    button_Font_all = tk.Button(tab7, text = 'Chọn Font\nAll slide', command = wtp.Fix_name_Font_All_silde, bg = 'indian red', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_Font_all.grid(row = 1, column = 3, padx = 2, pady = (10, 2))

    button_2_0 = tk.Button(tab7, text = 'word to PP\n (Select 1)', command = (lambda : wtp.w2p_select(word)), bg = 'olive', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_2_0.grid(row = 2, column = 0, padx = 2, pady = 2)

    button_format_all = tk.Button(tab7, text = 'Định dạng\nslide select', command = wtp.Fix_AllShape_in_silde, bg = 'lightgreen', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_format_all.grid(row = 2, column = 1, padx = 2, pady = 2)

    button_3_0 = tk.Button(tab7, text = 'Chọn size Font\nslide select ', command = wtp.Fix_size_Font_one_silde, bg = 'coral', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_3_0.grid(row = 2, column = 2, padx = 2, pady = 2)

    button_3_2 = tk.Button(tab7, text = 'Chọn Font\nslide select', command = wtp.Fix_name_Font_one_silde, bg = 'coral', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_3_2.grid(row = 2, column = 3, padx = 2, pady = 2)

    button_4_dong = tk.Button(tab7, text = 'Đáp án 4 dòng\nSlide select', command = wtp.Canh_dap_an_one_slide_4_dong, bg = 'turquoise', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_4_dong.grid(row = 4, column = 0, padx = 2, pady = 2)

    button_2_dong = tk.Button(tab7, text = 'Đáp án 2 dòng\nSlide select', command = wtp.Canh_dap_an_one_slide_2_dong, bg = 'turquoise', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_2_dong.grid(row = 4, column = 1, padx = 2, pady = 2)

    button_1_dong = tk.Button(tab7, text = 'Đáp án 1 dòng\nSlide select', command = wtp.Canh_dap_an_one_slide_1_dong, bg = 'turquoise', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_1_dong.grid(row = 4, column = 2, padx = 2, pady = 2)

    button_fix_kq = tk.Button(tab7, text = 'Fix kết quả\nslide Select', command = wtp.Fix_result_at, bg = 'orange red', fg = 'black', width = 11, height = 2, font = ('Arial', 8))

    button_fix_kq.grid(row = 4, column = 3, padx = 2, pady = 2)

    button_del_kq = tk.Button(tab7, text = 'Xóa kết quả', command = wtp.delete_result_box, bg = 'mistyrose', fg = 'mediumvioletred', width = 11, height = 1, font = ('Arial', 8))

    button_del_kq.grid(row = 5, column = 0, padx = 2, pady = (10, 2))

    button_show = tk.Button(tab7, text = 'Show start 1', command = wtp.slide_show_start_1, bg = 'lightgreen', fg = 'black', width = 11, height = 1, font = ('Arial', 8))

    button_show.grid(row = 5, column = 1, padx = 2, pady = (10, 2))

    button_at = tk.Button(tab7, text = 'Show At slide', command = wtp.slide_show_at_now, bg = 'lightgreen', fg = 'black', width = 11, height = 1, font = ('Arial', 8))

    button_at.grid(row = 5, column = 2, padx = 2, pady = (10, 2))

    button_next = tk.Button(tab7, text = 'Next', command = wtp.go_next_slide, bg = 'lightgreen', fg = 'black', width = 11, height = 1, font = ('Arial', 8))

    button_next.grid(row = 6, column = 0, padx = 2, pady = 2)

    button_previous = tk.Button(tab7, text = 'Previous', command = wtp.go_Previous_slide, bg = 'lightgreen', fg = 'black', width = 11, height = 1, font = ('Arial', 8))

    button_previous.grid(row = 6, column = 1, padx = 2, pady = 2)

    button_Exit = tk.Button(tab7, text = 'Exit Show', command = wtp.exit_slide_show, bg = 'lightgreen', fg = 'black', width = 11, height = 1, font = ('Arial', 8))

    button_Exit.grid(row = 6, column = 2, padx = 2, pady = 2)

    button_pen = tk.Button(tab7, text = 'PEN', command = wtp.open_pen, bg = 'lightgreen', fg = 'black', width = 11, height = 1, font = ('Arial', 8))

    button_pen.grid(row = 7, column = 0, padx = 2, pady = 2)

    button_Highlighter = tk.Button(tab7, text = 'Eraser', command = wtp.open_Eraser, bg = 'lightgreen', fg = 'black', width = 11, height = 1, font = ('Arial', 8))

    button_Highlighter.grid(row = 7, column = 1, padx = 2, pady = 2)

    button_mouse = tk.Button(tab7, text = 'Mouse', command = wtp.return_mouse, bg = 'lightgreen', fg = 'black', width = 11, height = 1, font = ('Arial', 8))

    button_mouse.grid(row = 7, column = 2, padx = 2, pady = 2)

    button_mouse = tk.Button(tab7, text = 'Đồng hồ: count down, stop watch', command = clockcd.clockmain, bg = '#CCFF00', fg = 'black', width = 11, height = 1, font = ('Arial', 8))

    button_mouse.grid(row = 8, column = 0, columnspan = 4, padx = 2, pady = 2, sticky = 'ew')

    frame_t3b = tk.Frame(tab5)

    frame_t3b.pack(padx = 10, pady = 10)

    listbox_pdf = tk.Listbox(frame_t3b, selectmode = tk.MULTIPLE, width = 45, height = 15)

    listbox_pdf.pack(side = tk.LEFT, padx = 5)

    scrollbar = tk.Scrollbar(frame_t3b, orient = tk.VERTICAL)

    scrollbar.config(command = listbox_pdf.yview)

    scrollbar.pack(side = tk.RIGHT, fill = tk.Y)

    listbox_pdf.config(yscrollcommand = scrollbar.set)

    frame_t32 = tk.Frame(tab5)

    frame_t32.pack(padx = 0, pady = 10)

    select_button = tk.Button(frame_t32, text = 'Select Files', command = (lambda : pdf.select_files(listbox_pdf)), bg = 'cyan', fg = 'black', width = 20)

    select_button.pack(side = tk.LEFT, padx = (0, 5), pady = (5, 10))

    clear_button = tk.Button(frame_t32, text = 'Clear All files', command = (lambda : pdf.clear_files(listbox_pdf)), bg = 'white', fg = 'blue', width = 20)

    clear_button.pack(side = tk.LEFT, padx = (0, 5), pady = (5, 10))

    frame_t33 = tk.Frame(tab5)

    frame_t33.pack(padx = 0, pady = 10)

    fix_button = tk.Button(frame_t33, text = 'Convert to PDF\n(Xã hội)', command = (lambda : pdf.convert_to_pdf_batch(word, listbox_pdf)), bg = 'lightgreen', fg = 'black', width = 20)

    fix_button.pack(side = tk.LEFT, padx = (0, 2), pady = (5, 10))

    fix_button = tk.Button(frame_t33, text = 'Convert to PDF\n(Chuẩn cho mathtype)', command = (lambda : pdf.run_print_pdf(word, listbox_pdf)), bg = 'lightpink', fg = 'black', width = 20)

    fix_button.pack(side = tk.LEFT, padx = (0, 2), pady = (5, 10))

    Thong_tin = tk.Label(tabA, text = '-----Dành cho tất cả các môn khác ngoại ngữ-----', foreground = 'blue', background = 'yellowgreen')

    Thong_tin.grid(row = 0, column = 0, columnspan = 3, padx = 2, pady = 5)

    info_frame = ttk.LabelFrame(tabA, text = 'I. Xử lí dữ liệu: Chỉ nên mở duy nhất 1 file word cần làm', padding = '2')

    info_frame.grid(row = 1, column = 0, columnspan = 2, sticky = 'ew', padx = (5, 2), pady = 2)

    info_frame.columnconfigure(2, weight = 1)

    button_xdong = tk.Button(info_frame, text = '(1) Xuống dòng', command = (lambda : chd.xuong_dong_nhanh_combine_docx(word)), height = 2, bg = 'palegoldenrod', fg = 'blue')

    button_xdong.grid(row = 0, column = 0, padx = 2, pady = 2)

    button_checkDATA = tk.Button(info_frame, text = '(2) Kiểm tra dữ liệu', command = (lambda : vme.check_data_nhanh_combine_docx(word)), height = 2, bg = 'mistyrose', fg = 'mediumvioletred')

    button_checkDATA.grid(row = 0, column = 1, padx = 2, pady = 2)

    button_checkDATA = tk.Button(info_frame, text = '(3) save và tắt file', command = (lambda : vbf.save_convert_sendpath_docx(word, entry_input)), height = 2, bg = '#C9E4D6', fg = 'black')

    button_checkDATA.grid(row = 0, column = 2, padx = 2, pady = 2)

    tronde_frame = ttk.LabelFrame(tabA, text = 'II. Trộn đề: Bắt buộc phải tắt file gốc', padding = '2')

    tronde_frame.grid(row = 2, column = 0, columnspan = 3, sticky = 'ew', padx = (5, 2), pady = 2)

    tronde_frame.columnconfigure(1, weight = 1)

    button_browse = tk.Button(tronde_frame, text = 'Chọn file gốc   ', command = browse_file_Me, width = 13, bg = 'cyan', fg = 'black')

    button_browse.grid(row = 0, column = 0, padx = 2, pady = 2)

    entry_input = tk.Entry(tronde_frame, width = 30)

    entry_input.grid(row = 0, column = 1, columnspan = 2, sticky = 'w', padx = 2, pady = 2)

    update_time = tk.Button(tronde_frame, text = 'Cập nhật time', command = cap_nhat_time, width = 13, bg = 'yellowgreen', fg = 'black')

    update_time.grid(row = 1, column = 0, padx = 2, pady = 2)

    entry_lantron = tk.Entry(tronde_frame, width = 7)

    entry_lantron.grid(row = 1, column = 1, padx = 2, sticky = 'w', pady = 2)

    entry_lantron.insert(0, datetime.datetime.now().strftime('%Hh%M'))

    label_sode = tk.Label(tronde_frame, text = 'Số đề:')

    label_sode.grid(row = 2, column = 0, padx = 2, pady = 2)

    entry_num = tk.Entry(tronde_frame, width = 10)

    entry_num.grid(row = 3, column = 0, padx = 2, pady = 2)

    entry_num.insert(0, '4')

    label_made = tk.Label(tronde_frame, text = 'Mã đề bắt đầu')

    label_made.grid(row = 2, column = 1, padx = 5, pady = 5)

    entry_made_start = tk.Entry(tronde_frame, width = 10)

    entry_made_start.grid(row = 3, column = 1, padx = 5, pady = 5)

    entry_made_start.insert(0, '1201')

    label_made = tk.Label(tronde_frame, text = 'Khoảng cách')

    label_made.grid(row = 2, column = 2, padx = 5, pady = 5)

    entry_khoang_cach_made = tk.Entry(tronde_frame, width = 10)

    entry_khoang_cach_made.grid(row = 3, column = 2, padx = 5, pady = 5)

    entry_khoang_cach_made.insert(0, '1')

    btn_tao_ma_de = tk.Button(tronde_frame, text = '(4) Tạo mã đề', command = tao_made_tudong, width = 13, bg = 'green', fg = 'white')

    btn_tao_ma_de.grid(row = 6, column = 0, padx = 15, pady = 5)

    entry_newname = tk.Entry(tronde_frame, width = 30)

    entry_newname.grid(row = 6, column = 1, columnspan = 2, sticky = 'w', padx = 5, pady = 5)

    entry_newname.insert(0, '1201,1202,1203,1204')

    label_phan1 = tk.Label(tronde_frame, text = 'Phần ABCD:')

    label_phan1.grid(row = 7, column = 0, padx = 2, pady = 2)

    check1_2_var = tk.BooleanVar(value = True)

    check1_2 = tk.Checkbutton(tronde_frame, text = 'Trộn câu       ', variable = check1_2_var)

    check1_2.grid(row = 7, column = 1, padx = 2, pady = 2)

    check1_1_var = tk.BooleanVar(value = True)

    check1_1 = tk.Checkbutton(tronde_frame, text = 'Trộn đáp án', variable = check1_1_var)

    check1_1.grid(row = 7, column = 2, padx = 2, pady = 2)

    label_phan2 = tk.Label(tronde_frame, text = 'Phần abcd:')

    label_phan2.grid(row = 8, column = 0, padx = 2, pady = 2)

    check2_2_var = tk.BooleanVar(value = True)

    check2_2 = tk.Checkbutton(tronde_frame, text = 'Trộn câu       ', variable = check2_2_var)

    check2_2.grid(row = 8, column = 1, padx = 2, pady = 2)

    check2_1_var = tk.BooleanVar()

    check2_1 = tk.Checkbutton(tronde_frame, text = 'Trộn đáp án', bg = 'yellow', variable = check2_1_var)

    check2_1.grid(row = 8, column = 2, padx = 2, pady = 2)

    label_phan3 = tk.Label(tronde_frame, text = 'TLN và TL:')

    label_phan3.grid(row = 9, column = 0, padx = 2, pady = 2)

    check3_1_var = tk.BooleanVar(value = True)

    check3_1 = tk.Checkbutton(tronde_frame, text = 'Trộn câu TLN', variable = check3_1_var)

    check3_1.grid(row = 9, column = 1, padx = 2, pady = 2)

    check4_1_var = tk.BooleanVar()

    check4_1 = tk.Checkbutton(tronde_frame, text = 'Trộn câu TL', variable = check4_1_var)

    check4_1.grid(row = 9, column = 2, padx = 2, pady = 2)

    check5_1_var = tk.BooleanVar(value = False)

    check5_1 = tk.Checkbutton(tronde_frame, text = 'Đáp án TNMaker', variable = check5_1_var)

    check5_1.grid(row = 10, column = 0, columnspan = 3, sticky = 'w', padx = 5, pady = 5)

    check6_1_var = tk.BooleanVar()

    check6_1 = tk.Checkbutton(tronde_frame, text = 'Đề 2 cột', variable = check6_1_var)

    check6_1.grid(row = 10, column = 1, padx = 2, pady = 5, sticky = 'w')

    check7_1_var = tk.BooleanVar()

    check7_1 = tk.Checkbutton(tronde_frame, text = 'Kèm lời giải', variable = check7_1_var)

    check7_1.grid(row = 10, column = 2, padx = 2, pady = 5)

    button_process = tk.Button(tronde_frame, text = '(5) Xuất đề', command = tron_and_xuat_de_Me, bg = 'lightpink', fg = 'black', width = 10)

    button_process.grid(row = 11, column = 0, padx = 5, pady = 5, sticky = 'ew')

    button_open = tk.Button(tronde_frame, text = 'Mở thư mục chứa đề', command = (lambda : vbf.open_folder(entry_input)), bg = '#99FF66', fg = 'black', width = 10)

    button_open.grid(row = 11, column = 1, columnspan = 2, padx = 5, pady = 5, sticky = 'ew')

    chucnangkhac_frame = ttk.LabelFrame(tabA, text = 'III. Chức năng khác: Tạo file mẫu và ...', padding = '2')

    chucnangkhac_frame.grid(row = 3, column = 0, columnspan = 2, sticky = 'ew', padx = (5, 2), pady = 2)

    chucnangkhac_frame.columnconfigure(1, weight = 1)

    button_khac = tk.Button(chucnangkhac_frame, text = 'Chức năng khác', command = (lambda : vme.chuc_nang_khac_mix(root, word, addin_name)), width = 20, bg = 'moccasin', fg = 'black')

    button_khac.grid(row = 1, column = 0, padx = 2, pady = 2)

    button_youtube_tron = tk.Button(chucnangkhac_frame, text = 'Xem youtube ', command = (lambda : open_link_url('https://youtu.be/NqJZIE855qE')), width = 20, bg = 'chocolate', fg = 'white')

    button_youtube_tron.grid(row = 1, column = 1, padx = 2, pady = 2)

    Thong_tin = tk.Label(tab_EN, text = '-----Dành cho môn ngoại ngữ-----', foreground = 'blue', background = 'hotpink')

    Thong_tin.grid(row = 0, column = 0, columnspan = 3, padx = 2, pady = 2)

    info_frame_EN = ttk.LabelFrame(tab_EN, text = '1. Xử lí dữ liệu: Chỉ nên mở duy nhất 1 file word cần làm', padding = '2')

    info_frame_EN.grid(row = 1, column = 0, columnspan = 3, sticky = 'ew', padx = (5, 2), pady = 2)

    info_frame_EN.columnconfigure(1, weight = 1)

    button_xdong = tk.Button(info_frame_EN, text = '(1) Xuống dòng phương án', command = (lambda : chd.xuong_dong_nhanh_combine_docx(word)), height = 1, bg = 'palegoldenrod', fg = 'blue')

    button_xdong.grid(row = 1, column = 0, padx = 3, pady = 5)

    button_duc_lo = tk.Button(info_frame_EN, text = 'Xử lí phần đục lỗ', command = (lambda : vme.thay_so_trong_phan_duc_lo(word)), height = 1, width = 20, bg = 'cyan', fg = 'black')

    button_duc_lo.grid(row = 1, column = 1, padx = 3, pady = 5)

    button_table = tk.Button(info_frame_EN, text = '(2) Kiểm tra dữ liệu', command = (lambda : vme.check_data_nhanh_combine_docx_EN(word)), width = 20, height = 1, bg = 'mistyrose', fg = 'mediumvioletred')

    button_table.grid(row = 2, column = 0, padx = 5, pady = 5)

    button_table = tk.Button(info_frame_EN, text = '(3) Save và tắt file', command = (lambda : vbf.save_convert_sendpath_docx(word, entry_input_EN)), width = 20, height = 1, bg = '#C9E4D6', fg = 'black')

    button_table.grid(row = 2, column = 1, padx = 5, pady = 5)

    tronde_frame_EN = ttk.LabelFrame(tab_EN, text = '2. Trộn đề: Nên save và tắt hết word trước khi làm', padding = '2')

    tronde_frame_EN.grid(row = 2, column = 0, columnspan = 3, sticky = 'ew', padx = (5, 2), pady = 2)

    tronde_frame_EN.columnconfigure(1, weight = 1)

    button_browse_EN = tk.Button(tronde_frame_EN, text = 'Chọn file gốc', command = browse_file_Me_EN, bg = 'cyan', fg = 'black')

    button_browse_EN.grid(row = 0, column = 0, padx = 5, pady = 5)

    entry_input_EN = tk.Entry(tronde_frame_EN, width = 30)

    entry_input_EN.grid(row = 0, column = 1, columnspan = 3, padx = 5, pady = 5)

    update_time_EN = tk.Button(tronde_frame_EN, text = 'Cập nhật time', command = cap_nhat_time_EN, bg = 'yellowgreen', fg = 'black')

    update_time_EN.grid(row = 1, column = 0, padx = 5, pady = 10)

    entry_lantron_EN = tk.Entry(tronde_frame_EN, width = 10)

    entry_lantron_EN.grid(row = 1, column = 1, padx = 5, pady = 5)

    entry_lantron_EN.insert(0, datetime.datetime.now().strftime('%Hh%M'))

    label_sode_EN = tk.Label(tronde_frame_EN, text = 'Số đề:')

    label_sode_EN.grid(row = 2, column = 0, padx = 5, pady = 5)

    entry_num_EN = tk.Entry(tronde_frame_EN, width = 10)

    entry_num_EN.grid(row = 3, column = 0, padx = 5, pady = 5)

    entry_num_EN.insert(0, '4')

    label_made_EN = tk.Label(tronde_frame_EN, text = 'Mã đề bắt đầu')

    label_made_EN.grid(row = 2, column = 1, padx = 5, pady = 5)

    entry_made_start_EN = tk.Entry(tronde_frame_EN, width = 10)

    entry_made_start_EN.grid(row = 3, column = 1, padx = 5, pady = 5)

    entry_made_start_EN.insert(0, '1201')

    label_made_EN = tk.Label(tronde_frame_EN, text = 'Khoảng cách')

    label_made_EN.grid(row = 2, column = 2, padx = 5, pady = 5)

    entry_khoang_cach_made_EN = tk.Entry(tronde_frame_EN, width = 10)

    entry_khoang_cach_made_EN.grid(row = 3, column = 2, padx = 5, pady = 5)

    entry_khoang_cach_made_EN.insert(0, '1')

    btn_tao_ma_de_EN = tk.Button(tronde_frame_EN, text = '(4) Tạo các mã đề', command = tao_made_tudong_EN, bg = 'green', fg = 'white')

    btn_tao_ma_de_EN.grid(row = 5, column = 0, padx = 15, pady = 5)

    entry_newname_EN = tk.Entry(tronde_frame_EN, width = 30)

    entry_newname_EN.grid(row = 5, column = 1, columnspan = 3, padx = 5, pady = 5)

    entry_newname_EN.insert(0, '1201,1202,1203,1204')

    check6_1_EN_var = tk.BooleanVar()

    check6_1_EN = tk.Checkbutton(tronde_frame_EN, text = 'Đề 2 cột', variable = check6_1_EN_var)

    check6_1_EN.grid(row = 6, column = 0, padx = 5, pady = 5)

    check7_1_EN_var = tk.BooleanVar()

    check7_1_EN = tk.Checkbutton(tronde_frame_EN, text = 'Kèm lời giải', variable = check7_1_EN_var)

    check7_1_EN.grid(row = 6, column = 1, padx = 5, pady = 5)

    check8_1_EN_var = tk.BooleanVar()

    check8_1_EN = tk.Checkbutton(tronde_frame_EN, text = 'Tiếng Nhật', variable = check8_1_EN_var)

    check8_1_EN.grid(row = 6, column = 2, padx = 5, pady = 5)

    button_process_EN = tk.Button(tronde_frame_EN, text = '(5) Xuất đề', command = tron_and_xuat_de_Me_NN, bg = 'lightpink', fg = 'black', height = 2)

    button_process_EN.grid(row = 7, column = 0, padx = 5, pady = 5, sticky = 'ew')

    button_open = tk.Button(tronde_frame_EN, text = 'Mở thư mục chứa đề', command = (lambda : vbf.open_folder(entry_input_EN)), bg = '#99FF66', fg = 'black', height = 2)

    button_open.grid(row = 7, column = 1, columnspan = 2, padx = 5, pady = 5, sticky = 'ew')

    chucnangkhac_frame = ttk.LabelFrame(tab_EN, text = '3. Chức năng khác: Tạo file mẫu và ...', padding = '2')

    chucnangkhac_frame.grid(row = 3, column = 0, columnspan = 3, sticky = 'ew', padx = (5, 2), pady = 2)

    chucnangkhac_frame.columnconfigure(1, weight = 1)

    button_khac = tk.Button(chucnangkhac_frame, text = 'Chức năng khác', command = (lambda : vme.chuc_nang_khac_EN_mix(root, word, addin_name)), width = 20, height = 2, bg = 'moccasin', fg = 'black')

    button_khac.grid(row = 8, column = 0, padx = 3, columnspan = 3, pady = 5)

    tabD.columnconfigure(0, weight = 1)

    tabD.columnconfigure(1, weight = 1)

    thongtin_dangki = tk.Label(tabD, text = thongtin_dk_full, fg = 'black', bg = 'yellowgreen')

    thongtin_dangki.grid(row = 0, column = 0, columnspan = 2, padx = 5, pady = (25, 5), sticky = 'ew')

    thongtin_label = tk.Label(tabD, text = Quang_cao, fg = 'black', bg = 'burlywood', anchor = 'w', justify = 'left')

    thongtin_label.grid(row = 1, column = 0, columnspan = 2, padx = 5, pady = 5, sticky = 'ew')

    button_dk1 = tk.Button(tabD, text = 'Đăng kí cho máy tính', command = (lambda : dangki.giaodien_dangki(Quang_cao, formatted_date_het_han, root)), bg = 'cyan', fg = 'black', width = 20, height = 2)

    button_dk1.grid(row = 7, column = 0, padx = 5, pady = 5)

    button_dk2 = tk.Button(tabD, text = 'Đăng kí cho USB', command = (lambda : dangkiUSB.giaodien_dangki(Quang_cao, formatted_date_het_han, root)), bg = 'cyan', fg = 'black', width = 20, height = 2)

    button_dk2.grid(row = 7, column = 1, padx = 5, pady = 5)

    thongtin_capnhat = tk.Label(tabD, text = 'Quý thầy cô thường xuyên bấm vào đường dẫn sau\nđể cập nhật bản mới')

    thongtin_capnhat.grid(row = 8, column = 0, columnspan = 2, padx = 5, pady = 5)

    button_update = tk.Button(tabD, text = 'Cập nhật bản mới', command = open_link, bg = 'lightgreen', fg = 'black', width = 25, height = 2)

    button_update.grid(row = 9, column = 0, columnspan = 2, padx = 5, pady = 5)

    root.wm_attributes('-topmost', True)

    root.mainloop()



kiem_tra_cho_su_dung(root)

