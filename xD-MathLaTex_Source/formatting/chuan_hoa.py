import os

import datetime

import win32com.client as win32com

from win32com.client import client as win32

import win32api

import re

import sys

import shutil

import pythoncom

import tempfile

import wmi

import docx

import random

from docx import Document

from docx.shared import Pt

from docx.enum.text import WD_PARAGRAPH_ALIGNMENT

from docx.shared import RGBColor

from openpyxl import Workbook

from openpyxl.styles import Font

from openpyxl.styles import Alignment

from openpyxl.styles import Border, Side

from openpyxl.styles import PatternFill

from openpyxl.styles import GradientFill

from openpyxl.utils import get_column_letter

from openpyxl import load_workbook

import tkinter as tk

from tkinter import ttk, Label, filedialog, Entry, Button, Frame, Listbox, Scrollbar, messagebox

from tkinter import simpledialog

import webbrowser

from core import functions as vbf

from tools import tool_by_docx as docxtool



def xoa_ky_tu_an(word):

    doc = word.ActiveDocument

    find = word.Selection.Find

    invisible_chars = [

        '​',

        '‬',

        '⁠',

        '﻿',

        '‌',

        '‍']

    for invisible_char in invisible_chars:

        word.Selection.HomeKey(Unit = 6)

        find = word.Selection.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = invisible_char

        find.Replacement.Text = ''

        find.Wrap = 1

        find.Forward = True

        find.MatchCase = True

        find.Format = True

        find.MatchWildcards = True

        find.MatchWholeWord = False

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def thay_ky_tu_gia_tao_latin(word):

    '''Chuẩn hóa các ký tự Greek/Cyrillic trông giống chữ Latin trong Word.'''

    doc = word.ActiveDocument

    char_map = {

        'Α': 'A',

        'Β': 'B',

        'А': 'A',

        'В': 'B',

        'С': 'C',

        'а': 'a',

        'в': 'b',

        'с': 'c' }

    for old_char, new_char in char_map.items():

        find = word.Selection.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = old_char

        find.Replacement.Text = new_char

        find.Forward = True

        find.MatchCase = True

        find.Wrap = 1

        find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def remove_all_hyperlinks(word):

    doc = word.ActiveDocument

    links = doc.Hyperlinks

    count = links.Count

    for i in range(count, 0, -1):

        links(i).Delete()

    return None

# WARNING: Decompyle incomplete





def remove_collapsible_headings(word):

    doc = word.ActiveDocument

    paras = doc.Paragraphs

    count = paras.Count

    for i in range(1, count + 1):

        p = paras(i)

        if not p.OutlineLevel != win32.constants.wdOutlineLevelBodyText:

            continue

        p.OutlineLevel = win32.constants.wdOutlineLevelBodyText

    return None

# WARNING: Decompyle incomplete





def trang_daucham_ngoac(word):

    doc = word.ActiveDocument

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '([ ]{1,})([.:\\)])'

    find.Replacement.Text = '\\2'

    find.Wrap = 1

    find.Forward = True

    find.MatchCase = True

    find.Format = True

    find.MatchWildcards = True

    find.MatchWholeWord = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def daucham_ngoac_trang(word):

    doc = word.ActiveDocument

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '([.:\\)])([ ]{1,})'

    find.Replacement.Text = '\\1 '

    find.Wrap = 1

    find.Forward = True

    find.MatchCase = True

    find.Format = True

    find.MatchWildcards = True

    find.MatchWholeWord = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def nhan_dien_dap_an_sau_Space(word):

    vbf.thay_the_replace(word, '([ ^t]{3,})([ABCDabcd][.\\)])', '^p\\2')





def xuongdong_ABCD(word):

    doc = word.ActiveDocument

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '([^13^9])([ABCD][.])'

    find.Replacement.Text = '^p\\2'

    find.Wrap = 1

    find.Forward = True

    find.MatchCase = True

    find.Format = True

    find.MatchWildcards = True

    find.MatchWholeWord = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Execute(Replace = 2)

    find.Text = '([^13^9])([abcd][\\)])'

    find.Replacement.Text = '^p\\2'

    find.Wrap = 1

    find.Forward = True

    find.MatchCase = True

    find.Format = True

    find.MatchWildcards = True

    find.MatchWholeWord = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Execute(Replace = 2)

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '([^13^9])([  ]{1,})([ABCDabcd][.\\)])'

    find.Replacement.Text = '^p\\3'

    find.Wrap = 1

    find.Forward = True

    find.MatchCase = True

    find.Format = True

    find.MatchWildcards = True

    find.MatchWholeWord = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def dinh_cham_ABCD(word):

    doc = word.ActiveDocument

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '(^13)([ABCDabcd])([.\\)])'

    find.Replacement.Text = '^p\\2\\3 '

    find.Wrap = 1

    find.Forward = True

    find.MatchCase = True

    find.Format = True

    find.MatchWildcards = True

    find.MatchWholeWord = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Execute(Replace = 2)

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '([ABCDabcd])([.\\)])([ ^t]{1,})'

    find.Replacement.Text = '\\1\\2 '

    find.Wrap = 1

    find.Forward = True

    find.MatchCase = True

    find.Format = True

    find.MatchWildcards = True

    find.MatchWholeWord = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def dinh_cham_ABCD_all(word):

    doc = word.ActiveDocument

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '([ABCDabcd])([.\\)])'

    find.Replacement.Text = '\\1\\2 '

    find.Wrap = 1

    find.Forward = True

    find.MatchCase = True

    find.Format = True

    find.MatchWildcards = True

    find.MatchWholeWord = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def xoa_tab(word):

    doc = word.ActiveDocument

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '[^t]{1,}^13'

    find.Replacement.Text = '^p'

    find.Wrap = 1

    find.Forward = True

    find.MatchCase = True

    find.Format = True

    find.MatchWildcards = True

    find.MatchWholeWord = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def khoangtrang_13(word):

    doc = word.ActiveDocument

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '[ ]{1,}^13'

    find.Replacement.Text = '^p'

    find.Wrap = 1

    find.Forward = True

    find.MatchCase = True

    find.Format = True

    find.MatchWildcards = True

    find.MatchWholeWord = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def xoa_2_khoangtrang(word):

    vbf.xoa_multi_space(word)





def xoa_gach_chan_tab(doc, word):

    find = word.Selection.Find

    find.ClearFormatting()

    find.Font.Underline = True

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Underline = False

    word.Selection.Find.Execute(FindText = '(^t)', ReplaceWith = '\\1', Replace = 2, Forward = True, MatchCase = True, MatchWholeWord = False, MatchWildcards = True, MatchSoundsLike = False, MatchAllWordForms = False, Wrap = 1, Format = True)





def maudo_to_gachchan_chd(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange = mo_rong_vung_chon(doc, myrange)

    else:

        myrange = doc.Range()

    for i in range(1, 100):

        myrange_cau = myrange.Duplicate

        find = myrange_cau.Find

        find.ClearFormatting()

        find.Text = '([^13^9][ABCDabcd][.\\)])'

        find.MatchWildcards = True

        if find.Execute():

            myrange = doc.Range(myrange_cau.End, myrange.End)

            find = myrange_cau.Find

            find.ClearFormatting()

            find.Font.Color = win32api.RGB(255, 0, 0)

            find.Replacement.ClearFormatting()

            find.Replacement.Font.Bold = True

            find.Replacement.Font.Underline = 1

            find.Text = '([ABCDabcd])'

            find.Replacement.Text = '\\1'

            find.Forward = True

            find.MatchCase = True

            find.MatchWholeWord = False

            find.MatchWildcards = True

            find.MatchSoundsLike = False

            find.MatchAllWordForms = False

            find.Wrap = 0

            find.Format = True

            find.Execute(Replace = 1)

            continue

        range(1, 100)

        return None

    return None

# WARNING: Decompyle incomplete





def fix_dap_an_xanh_do_gachchan(word):

    doc = word.ActiveDocument

    myrange = doc.Range()

    for i in range(1, 501):

        myrange_find = myrange.Duplicate

        find = myrange_find.Find

        find.ClearFormatting()

        find.Text = '(^13)([ABCDabcd][.\\)][ ]{1,})'

        find.Replacement.Text = '^p\\2'

        find.MatchWildcards = True

        find.Format = False

        if find.Execute(Replace = 1):

            myrange.SetRange(Start = myrange_find.Start + 1, End = myrange.End)

            rng = myrange_find.Duplicate

            if rng.Tables.Count > 0:

                table = rng.Tables(1)

                table_range = table.Range

                table_end = table_range.End

                myrange.SetRange(Start = table_end, End = myrange.End)

            find = rng.Find

            find.ClearFormatting()

            find.Replacement.ClearFormatting()

            find.Text = '([ABCDabcd])'

            find.Forward = True

            find.MatchCase = True

            find.MatchWholeWord = False

            find.MatchWildcards = True

            find.MatchSoundsLike = False

            find.MatchAllWordForms = False

            find.Wrap = 0

            find.Format = True

            find.Replacement.Text = '\\1'

            find.Replacement.Font.Underline = 1

            find.Replacement.Font.Color = win32api.RGB(255, 0, 0)

            find.Replacement.Font.Bold = True

            if find.Execute():

                if rng.Font.Color == win32api.RGB(255, 0, 0) and rng.Font.Color == win32api.RGB(238, 0, 0) or rng.Font.Color == win32api.RGB(192, 0, 0):

                    find.Execute(Replace = 1)

            rng = myrange_find.Duplicate

            find = rng.Find

            find.ClearFormatting()

            find.Replacement.ClearFormatting()

            find.Text = '([ABCD][.])'

            find.Replacement.Text = '\\1'

            find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

            find.Replacement.Font.Bold = True

            find.Forward = True

            find.MatchCase = True

            find.MatchWholeWord = False

            find.MatchWildcards = True

            find.MatchSoundsLike = False

            find.MatchAllWordForms = False

            find.Wrap = 0

            find.Format = True

            find.Execute(Replace = 1)

            rng = myrange_find.Duplicate

            find = rng.Find

            find.ClearFormatting()

            find.Replacement.ClearFormatting()

            find.Text = '([abcd][\\)])'

            find.Replacement.Text = '\\1'

            find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

            find.Replacement.Font.Bold = True

            find.Forward = True

            find.MatchCase = True

            find.MatchWholeWord = False

            find.MatchWildcards = True

            find.MatchSoundsLike = False

            find.MatchAllWordForms = False

            find.Wrap = 0

            find.Format = True

            find.Execute(Replace = 1)

            for Under_line in (1, 3, 6):

                rng = myrange_find.Duplicate

                find = rng.Find

                find.ClearFormatting()

                find.Replacement.ClearFormatting()

                find.Font.Underline = Under_line

                find.Text = '([ABCDabcd])'

                find.Replacement.Text = '\\1'

                find.Replacement.Font.Underline = 1

                find.Replacement.Font.Color = win32api.RGB(255, 0, 0)

                find.Replacement.Font.Bold = True

                find.Forward = True

                find.MatchCase = True

                find.MatchWholeWord = False

                find.MatchWildcards = True

                find.MatchSoundsLike = False

                find.MatchAllWordForms = False

                find.Wrap = 0

                find.Format = True

                find.Execute(Replace = 1)

                rng = myrange_find.Duplicate

                find = rng.Find

                find.ClearFormatting()

                find.Replacement.ClearFormatting()

                find.Text = '([.\\)][ ]{1,})'

                find.Replacement.Text = '\\1'

                find.Replacement.Font.Underline = False

                find.Forward = True

                find.MatchCase = True

                find.MatchWholeWord = False

                find.MatchWildcards = True

                find.MatchSoundsLike = False

                find.MatchAllWordForms = False

                find.Wrap = 0

                find.Format = True

                find.Execute(Replace = 1)

            continue

        return None

    return None

# WARNING: Decompyle incomplete





def xoa_khoangtrang_dapan(word):

    doc = word.ActiveDocument

    vbf.thay_the_replace(word, '(ĐS:)', '^13\\1')

    vbf.thay_the_replace(word, '(ĐS:)([ ]{1,})', '\\1')

    vbf.thay_the_replace(word, '(ĐS:)([0-9]{1,})(.)([0-9]{1,})', '\\1\\2,\\4')

    vbf.thay_the_replace(word, '(ĐS:)([0-9]{1,})(,)([0-9]{1,})(.)', '\\1\\2\\3\\4')

    vbf.thay_the_replace(word, '(ĐS:)([0-9]{1,})(.)', '\\1\\2')

    return None

# WARNING: Decompyle incomplete





def page_A4_setup_Mix(doc, word):

    doc.Content.WholeStory()

    doc.Range().Select()

    page_setup = doc.PageSetup

    page_setup.TopMargin = vbf.InchesToPoints(0.4)

    page_setup.BottomMargin = vbf.InchesToPoints(0.4)

    page_setup.LeftMargin = vbf.InchesToPoints(0.4)

    page_setup.RightMargin = vbf.InchesToPoints(0.4)

    page_setup.Gutter = vbf.InchesToPoints(0)

    page_setup.HeaderDistance = vbf.InchesToPoints(0.24)

    page_setup.FooterDistance = vbf.InchesToPoints(0.24)

    page_setup.PageWidth = vbf.InchesToPoints(8.27)

    page_setup.PageHeight = vbf.InchesToPoints(11.69)





def tim_para_chua_cau_first(doc):

    myrange = doc.Range()

    find = myrange.Find

    find.Text = '^13Câu'

    find.MatchWildcards = True

    find.Wrap = 0

    if find.Execute():

        myrange.SetRange(Start = myrange.Start, End = doc.Range().End)

        myrange_find = myrange.Duplicate

        find = myrange_find.Find

        find.Text = 'HẾT'

        find.MatchWildcards = True

        find.Wrap = 1

        if find.Execute():

            myrange.SetRange(Start = myrange.Start, End = myrange_find.End)

        return myrange

    myrange = None.Range()

    find = myrange.Find

    find.Text = '^13Question'

    find.MatchWildcards = True

    find.Wrap = 0

    if find.Execute():

        myrange.SetRange(Start = myrange.Start, End = doc.Range().End)

        myrange_find = myrange.Duplicate

        find = myrange_find.Find

        find.Text = 'END'

        find.MatchWildcards = True

        find.Wrap = 1

        if find.Execute():

            myrange.SetRange(Start = myrange.Start, End = myrange_find.End)

    return myrange





def tim_para_chua_phan(doc):

    paras_with_phan = []

    for i, para in enumerate(doc.Paragraphs, start = 1):

        if not para.Range.Text.startswith('PHẦN') and 'HẾT' in para.Range.Text:

            continue

        paras_with_phan.append(i)

    return paras_with_phan





def xoa_dong_trang_new_Mix(word):

    doc = word.ActiveDocument

    myrange = doc.Range()

    myrange_find = myrange.Duplicate

    find = myrange_find.Find

    find.ClearFormatting()

    find.Text = 'PHẦN'

    find.MatchCase = False

    if find.Execute():

        myrange.SetRange(Start = myrange_find.Start, End = myrange.End)

    else:

        find.Text = 'Câu [0-9]{1,2}'

        find.MatchWildcards = True

        if find.Execute():

            myrange.SetRange(Start = myrange_find.Start, End = myrange.End)

        else:

            find.Text = 'Question [0-9]{1,2}'

            find.MatchWildcards = True

            if find.Execute():

                myrange.SetRange(Start = myrange_find.Start, End = myrange.End)

            else:

                return None

    for para in myrange.Paragraphs:

        if not para.Range.Text.strip() == '':

            continue

        para.Range.Delete()

    return None

# WARNING: Decompyle incomplete





def canh_befor_after_phan_het(doc, word):

    Danh_sach_phan = tim_para_chua_phan(doc)

    paragraphs = doc.Paragraphs

    if len(Danh_sach_phan) > 0:

        for i in Danh_sach_phan:

            paragraph_format = paragraphs(i).Range.ParagraphFormat

            paragraph_format.SpaceBefore = 6

            paragraph_format.SpaceBeforeAuto = False

        return None





def canh_befor_after(doc, word):

    myrange = tim_para_chua_cau_first(doc)

    paragraph_format = myrange.ParagraphFormat

    paragraph_format.SpaceBefore = 0

    paragraph_format.SpaceAfter = 0

    paragraph_format.SpaceBeforeAuto = False

    paragraph_format.SpaceAfterAuto = False

    paragraph_format.LineUnitBefore = 0

    paragraph_format.LineUnitAfter = 0

    paragraph_format.FirstLineIndent = vbf.InchesToPoints(0)

    paragraph_format.LeftIndent = vbf.InchesToPoints(0)

    paragraph_format.RightIndent = vbf.InchesToPoints(0)

    paragraph_format.LineSpacing = vbf.LinesToPoints(1.15)

    canh_befor_after_phan_het(doc, word)





def canh_deu_2ben(doc, word):

    for para in doc.Paragraphs:

        if not para.Alignment == 0:

            continue

        para.Alignment = 3





def page_A4_setup(word):

    doc = word.ActiveDocument

    page_A4_setup_Mix(doc, word)

    doc.Content.WholeStory()

    doc.Range().Select()

    canh_befor_after(doc, word)

    doc.Range().Select()

    canh_deu_2ben(doc, word)

    return None

# WARNING: Decompyle incomplete





def page_2cot_Mix(doc, word):

    doc.Content.WholeStory()

    doc.Range().Select()

    columns = doc.PageSetup.TextColumns

    columns.SetCount(NumColumns = 2)

    columns.EvenlySpaced = True

    columns.LineBetween = True

    columns.Width = vbf.InchesToPoints(3.59)

    columns.Spacing = vbf.InchesToPoints(0.1)

    doc.Content.WholeStory()

    doc.Range().Select()

    page_setup = doc.PageSetup

    page_setup.TopMargin = vbf.InchesToPoints(0.24)

    page_setup.BottomMargin = vbf.InchesToPoints(0.24)

    page_setup.LeftMargin = vbf.InchesToPoints(0.34)

    page_setup.RightMargin = vbf.InchesToPoints(0.34)

    page_setup.Gutter = vbf.InchesToPoints(0)

    page_setup.HeaderDistance = vbf.InchesToPoints(0.24)

    page_setup.FooterDistance = vbf.InchesToPoints(0.24)

    page_setup.PageWidth = vbf.InchesToPoints(8.27)

    page_setup.PageHeight = vbf.InchesToPoints(11.69)

    doc.Content.WholeStory()

    doc.Range().Select()

    paragraph_format = word.Selection.ParagraphFormat

    paragraph_format.SpaceBefore = 0

    paragraph_format.SpaceBeforeAuto = False

    paragraph_format.SpaceAfterAuto = False

    paragraph_format.LineUnitBefore = 0

    paragraph_format.SpaceBeforeAuto = False

    paragraph_format.SpaceAfter = 0

    paragraph_format.SpaceAfterAuto = False

    paragraph_format.LineUnitAfter = 0

    paragraph_format.FirstLineIndent = vbf.InchesToPoints(0)

    paragraph_format.LeftIndent = vbf.InchesToPoints(0)

    paragraph_format.RightIndent = vbf.InchesToPoints(0)

    paragraph_format.LineSpacing = vbf.LinesToPoints(1.15)





def page_2cot(word):

    doc = word.ActiveDocument

    page_2cot_Mix(doc, word)

    return None

# WARNING: Decompyle incomplete





def mu11_qua_mu13_Mix(doc, word):

    vbf.thay_the_replace(word, '^l', '^p')

    vbf.thay_the_replace(word, '^13', '^p')

    vbf.thay_the_replace(word, '^m', '^p')





def xuongdong_ABCD_bold(word):

    doc = word.ActiveDocument

    myrange = doc.Range()

    for i in range(1, 200):

        myrange_find = myrange.Duplicate

        find = myrange_find.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = '([ ]{2,})([ABCD].)'

        find.Wrap = 0

        find.Forward = True

        find.MatchCase = True

        find.Format = True

        find.MatchWildcards = True

        find.MatchWholeWord = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        if find.Execute():

            myrange = doc.Range(myrange_find.End, myrange.End)

            find = myrange_find.Find

            find.ClearFormatting()

            find.Replacement.ClearFormatting()

            find.Text = '([ABCD])'

            find.Font.Bold = True

            find.Replacement.Text = '^p\\1'

            find.Wrap = 0

            find.Forward = True

            find.MatchCase = True

            find.Format = True

            find.MatchWildcards = True

            find.MatchWholeWord = True

            find.MatchSoundsLike = False

            find.MatchAllWordForms = False

            find.Execute(Replace = 1)

            continue

        range(1, 200)

    khoangtrang_13(word)

    return None

# WARNING: Decompyle incomplete





def xuongdong_phuongan_lite(word):

    doc = word.ActiveDocument

    khoangtrang_13(word)

    vbf.add_blank_line_after_table(word)

    fix_dap_an_xanh_do_gachchan(word)

    xoa_dong_trang_new_Mix(word)

    return None

# WARNING: Decompyle incomplete





def xuongdong_phuongan(word):

    doc = word.ActiveDocument

    vbf.add_blank_line_after_table(word)

    remove_all_hyperlinks(word)

    remove_collapsible_headings(word)

    delete_bookmarks(doc, word)

    vbf.thay_the_replace(word, chr(160), ' ')

    xoa_ky_tu_an(word)

    thay_ky_tu_gia_tao_latin(word)

    vbf.thay_the_replace(word, '^l', '^p')

    vbf.thay_the_replace(word, '^13', '^p')

    vbf.thay_the_replace(word, '^m', '^p')

    vbf.thay_the_replace(word, '(^13)([^9 ]{1,})(PHẦN)', '\\1\\3')

    vbf.thay_the_replace(word, '(^13PHẦN 1.)', '^13PHẦN I.')

    vbf.thay_the_replace(word, '(^13PHẦN 2.)', '^13PHẦN II.')

    vbf.thay_the_replace(word, '(^13PHẦN 3.)', '^13PHẦN III.')

    vbf.thay_the_replace(word, '(^13PHẦN 4.)', '^13PHẦN IV.')

    trang_daucham_ngoac(word)

    daucham_ngoac_trang(word)

    vbf.fix_cau(word)

    nhan_dien_dap_an_sau_Space(word)

    xuongdong_ABCD_bold(word)

    xuongdong_ABCD(word)

    dinh_cham_ABCD(word)

    xoa_tab(word)

    khoangtrang_13(word)

    xoa_2_khoangtrang(word)

    fix_dap_an_xanh_do_gachchan(word)

    canh_befor_after(doc, word)

    canh_deu_2ben(doc, word)

    vbf.canh_pict_giua(doc, word)

    vbf.chuanhoa_loigiai(word)

    xoa_dong_trang_new_Mix(word)

    return None

# WARNING: Decompyle incomplete





def xuongdong_phuongan_Tool(word):

    doc = word.ActiveDocument

    xuongdong_phuongan(word)

    vbf.chuanhoa_dap_so_py32(word)

    dem_so = vbf.check_pic_float(doc, word)

    vbf.pic_inline_center(word)

    vbf.autofit_table(word)

    messages_sentences = []

# WARNING: Decompyle incomplete





def xuongdong_phuongan_for_chuanhoa_tool(word, messages_sentences):

    doc = word.ActiveDocument

    xuongdong_phuongan(word)

    dem_so = vbf.check_pic_float(doc, word)

    vbf.pic_inline_center(word)

    vbf.autofit_table(word)

    if dem_so > 0:

        messages_sentences.append(f'''Có {dem_so} hình ảnh không ở chế độ pict in line.\n Tôi đã giúp bạn về dạng pict in line nhưng bạn phải kiểm tra lại vị trí xuất hiện của nó để chỉnh lại phù hợp''')

    vbf.check_sentences_in_tables(word, messages_sentences)

    return None

# WARNING: Decompyle incomplete





def xuong_dong_nhanh_combine_docx(word):

    doc = word.ActiveDocument

    doc_path = doc.FullName

    doc_path = os.path.normpath(os.path.abspath(doc_path))

    bad_part = '\\Python\\VBA\\PC7\\https:\\tcdcnh-my.sharepoint.com\\personal\\hungnn_giahoi_cee_edu_vn\\Documents'

    if bad_part in doc_path:

        doc_path = doc_path.replace(bad_part, '')

    if not os.path.exists(doc_path):

        messagebox.showerror('Lỗi', 'File không tồn tại!')

        return None

    doc.Save()

    doc_path = doc.FullName

    doc_path = os.path.normpath(os.path.abspath(doc_path))

    bad_part = '\\Python\\VBA\\PC7\\https:\\tcdcnh-my.sharepoint.com\\personal\\hungnn_giahoi_cee_edu_vn\\Documents'

    if bad_part in doc_path:

        doc_path = doc_path.replace(bad_part, '')

    doc_path_name = os.path.splitext(doc_path)[0]

    doc_ext = os.path.splitext(doc_path)[1]

    if doc_ext.lower() != '.docx':

        new_file_name = file_name + '.docx'

        doc.SaveAs(FileName = new_file_name, FileFormat = 12)

        doc.Convert()

        doc.Save()

    doc_path = doc.FullName

    doc_path = os.path.normpath(os.path.abspath(doc_path))

    if bad_part in doc_path:

        doc_path = doc_path.replace(bad_part, '')

    doc.Close(SaveChanges = True)

    file_docx = docxtool.open_doc_off_python_docx(doc_path)

    messages = []

    docxtool.xuong_dong_phuong_an_docx(file_docx, messages)

    file_docx.save(doc_path)

    word.Documents.Open(doc_path)

    print('Đã xong xuống dòng nhanh by docx')

    messages.append('Đã xong!')

    if messages:

        thong_bao = '\n'.join(messages)

        messagebox.showinfo('Thông báo', thong_bao)

        return None

    return None

# WARNING: Decompyle incomplete





def Fix_du_lieu_xuong_dong(root, word):

    pass

# WARNING: Decompyle incomplete





def tab_btp(doc, word):

    selection = word.Selection

    selection.ParagraphFormat.TabStops.ClearAll()

    doc.DefaultTabStop = vbf.InchesToPoints(0.2)

    wdAlignTabLeft = 0

    wdTabLeaderSpaces = 0

    tab_stops = [

        0.21,

        2.07,

        3.93,

        5.79]

    for pos in tab_stops:

        selection.ParagraphFormat.TabStops.Add(Position = vbf.InchesToPoints(pos), Alignment = wdAlignTabLeft, Leader = wdTabLeaderSpaces)





def tab_btp_2_1_1(doc, word):

    selection = word.Selection

    selection.ParagraphFormat.TabStops.ClearAll()

    doc.DefaultTabStop = vbf.InchesToPoints(0.2)

    wdAlignTabLeft = 0

    wdTabLeaderSpaces = 0

    tab_stops = [

        0.21,

        3.93]

    for pos in tab_stops:

        selection.ParagraphFormat.TabStops.Add(Position = vbf.InchesToPoints(pos), Alignment = wdAlignTabLeft, Leader = wdTabLeaderSpaces)





def tab_btp_2cot(doc, word):

    selection = word.Selection

    selection.ParagraphFormat.TabStops.ClearAll()

    doc.DefaultTabStop = vbf.InchesToPoints(0.04)

    wdAlignTabLeft = 0

    wdTabLeaderSpaces = 0

    tab_stops = [

        0.1,

        0.98,

        1.86,

        2.74]

    for pos in tab_stops:

        selection.ParagraphFormat.TabStops.Add(Position = vbf.InchesToPoints(pos), Alignment = wdAlignTabLeft, Leader = wdTabLeaderSpaces)





def tab_btp_2cot_2_1_1(doc, word):

    selection = word.Selection

    selection.ParagraphFormat.TabStops.ClearAll()

    doc.DefaultTabStop = vbf.InchesToPoints(0.04)

    wdAlignTabLeft = 0

    wdTabLeaderSpaces = 0

    tab_stops = [

        0.1,

        1.86]

    for pos in tab_stops:

        selection.ParagraphFormat.TabStops.Add(Position = vbf.InchesToPoints(pos), Alignment = wdAlignTabLeft, Leader = wdTabLeaderSpaces)





def delete_bookmarks(doc, word):

    for bookmark in doc.Bookmarks:

        if not bookmark.Name != 'MDH':

            continue

        if not bookmark.Name != 'MDH2':

            continue

        if not bookmark.Name != 'num_page':

            continue

        bookmark.Delete()





def add_bookmark(doc, word, paragraph, bookmark_name):

    bookmark = doc.Bookmarks.Add(bookmark_name, paragraph.Range)





def create_bookmarks_cau(doc, word):

    delete_bookmarks(doc, word)

    bookmark_count = 1

    sobmabcd = 1

    sobman = 1

    sobmbn = 1

    sobmcn = 1

    sobmdn = 1

    for paragraph in doc.Paragraphs:

        if paragraph.Range.Text.startswith('Câu') or paragraph.Range.Text.startswith('Question'):

            bookmark_name = f'''c{bookmark_count}q'''

            add_bookmark(doc, word, paragraph, bookmark_name)

            bookmark_count += 1

            continue

        if paragraph.Range.Text.startswith('A.'):

            bookmark_name = f'''c{sobmabcd}a'''

            add_bookmark(doc, word, paragraph, bookmark_name)

            continue

        if paragraph.Range.Text.startswith('B.'):

            bookmark_name = f'''c{sobmabcd}b'''

            add_bookmark(doc, word, paragraph, bookmark_name)

            continue

        if paragraph.Range.Text.startswith('C.'):

            bookmark_name = f'''c{sobmabcd}c'''

            add_bookmark(doc, word, paragraph, bookmark_name)

            continue

        if paragraph.Range.Text.startswith('D.'):

            bookmark_name = f'''c{sobmabcd}d'''

            add_bookmark(doc, word, paragraph, bookmark_name)

            sobmabcd = sobmabcd + 1

            continue

        if paragraph.Range.Text.startswith('a)'):

            bookmark_name = f'''c{sobman}an'''

            add_bookmark(doc, word, paragraph, bookmark_name)

            sobman = sobman + 1

            continue

        if paragraph.Range.Text.startswith('b)'):

            bookmark_name = f'''c{sobmbn}bn'''

            add_bookmark(doc, word, paragraph, bookmark_name)

            sobmbn = sobmbn + 1

            continue

        if paragraph.Range.Text.startswith('c)'):

            bookmark_name = f'''c{sobmcn}cn'''

            add_bookmark(doc, word, paragraph, bookmark_name)

            sobmcn = sobmcn + 1

            continue

        if not paragraph.Range.Text.startswith('d)'):

            continue

        bookmark_name = f'''c{sobmdn}dn'''

        add_bookmark(doc, word, paragraph, bookmark_name)

        sobmdn = sobmdn + 1

    return (sobmabcd, sobman, sobmbn, sobmcn, sobmdn)





def len_option(doc, word, bookmark_name):

    bookmark = doc.Bookmarks(bookmark_name).Range

    text_length = len(bookmark.Text.strip())

    image_count = len(list(bookmark.InlineShapes))

    image_length = 0

    for shape in bookmark.InlineShapes:

        image_length += int(shape.Width // 6)

    total_length = text_length + image_length

    return total_length





def format_options_1(doc, word, i):

    bookmark = doc.Bookmarks('c' + str(i) + 'a')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeText(Text = '\t')

    tab_btp(doc, word)

    bookmark = doc.Bookmarks('c' + str(i) + 'b')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeBackspace()

    word.Selection.TypeText(Text = '\t')

    bookmark = doc.Bookmarks('c' + str(i) + 'c')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeBackspace()

    word.Selection.TypeText(Text = '\t')

    bookmark = doc.Bookmarks('c' + str(i) + 'd')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeBackspace()

    word.Selection.TypeText(Text = '\t')





def format_options_1_2cot(doc, word, i):

    bookmark = doc.Bookmarks('c' + str(i) + 'a')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeText(Text = '\t')

    tab_btp_2cot(doc, word)

    bookmark = doc.Bookmarks('c' + str(i) + 'b')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeBackspace()

    word.Selection.TypeText(Text = '\t')

    bookmark = doc.Bookmarks('c' + str(i) + 'c')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeBackspace()

    word.Selection.TypeText(Text = '\t')

    bookmark = doc.Bookmarks('c' + str(i) + 'd')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeBackspace()

    word.Selection.TypeText(Text = '\t')





def format_options_2(doc, word, i):

    bookmark = doc.Bookmarks('c' + str(i) + 'a')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeText(Text = '\t')

    tab_btp_2_1_1(doc, word)

    bookmark = doc.Bookmarks('c' + str(i) + 'b')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeBackspace()

    word.Selection.TypeText(Text = '\t')

    bookmark = doc.Bookmarks('c' + str(i) + 'c')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeText(Text = '\t')

    tab_btp_2_1_1(doc, word)

    bookmark = doc.Bookmarks('c' + str(i) + 'd')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeBackspace()

    word.Selection.TypeText(Text = '\t')





def format_options_2_2cot(doc, word, i):

    bookmark = doc.Bookmarks('c' + str(i) + 'a')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeText(Text = '\t')

    tab_btp_2cot_2_1_1(doc, word)

    bookmark = doc.Bookmarks('c' + str(i) + 'b')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeBackspace()

    word.Selection.TypeText(Text = '\t')

    bookmark = doc.Bookmarks('c' + str(i) + 'c')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeText(Text = '\t')

    tab_btp_2cot_2_1_1(doc, word)

    bookmark = doc.Bookmarks('c' + str(i) + 'd')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeBackspace()

    word.Selection.TypeText(Text = '\t')





def format_options_4(doc, word, i):

    bookmark = doc.Bookmarks('c' + str(i) + 'a')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeText(Text = '\t')

    tab_btp(doc, word)

    bookmark = doc.Bookmarks('c' + str(i) + 'b')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeText(Text = '\t')

    tab_btp(doc, word)

    bookmark = doc.Bookmarks('c' + str(i) + 'c')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeText(Text = '\t')

    tab_btp(doc, word)

    bookmark = doc.Bookmarks('c' + str(i) + 'd')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeText(Text = '\t')

    tab_btp(doc, word)





def format_options_4_2cot(doc, word, i):

    bookmark = doc.Bookmarks('c' + str(i) + 'a')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeText(Text = '\t')

    tab_btp_2cot(doc, word)

    bookmark = doc.Bookmarks('c' + str(i) + 'b')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeText(Text = '\t')

    tab_btp_2cot(doc, word)

    bookmark = doc.Bookmarks('c' + str(i) + 'c')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeText(Text = '\t')

    tab_btp_2cot(doc, word)

    bookmark = doc.Bookmarks('c' + str(i) + 'd')

    bookmark.Range.Select()

    word.Selection.MoveLeft(Unit = 1)

    word.Selection.TypeText(Text = '\t')

    tab_btp_2cot(doc, word)





def canhtap_lamdep(doc, word):

    (sobmabcd, sobman, sobmbn, sobmcn, sobmdn) = create_bookmarks_cau(doc, word)

    for i in range(1, sobmabcd):

        L1max = max(len_option(doc, word, f'''c{i}a'''), len_option(doc, word, f'''c{i}b'''), len_option(doc, word, f'''c{i}c'''), len_option(doc, word, f'''c{i}d'''))

        if L1max < 20:

            format_options_1(doc, word, i)

        elif L1max < 45:

            format_options_2(doc, word, i)

        else:

            format_options_4(doc, word, i)

    continue

    vbf.thay_the_replace_Mix(doc, word, '(^13)([abcd])', '\\1^9\\2')

    for i in range(1, sobman):

        bookmark = doc.Bookmarks('c' + str(i) + 'an')

        bookmark.Range.Select()

        word.Selection.MoveLeft(Unit = 1)

        tab_btp(doc, word)

    for i in range(1, sobmbn):

        bookmark = doc.Bookmarks('c' + str(i) + 'bn')

        bookmark.Range.Select()

        word.Selection.MoveLeft(Unit = 1)

        tab_btp(doc, word)

    for i in range(1, sobmcn):

        bookmark = doc.Bookmarks('c' + str(i) + 'cn')

        bookmark.Range.Select()

        word.Selection.MoveLeft(Unit = 1)

        tab_btp(doc, word)

    for i in range(1, sobmdn):

        bookmark = doc.Bookmarks('c' + str(i) + 'dn')

        bookmark.Range.Select()

        word.Selection.MoveLeft(Unit = 1)

        tab_btp(doc, word)

    delete_bookmarks(doc, word)

    return None

# WARNING: Decompyle incomplete





def canhtap_lamdep_2cot(doc, word):

    (sobmabcd, sobman, sobmbn, sobmcn, sobmdn) = create_bookmarks_cau(doc, word)

    for i in range(1, sobmabcd):

        L1max = max(len_option(doc, word, f'''c{i}a'''), len_option(doc, word, f'''c{i}b'''), len_option(doc, word, f'''c{i}c'''), len_option(doc, word, f'''c{i}d'''))

        if L1max < 11:

            format_options_1_2cot(doc, word, i)

        elif L1max < 22:

            format_options_2_2cot(doc, word, i)

        else:

            format_options_4_2cot(doc, word, i)

    continue

    delete_bookmarks(doc, word)

    return None

# WARNING: Decompyle incomplete





def chia_mot_cot(word):

    doc = word.ActiveDocument

    doc.Content.WholeStory()

    columns = doc.PageSetup.TextColumns

    columns.SetCount(NumColumns = 1)

    return None

# WARNING: Decompyle incomplete





def chia_hai_cot_H(word):

    doc = word.ActiveDocument

    page_A4_setup(word)

    page_2cot(word)

    tab_btp_2cot(doc, word)

    vbf.pic_inline_center(word)

    resize_images_to_column_width(word)

    return None

# WARNING: Decompyle incomplete





def chinhlai_boder(doc, word):

    for paragraph in doc.Paragraphs:

        if paragraph.Range.Shading.BackgroundPatternColor != -16777216:

            paragraph.Range.Select()

            paragraph_format = word.Selection.ParagraphFormat

            paragraph_format.LeftIndent = vbf.InchesToPoints(0.06)

            paragraph_format.RightIndent = vbf.InchesToPoints(0.06)

            paragraph_format.FirstLineIndent = vbf.InchesToPoints(0)

    continue

    return None

# WARNING: Decompyle incomplete





def Chuanhoade(word):

    doc = word.ActiveDocument

    messages_sentences = []

    xuongdong_phuongan_for_chuanhoa_tool(word, messages_sentences)

    page_A4_setup(word)

    chia_mot_cot(word)

    vbf.add_blank_line_after_table(word)

    canhtap_lamdep(doc, word)

    xoa_gach_chan_tab(doc, word)

    xoa_dong_trang_new_Mix(word)

    vbf.pic_inline_center(word)

    messages_sentences.append('Đã hoàn thành')

    if messages_sentences:

        thong_bao = '\n'.join(messages_sentences)

        messagebox.showinfo('Thông báo', thong_bao)

        return None

    return None

# WARNING: Decompyle incomplete





def Chuanhoade_2cot(word):

    doc = word.ActiveDocument

    messages_sentences = []

    xuongdong_phuongan_for_chuanhoa_tool(word, messages_sentences)

    page_A4_setup(word)

    page_2cot(word)

    vbf.autofit_table(word)

    vbf.add_blank_line_after_table(word)

    canhtap_lamdep_2cot(doc, word)

    xoa_gach_chan_tab(doc, word)

    chinhlai_boder(doc, word)

    canh_befor_after(doc, word)

    canh_deu_2ben(doc, word)

    xoa_dong_trang_new_Mix(word)

    vbf.resize_images_to_column_width_Mix(doc, word)

    vbf.pic_inline_center(word)

    messages_sentences.append('Đã hoàn thành')

    if messages_sentences:

        thong_bao = '\n'.join(messages_sentences)

        messagebox.showinfo('Thông báo', thong_bao)

        return None

    return None

# WARNING: Decompyle incomplete





def Chuanhoade_2cot_note(word):

    doc = word.ActiveDocument

    messages_sentences = []

    xuongdong_phuongan_for_chuanhoa_tool(word, messages_sentences)

    vbf.autofit_table(word)

    vbf.add_blank_line_after_table(word)

    canhtap_lamdep_2cot(doc, word)

    xoa_gach_chan_tab(doc, word)

    canh_befor_after(doc, word)

    canh_deu_2ben(doc, word)

    xoa_dong_trang_new_Mix(word)

    vbf.resize_images_to_column_width_Mix(doc, word)

    vbf.pic_inline_center(word)

    messages_sentences.append('Đã hoàn thành')

    if messages_sentences:

        thong_bao = '\n'.join(messages_sentences)

        messagebox.showinfo('Thông báo', thong_bao)

        return None

    return None

# WARNING: Decompyle incomplete





def chuan_hoa_note(word):

    

    def inches(inch):

        return inch * 72



    doc = word.ActiveDocument

    so_dong = 40

    page_setup = doc.PageSetup

    page_setup.TopMargin = inches(0.34)

    page_setup.BottomMargin = inches(0.24)

    page_setup.LeftMargin = inches(0.34)

    page_setup.RightMargin = inches(4.19)

    page_setup.Gutter = 0

    page_setup.HeaderDistance = inches(0.24)

    page_setup.FooterDistance = inches(0.24)

    page_setup.PageWidth = inches(8.27)

    page_setup.PageHeight = inches(11.69)

    textbox_width = page_setup.RightMargin - inches(0.34)

    textbox_left = (page_setup.PageWidth - page_setup.RightMargin) + inches(0.1)

    textbox_height = page_setup.PageHeight - page_setup.TopMargin - page_setup.BottomMargin

    header = doc.Sections(1).Headers(1)

    shape = header.Shapes.AddTextbox(Orientation = 1, Left = textbox_left, Top = page_setup.TopMargin, Width = textbox_width, Height = textbox_height)

    shape.WrapFormat.Type = 5

    shape.Line.Visible = False

    shape.Fill.Visible = False

# WARNING: Decompyle incomplete





def canh_le_tab_tool(root, word):

    pass

# WARNING: Decompyle incomplete





def chuan_hoa_nhanh_A4(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        rng = word.Selection.Range

        rng.Select()

        tab_btp(doc, word)

        rng.Select()

        vbf.autofit_table(word)

        paragraph_format = rng.ParagraphFormat

        paragraph_format.LeftIndent = vbf.InchesToPoints(0)

        paragraph_format.RightIndent = vbf.InchesToPoints(0)

        paragraph_format.FirstLineIndent = vbf.InchesToPoints(0)

        rng.Select()

        vbf.thay_the_replace_stop(word, '(^13)([ABCDabcd][.)])', '\\1^9\\2')

        rng.Select()

        vbf.thay_the_replace_stop(word, '(^9)([ ]{1,})([ABCDabcd][.)])', '\\1\\3')

        rng.Select()

        vbf.thay_the_replace_stop(word, '([ ]{2,})([ABCDabcd][.)])', '^9\\2')

        canh_deu_2ben(doc, word)

        return None

    rng = doc.Content

    page_A4_setup(word)

    rng.Select()

    chia_mot_cot(word)

    rng.Select()

    tab_btp(doc, word)

    vbf.autofit_table(word)

    vbf.add_blank_line_after_table(word)

    vbf.thay_the_replace(word, '(^13)([ABCDabcd][.)])', '\\1^9\\2')

    vbf.thay_the_replace(word, '(^9)([ ]{1,})([ABCDabcd][.)])', '\\1\\3')

    vbf.thay_the_replace(word, '([ ]{2,})([ABCDabcd][.)])', '^9\\2')

    canh_befor_after(doc, word)

    canh_deu_2ben(doc, word)

    xoa_dong_trang_new_Mix(word)

    return None

# WARNING: Decompyle incomplete





def chuan_hoa_nhanh_2cot(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        rng = word.Selection.Range

        rng.Select()

        tab_btp_2cot(doc, word)

        rng.Select()

        vbf.autofit_table(word)

        paragraph_format = rng.ParagraphFormat

        paragraph_format.LeftIndent = vbf.InchesToPoints(0)

        paragraph_format.RightIndent = vbf.InchesToPoints(0)

        paragraph_format.FirstLineIndent = vbf.InchesToPoints(0)

        canh_deu_2ben(doc, word)

        return None

    rng = doc.Content

    page_A4_setup(word)

    page_2cot(word)

    rng.Select()

    tab_btp_2cot(doc, word)

    rng.Select()

    vbf.autofit_table(word)

    vbf.pic_inline_center(word)

    vbf.resize_images_to_column_width(word)

    chinhlai_boder(doc, word)

    vbf.add_blank_line_after_table(word)

    canh_befor_after(doc, word)

    canh_deu_2ben(doc, word)

    xoa_dong_trang_new_Mix(word)

    return None

# WARNING: Decompyle incomplete





def chuan_hoa_nhanh_2c_by_docx(word):

    doc = word.ActiveDocument

    print('1.doc.active rồi')

    doc_path = doc.FullName

    doc_path = os.path.normpath(os.path.abspath(doc_path))

    bad_part = '\\Python\\VBA\\PC7\\https:\\tcdcnh-my.sharepoint.com\\personal\\hungnn_giahoi_cee_edu_vn\\Documents'

    if bad_part in doc_path:

        doc_path = doc_path.replace(bad_part, '')

    print(doc_path)

    if not os.path.exists(doc_path):

        messagebox.showerror('Lỗi', 'File không tồn tại!')

        return None

    doc.Save()

    doc_path = doc.FullName

    doc_path = os.path.normpath(os.path.abspath(doc_path))

    bad_part = '\\Python\\VBA\\PC7\\https:\\tcdcnh-my.sharepoint.com\\personal\\hungnn_giahoi_cee_edu_vn\\Documents'

    if bad_part in doc_path:

        doc_path = doc_path.replace(bad_part, '')

    doc_path_name = os.path.splitext(doc_path)[0]

    doc_ext = os.path.splitext(doc_path)[1]

    if doc_ext.lower() != '.docx':

        new_file_name = file_name + '.docx'

        doc.SaveAs(FileName = new_file_name, FileFormat = 12)

        doc.Convert()

        doc.Save()

    doc_path = doc.FullName

    doc_path = os.path.normpath(os.path.abspath(doc_path))

    if bad_part in doc_path:

        doc_path = doc_path.replace(bad_part, '')

    doc_path_name = os.path.splitext(doc_path)[0]

    doc_ext = os.path.splitext(doc_path)[1]

    doc.Close(SaveChanges = True)

    file_docx = docxtool.open_doc_off_python_docx(doc_path)

    messages = []

    docxtool.chuanhoa_bo_DA_2_cot(file_docx, messages)

    doc_path_new = doc_path_name + '_HS.docx'

    file_docx.save(doc_path_new)

    word.Documents.Open(doc_path_new)

    print('Đã xong nhanh chuẩn hóa 2 cột')

    return None

# WARNING: Decompyle incomplete





def chuan_hoa_nhanh_A4_by_docx(word):

    doc = word.ActiveDocument

    print('1.doc.active rồi')

    doc_path = doc.FullName

    doc_path = os.path.normpath(os.path.abspath(doc_path))

    bad_part = '\\Python\\VBA\\PC7\\https:\\tcdcnh-my.sharepoint.com\\personal\\hungnn_giahoi_cee_edu_vn\\Documents'

    if bad_part in doc_path:

        doc_path = doc_path.replace(bad_part, '')

    print(doc_path)

    if not os.path.exists(doc_path):

        messagebox.showerror('Lỗi', 'File không tồn tại!')

        return None

    doc.Save()

    doc_path = doc.FullName

    doc_path = os.path.normpath(os.path.abspath(doc_path))

    bad_part = '\\Python\\VBA\\PC7\\https:\\tcdcnh-my.sharepoint.com\\personal\\hungnn_giahoi_cee_edu_vn\\Documents'

    if bad_part in doc_path:

        doc_path = doc_path.replace(bad_part, '')

    doc_path_name = os.path.splitext(doc_path)[0]

    doc_ext = os.path.splitext(doc_path)[1]

    if doc_ext.lower() != '.docx':

        new_file_name = file_name + '.docx'

        doc.SaveAs(FileName = new_file_name, FileFormat = 12)

        doc.Convert()

        doc.Save()

    doc_path = doc.FullName

    doc_path = os.path.normpath(os.path.abspath(doc_path))

    if bad_part in doc_path:

        doc_path = doc_path.replace(bad_part, '')

    doc_path_name = os.path.splitext(doc_path)[0]

    doc_ext = os.path.splitext(doc_path)[1]

    doc.Close(SaveChanges = True)

    file_docx = docxtool.open_doc_off_python_docx(doc_path)

    messages = []

    docxtool.chuanhoa_bo_DA_docx(file_docx, messages)

    doc_path_new = doc_path_name + '_HS.docx'

    file_docx.save(doc_path_new)

    word.Documents.Open(doc_path_new)

    print('Đã xong nhanh chuẩn hóa A4')

    return None

# WARNING: Decompyle incomplete





def Check_ABCD_select(word):

    doc = word.ActiveDocument

    messages = []

    if word.Selection.Type != 1:

        vbf.add_blank_line_after_table(word)

        myrange = word.Selection.Range

        for i in range(1, 50):

            count_DA = 0

            count_DAD = 0

            myrange_cau = myrange.Duplicate

            find = myrange_cau.Find

            find.ClearFormatting()

            find.Text = '(Câu [0-9]{1,2}.)(*)(Câu [0-9]{1,2}.)'

            find.MatchWildcards = True

            if find.Execute():

                myrange.SetRange(Start = myrange_cau.End - 8, End = myrange.End)

                for k in range(1, 6):

                    myrange_find = myrange_cau.Duplicate

                    find = myrange_find.Find

                    find.ClearFormatting()

                    find.Text = '([^13^9])([A-D].)'

                    find.MatchWildcards = True

                    if not find.Execute():

                        continue

                    count_DA = count_DA + 1

                    myrange_cau.SetRange(Start = myrange_find.End, End = myrange_cau.End)

                    myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                    text = myrange_find.Text.strip()

                    if not myrange_find.Font.Underline == True and myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                        continue

                    count_DAD = count_DAD + 1

                if count_DA != 4:

                    messages.append(f'''Câu số {i} của bạn có số lượng đáp án là {count_DA} - KHÔNG hợp lệ''')

                if not count_DAD != 1:

                    continue

                messages.append(f'''Câu số {i} của bạn có {count_DAD} đáp án ĐÚNG - KHÔNG hợp lệ''')

                continue

        count_DA = 0

        count_DAD = 0

        myrange_cau = myrange.Duplicate

        find = myrange_cau.Find

        find.ClearFormatting()

        find.Text = '(Câu [0-9]{1,2}.)'

        find.MatchWildcards = True

        if find.Execute():

            myrange_cau.SetRange(Start = myrange_cau.End, End = myrange.End)

            for k in range(1, 6):

                myrange_find = myrange_cau.Duplicate

                find = myrange_find.Find

                find.ClearFormatting()

                find.Text = '([^13^9])([A-D].)'

                find.MatchWildcards = True

                if not find.Execute():

                    continue

                count_DA = count_DA + 1

                myrange_cau.SetRange(Start = myrange_find.End, End = myrange_cau.End)

                myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                text = myrange_find.Text.strip()

                if not myrange_find.Font.Underline == True and myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                    continue

                count_DAD = count_DAD + 1

            if count_DA != 4:

                messages.append(f'''Câu cuối của bạn có số lượng đáp án là {count_DA} - KHÔNG hợp lệ''')

            if count_DAD != 1:

                messages.append(f'''Câu cuối của bạn có {count_DAD} đáp án ĐÚNG - KHÔNG hợp lệ''')

        if messages:

            messages.append('SAU KHI SỬA CHỮA HÃY KIỂM TRA LẠI')

            thong_bao = '\n'.join(messages)

            messagebox.showinfo('Thông báo', thong_bao)

            return None

        messagebox.showinfo('Thông báo', 'Chưa phát hiện lỗi')

        return None

    messagebox.showinfo('Thông báo', 'Hãy chọn vùng chứa Câu ABCD')

    return None

# WARNING: Decompyle incomplete



if __name__ == '__main__':

    root = tk.Tk()

    word = vbf.khoi_tao_word()

    chuan_hoa_nhanh_2c_by_docx(word)

    return None

