from spire.doc import Section, Paragraph, Table, TextRange, Shape, FileFormat, HorizontalAlignment, Regex

from spire.doc import UnderlineStyle

from spire.doc import Document as SpireDocument

from spire.doc import Color

from spire.doc import DocumentObjectType

from spire.doc import OfficeMath

from docx import Document

from docx.text.paragraph import Paragraph

from docx.shared import RGBColor

from docx.shared import Inches, Pt

from docx.oxml import OxmlElement

from docx.oxml.ns import qn

from docx.oxml import parse_xml

from lxml import etree



ElementTree

from xml.etree.ElementTree import QName

import xml.etree.ElementTree, etree

import os

import re

import random

import zipfile

import bisect

import tkinter as tk

from tkinter import ttk, Label, filedialog, Entry, Button, Frame, Listbox, Scrollbar, messagebox

from tkinter import simpledialog

import openpyxl

from openpyxl import Workbook

from openpyxl.styles import Font

from openpyxl.styles import Alignment

from openpyxl.styles import Border, Side

from openpyxl.styles import PatternFill

from openpyxl.styles import GradientFill

from openpyxl.utils import get_column_letter

from openpyxl import load_workbook

from formatting import chuan_hoa_docx as chdocx

from mixing import mix_docx_func as mdocx_f

from tools import tool_by_docx as docxtool



def open_doc_off_spire(doc_path):

    doc = SpireDocument()

    doc.LoadFromFile(doc_path)

    return doc





def save_and_close_spire(doc, output_file):

    doc.SaveToFile(output_file)

    doc.Close()

    return None

# WARNING: Decompyle incomplete





def update_so_trang_spr(doc_path):

    doc = SpireDocument()

    doc.LoadFromFile(doc_path)

    num_pages = doc.GetPageCount()

    new_text = f'''{num_pages:02}'''

    doc.Replace('<sotrang>', new_text, True, False)

    save_and_close_spire(doc, doc_path)





def delete_only_bookmark_tags(doc):

    body = doc._element.body

# WARNING: Decompyle incomplete





def delete_bookmarks(doc):

    excluded_names = {

        'MDH',

        'num_page'}

    bookmark_start_ids_to_keep = set()

    for elem in doc.element.body.iter():

        if not elem.tag == qn('w:bookmarkStart'):

            continue

        name = elem.get(qn('w:name'))

        bookmark_id = elem.get(qn('w:id'))

        if not name in excluded_names:

            continue

        bookmark_start_ids_to_keep.add(bookmark_id)

# WARNING: Decompyle incomplete





def find_phan_one(doc, text_find):

    '''Tìm chỉ mục của các đoạn văn bắt đầu bằng text_find và text_end'''

    indices = []

    body = doc.element.body

    paragraphs = list(body.iterchildren())

    for i, para in enumerate(paragraphs):

        if not para.tag.endswith('p'):

            continue

        paragraph = Paragraph(para, doc)

        text = paragraph.text

        if not text.startswith(text_find):

            continue

        indices.append(i)

    return indices





def Tao_group_tu_indices(doc, indices):

    '''Tạo nhóm đoạn văn từ danh sách chỉ mục indices'''

    groups = []

    body = doc.element.body

    paragraphs = list(body.iterchildren())

    items = paragraphs[indices[0]:indices[-1]]

    for k in range(len(indices) - 1):

        start = indices[k] - indices[0]

        end = indices[k + 1] - indices[0]

        groups.append(items[start:end])

    return groups





def tron_group_tu_indices(doc, indices):

    '''Trộn ngẫu nhiên các nhóm đoạn văn dựa trên danh sách indices'''

    if len(indices) > 2:

        groups = Tao_group_tu_indices(doc, indices)

        random.shuffle(groups)

        body = doc.element.body

        paragraphs = list(body.iterchildren())

        for i in range(indices[-1] - 1, indices[0] - 1, -1):

            body.remove(paragraphs[i])

        insert_index = indices[0]

        for group in groups:

            for element in group:

                body.insert(insert_index, element)

                insert_index += 1

        return None





def tim_chum_tu_a_den_b(doc, from_a, to_b):

    '''Tìm chỉ số bắt đầu và kết thúc của các cụm SCHUM trong đoạn từ from_a đến to_b'''

    text_start = '<SCHUM@>'

    text_end = '<ECHUM@>'

    body = doc.element.body

    paragraphs = list(body.iterchildren())

    indices_S = []

    indices_E = []

    for j in range(from_a, min(to_b, len(paragraphs))):

        para = paragraphs[j]

        if not para.tag.endswith('p'):

            continue

        paragraph = Paragraph(para, doc)

        text = paragraph.text

        if text.startswith(text_start):

            indices_S.append(j)

            continue

        if not text.startswith(text_end):

            continue

        indices_E.append(j)

    return (indices_S, indices_E)





def chon_mot_group_tu_indices(doc, indices):

    '''Chọn ngẫu nhiên một nhóm câu từ danh sách indices'''

    if len(indices) > 2:

        groups = Tao_group_tu_indices(doc, indices)

        selected_group = random.choice(groups)

        body = doc.element.body

        paragraphs = list(body.iterchildren())

        for i in range(indices[-1] - 1, indices[0] - 1, -1):

            body.remove(paragraphs[i])

        insert_index = indices[0]

        for item in selected_group:

            body.insert(insert_index, item)

            insert_index += 1

        return None





def chon_cau_CHUMCHON(doc):

    '''Chọn ngẫu nhiên một câu trong từng cụm <SCHUM@><CHON> ... <ECHUM@><CHON>'''

    (Schum, Echum) = mdocx_f.find_phan(doc, '<SCHUM@><CHON>', '<ECHUM@><CHON>')

    if len(Schum) != len(Echum):

        messagebox.showinfo('Thông báo', 'Số ký hiệu <SCHUM@><CHON> và <ECHUM@><CHON> không bằng nhau, hãy chỉnh sửa phù hợp')

        return None

    if len(Schum) > 0:

        for start, end in zip(reversed(Schum), reversed(Echum)):

            indices_cau = tim_cau_tu_a_den_b_tron_cau(doc, start + 1, end)

            chon_mot_group_tu_indices(doc, indices_cau)

        return None





def tim_cau_tu_a_den_b_tron_phuong_an(doc, from_a, to_b):

    '''Tìm danh sách câu hỏi hoặc các đoạn chứa <SCHUM@>/<ECHUM@> trong phạm vi từ from_a đến to_b'''

    pattern_cau = '^Câu [0-9]{1,}[.:]'

    pattern_hum = '<(?:SCHUM@|ECHUM@)>'

    body = doc.element.body

    paragraphs = list(body.iterchildren())

    indices_cau = []

    for j in range(from_a, min(to_b, len(paragraphs))):

        para = paragraphs[j]

        if not para.tag.endswith('p'):

            continue

        paragraph = Paragraph(para, doc)

        text = paragraph.text

        if not re.match(pattern_cau, text) and re.search(pattern_hum, text, re.IGNORECASE):

            continue

        indices_cau.append(j)

    indices_cau.append(to_b)

    return indices_cau





def tron_phuong_an_phan_I(doc):

    '''Trộn phương án trong phần S1@ đến E1@ (Phương án kiểu I)'''

    (indices_S, indices_E) = mdocx_f.find_phan(doc, 'S1@', 'E1@')

    if len(indices_S) != len(indices_E):

        messagebox.showinfo('Thông báo', 'Số kí hiệu S1@ và E1@ không bằng nhau, hãy chỉnh sửa phù hợp')

        return None

    if len(indices_S) > 0:

        for start, end in zip(indices_S, indices_E):

            cau = tim_cau_tu_a_den_b_tron_phuong_an(doc, start, end)

            for k in range(len(cau) - 1):

                phuong_an = mdocx_f.tim_phuong_an_trong_cau_phan_I(doc, cau[k], cau[k + 1])

                tron_group_tu_indices(doc, phuong_an)

        return None





def tron_phuong_an_phan_II(doc):

    '''Trộn phương án trong phần S2@ đến E2@ (Phương án kiểu II)'''

    (indices_S, indices_E) = mdocx_f.find_phan(doc, 'S2@', 'E2@')

    if len(indices_S) != len(indices_E):

        messagebox.showinfo('Thông báo', 'Số kí hiệu S2@ và E2@ không bằng nhau, hãy chỉnh sửa phù hợp')

        return None

    if len(indices_S) > 0:

        for start, end in zip(indices_S, indices_E):

            cau = tim_cau_tu_a_den_b_tron_phuong_an(doc, start, end)

            for k in range(len(cau) - 1):

                phuong_an = mdocx_f.tim_phuong_an_trong_cau_phan_II(doc, cau[k], cau[k + 1])

                tron_group_tu_indices(doc, phuong_an)

        return None





def tim_cau_tu_a_den_b_tron_cau(doc, from_a, to_b):

    '''Tìm danh sách câu hỏi từ from_a đến to_b, bỏ qua nội dung trong <SCHUM@> ... <ECHUM@>'''

    pattern_cau = '^Câu [0-9]{1,}[.:]'

    pattern_chum_start = '<SCHUM@>'

    pattern_chum_end = '<ECHUM@>'

    body = doc.element.body

    paragraphs = list(body.iterchildren())

    inside_chum = False

    indices_cau = []

    for j in range(from_a, min(to_b, len(paragraphs))):

        para = paragraphs[j]

        if not para.tag.endswith('p'):

            continue

        paragraph = Paragraph(para, doc)

        text = paragraph.text

        if re.search(pattern_chum_start, text, re.IGNORECASE):

            indices_cau.append(j)

            inside_chum = True

            continue

        if re.search(pattern_chum_end, text, re.IGNORECASE):

            inside_chum = False

            continue

        if inside_chum:

            continue

        if not re.match(pattern_cau, text):

            continue

        indices_cau.append(j)

    indices_cau.append(to_b)

    return indices_cau





def tron_cau_trong_chum(doc, from_a, to_b):

    '''Trộn thứ tự câu hỏi trong từng cụm <SCHUM@> ... <ECHUM@>'''

    (Schum, Echum) = tim_chum_tu_a_den_b(doc, from_a, to_b)

    if len(Schum) != len(Echum):

        messagebox.showinfo('Thông báo', 'Số kí hiệu <SCHUM@> và <ECHUM@> không bằng nhau, hãy chỉnh sửa phù hợp')

        return None

    if len(Schum) > 0:

        for start, end in zip(Schum, Echum):

            indices_cau = tim_cau_tu_a_den_b_tron_cau(doc, start + 1, end)

            tron_group_tu_indices(doc, indices_cau)

        return None





def tron_cau_trong_vung(doc, from_a, to_b):

    '''Trộn câu trong vùng từ from_a đến to_b'''

    indices_cau = tim_cau_tu_a_den_b_tron_cau(doc, from_a, to_b)

    tron_group_tu_indices(doc, indices_cau)





def tron_cau(doc, text_start, text_end):

    '''Trộn câu trong toàn bộ phạm vi từ text_start đến text_end'''

    (indices_S, indices_E) = mdocx_f.find_phan(doc, text_start, text_end)

    if len(indices_S) != len(indices_E):

        messagebox.showinfo('Thông báo', f'''Số kí hiệu {text_start} và {text_end} không bằng nhau, hãy chỉnh sửa phù hợp''')

        return None

    if len(indices_S) > 0:

        for start, end in zip(indices_S, indices_E):

            tron_cau_trong_chum(doc, start, end)

            tron_cau_trong_vung(doc, start, end)

        return None





def tron_cau_phan_I(doc):

    tron_cau(doc, 'S1@', 'E1@')





def tron_cau_phan_II(doc):

    tron_cau(doc, 'S2@', 'E2@')





def tron_cau_phan_III(doc):

    tron_cau(doc, 'S3@', 'E3@')





def tron_cau_phan_IV(doc):

    tron_cau(doc, 'S4@', 'E4@')





def dem_so_cau(doc):

    '''Đếm số câu hỏi trong từng phần S1@ - E1@, S2@ - E2@, S3@ - E3@'''

    cau_p1 = mdocx_f.tim_cau_trong_phan(doc, 'S1@', 'E1@')

    so_cau_p1 = (lambda .0: pass# WARNING: Decompyle incomplete

)(cau_p1()) if cau_p1 else 0

    cau_p2 = mdocx_f.tim_cau_trong_phan(doc, 'S2@', 'E2@')

    so_cau_p2 = (lambda .0: pass# WARNING: Decompyle incomplete

)(cau_p2()) if cau_p2 else 0

    cau_p3 = mdocx_f.tim_cau_trong_phan(doc, 'S3@', 'E3@')

    so_cau_p3 = (lambda .0: pass# WARNING: Decompyle incomplete

)(cau_p3()) if cau_p3 else 0

    return (so_cau_p1, so_cau_p2, so_cau_p3)





def lay_dap_an_vao_excel_TN(doc, ws, column = (1,)):

    '''Lấy đáp án đúng trong tài liệu Word và ghi vào file Excel'''

    labels = [

        'A.',

        'B.',

        'C.',

        'D.']

    labels2 = [

        'A',

        'B',

        'C',

        'D']

    body = doc.element.body

    paragraphs = list(body.iterchildren())

    all_indices_cau = mdocx_f.tim_cau_trong_phan(doc, 'S1@', 'E1@')

    dong = 0

    if len(all_indices_cau) > 0:

        for idx in range(len(all_indices_cau)):

            cau = all_indices_cau[idx]

            for k in range(len(cau) - 1):

                phuong_an = mdocx_f.tim_phuong_an_trong_cau_phan_I(doc, cau[k], cau[k + 1])

                encounter_count = 0

                for h in range(phuong_an[0], phuong_an[-1]):

                    para = paragraphs[h]

                    if not para.tag.endswith('p'):

                        continue

                    paragraph = Paragraph(para, doc)

                    text = paragraph.text.strip()

                    if not text.startswith(tuple(labels)):

                        continue

                    current_label = labels2[encounter_count % len(labels2)]

                    encounter_count += 1

                    if not paragraph.runs:

                        continue

                    first_run = paragraph.runs[0]

                    if not first_run.font.underline:

                        first_run.font.underline

                        if first_run.font.color:

                            first_run.font.color

                    is_correct = first_run.font.color.rgb == 'FF0000'

                    if not is_correct:

                        continue

                    ws.cell(row = dong + 4, column = column, value = current_label)

                    range(phuong_an[0], phuong_an[-1])

            dong += 1

        continue

        return None





def lay_dap_an_vao_excel_DS_TNmaker(doc, ws, column = (1,)):

    '''Lấy đáp án Đ/S từ tài liệu Word và ghi vào file Excel'''

    (so_cau_p1, so_cau_p2, so_cau_p3) = dem_so_cau(doc)

    dong = so_cau_p1

    labels = [

        'a)',

        'b)',

        'c)',

        'd)']

    all_indices_cau = mdocx_f.tim_cau_trong_phan(doc, 'S2@', 'E2@')

    body = doc.element.body

    paragraphs = list(body.iterchildren())

    if len(all_indices_cau) > 0:

        for idx in range(len(all_indices_cau)):

            cau = all_indices_cau[idx]

            for k in range(len(cau) - 1):

                phuong_an = mdocx_f.tim_phuong_an_trong_cau_phan_II(doc, cau[k], cau[k + 1])

                dap_an = []

                for h in range(phuong_an[0], phuong_an[-1]):

                    para = paragraphs[h]

                    if not para.tag.endswith('p'):

                        continue

                    paragraph = Paragraph(para, doc)

                    text = paragraph.text.strip()

                    if not text.startswith(tuple(labels)):

                        continue

                    if not paragraph.runs:

                        continue

                    first_run = paragraph.runs[0]

                    if not first_run.font.underline:

                        first_run.font.underline

                        if first_run.font.color:

                            first_run.font.color

                    is_correct = first_run.font.color.rgb == 'FF0000'

                    if is_correct:

                        dap_an.append('Đ')

                        continue

                    dap_an.append('S')

                answer = ''.join(dap_an)

                ws.cell(row = dong + 4, column = column, value = answer)

                dong += 1

        return None





def lay_dap_an_vao_excel_DS_Testpro(doc, ws, column = (1,)):

    '''Lấy đáp án Đ/S từ tài liệu Word và ghi vào file Excel'''

    (so_cau_p1, so_cau_p2, so_cau_p3) = dem_so_cau(doc)

    dong = so_cau_p1

    labels = [

        'a)',

        'b)',

        'c)',

        'd)']

    all_indices_cau = mdocx_f.tim_cau_trong_phan(doc, 'S2@', 'E2@')

    body = doc.element.body

    paragraphs = list(body.iterchildren())

    if len(all_indices_cau) > 0:

        for idx in range(len(all_indices_cau)):

            cau = all_indices_cau[idx]

            for k in range(len(cau) - 1):

                phuong_an = mdocx_f.tim_phuong_an_trong_cau_phan_II(doc, cau[k], cau[k + 1])

                dap_an = []

                for h in range(phuong_an[0], phuong_an[-1]):

                    para = paragraphs[h]

                    if not para.tag.endswith('p'):

                        continue

                    paragraph = Paragraph(para, doc)

                    text = paragraph.text.strip()

                    if not text.startswith(tuple(labels)):

                        continue

                    if not paragraph.runs:

                        continue

                    first_run = paragraph.runs[0]

                    if not first_run.font.underline:

                        first_run.font.underline

                        if first_run.font.color:

                            first_run.font.color

                    is_correct = first_run.font.color.rgb == 'FF0000'

                    if is_correct:

                        dap_an.append('D')

                        continue

                    dap_an.append('S')

                answer = ''.join(dap_an)

                ws.cell(row = dong + 4, column = column, value = answer)

                dong += 1

        return None





def lay_dap_an_vao_excel_TLN(doc, ws, column = (1,)):

    '''Lấy đáp án tự luận từ tài liệu Word và ghi vào file Excel'''

    (so_cau_p1, so_cau_p2, so_cau_p3) = dem_so_cau(doc)

    dong = so_cau_p1 + so_cau_p2

    all_indices_cau = mdocx_f.tim_cau_trong_phan(doc, 'S3@', 'E3@')

    body = doc.element.body

    paragraphs = list(body.iterchildren())

    if len(all_indices_cau) > 0:

        for idx in range(len(all_indices_cau)):

            cau = all_indices_cau[idx]

            answer = 'NO'

            for k in range(len(cau) - 1):

                for h in range(cau[k], cau[k + 1]):

                    para = paragraphs[h]

                    if not para.tag.endswith('p'):

                        continue

                    paragraph = Paragraph(para, doc)

                    text = paragraph.text.strip()

                    if not text.startswith('ĐS:'):

                        continue

                    answer = text.replace('ĐS:', '').strip()

                ws.cell(row = dong + 4, column = column, value = answer)

                dong += 1

        return None





def tron_and_xuat_de(check1_1_var, check1_2_var, check2_1_var, check2_2_var, check3_1_var, check4_1_var, check5_1_var, check6_1_var, check7_1_var, entry_input, entry_lantron, entry_newname):

    pass

# WARNING: Decompyle incomplete



