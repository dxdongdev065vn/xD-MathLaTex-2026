from spire.doc import Section, Paragraph, Table, TextRange, Shape, FileFormat, HorizontalAlignment, Regex, CharacterFormat

from spire.doc import BookmarkStart, BookmarkEnd

from spire.doc import UnderlineStyle

from spire.doc import Document as SpireDocument

from spire.doc import Color

import os

import re

import random

import tkinter as tk

from tkinter import ttk, Label, filedialog, Entry, Button, Frame, Listbox, Scrollbar, messagebox

from tkinter import simpledialog

from openpyxl import Workbook

from openpyxl.styles import Font

from openpyxl.styles import Alignment

from openpyxl.styles import Border, Side

from openpyxl.styles import PatternFill

from openpyxl.styles import GradientFill

from openpyxl.utils import get_column_letter

from openpyxl import load_workbook

from formatting import chuan_hoa_spr as chspr

from tools import tool_by_docx as docxtool

from mixing import mix_spr_func as mspr_f



def delete_only_bookmark_tags(doc):

    excluded_names = {

        'MDH',

        'num_page'}

    names_to_keep = set()

    bookmark_ends_to_keep = set()

    for i in range(doc.Bookmarks.Count):

        bookmark = doc.Bookmarks[i]

        if not bookmark.Name in excluded_names:

            continue

        names_to_keep.add(bookmark.Name)

        bookmark_ends_to_keep.add(bookmark.BookmarkEnd)

    for i in range(doc.Sections.Count):

        section = doc.Sections[i]

        for j in range(section.Paragraphs.Count):

            para = section.Paragraphs[j]

            index = para.ChildObjects.Count - 1

            if not index >= 0:

                continue

            child = para.ChildObjects[index]

            if isinstance(child, BookmarkStart) or child.Name not in names_to_keep:

                para.ChildObjects.RemoveAt(index)

            elif isinstance(child, BookmarkEnd) and child not in bookmark_ends_to_keep:

                para.ChildObjects.RemoveAt(index)

            index -= 1

            if index >= 0:

                continue

    continue





def delete_bookmarks(doc):

    excluded_names = {

        'MDH',

        'num_page'}

    names_to_keep = set()

    bookmark_ends_to_keep = set()

    for i in range(doc.Bookmarks.Count):

        bookmark = doc.Bookmarks[i]

        if not bookmark.Name in excluded_names:

            continue

        names_to_keep.add(bookmark.Name)

        bookmark_ends_to_keep.add(bookmark.BookmarkEnd)

    for i in range(doc.Sections.Count):

        section = doc.Sections[i]

        for j in range(section.Paragraphs.Count):

            para = section.Paragraphs[j]

            index = para.ChildObjects.Count - 1

            if not index >= 0:

                continue

            child = para.ChildObjects[index]

            if isinstance(child, BookmarkStart) or child.Name not in names_to_keep:

                para.ChildObjects.RemoveAt(index)

            elif isinstance(child, BookmarkEnd) and child not in bookmark_ends_to_keep:

                para.ChildObjects.RemoveAt(index)

            index -= 1

            if index >= 0:

                continue

    continue





def delete_run0_empty(doc, indices):

    section = doc.Sections[0]

    for p in indices:

        para = section.Body.Paragraphs[p]

        if not para != None:

            continue

        if not para.ChildObjects.Count > 0:

            continue

        first = para.ChildObjects[0]

        if not isinstance(first, TextRange):

            continue

        if not first.Text.strip() == '':

            continue

        para.ChildObjects.Remove(first)





def tim_chi_so_cau_dap_an_phan_chum(doc):

    indices = []

    section = doc.Sections[0]

    pattern1 = '^Câu [0-9]{1,}[.:]'

    pattern2 = '^Question [0-9]{1,}[.:]'

    labels = [

        'A.',

        'B.',

        'C.',

        'D.',

        'a)',

        'b)',

        'c)',

        'd)',

        'S1@',

        'E1@',

        'S2@',

        'E2@',

        'S3@',

        'E3@',

        'S4@',

        'E4@',

        '<S',

        '<E',

        'ĐS:']

    for p in range(section.Body.Paragraphs.Count):

        para = section.Body.Paragraphs[p]

        text = para.Text.strip()

        if not re.match(pattern1, text) and re.match(pattern2, text) and text.startswith(tuple(labels)):

            continue

        indices.append(p)

    return indices





def xoa_run_0_all(doc):

    indices = tim_chi_so_cau_dap_an_phan_chum(doc)

    delete_run0_empty(doc, indices)





def replace_mu_11(doc):

    for i in range(doc.Sections.Count):

        section = doc.Sections[i]

        for p in range(section.Body.Paragraphs.Count):

            para = section.Body.Paragraphs[p]

            para.Replace('\x0b', '\n', False, False)





def check_mu_11(doc):

    section = doc.Sections[0]

    for p in range(section.Body.Paragraphs.Count):

        para = section.Body.Paragraphs[p]

        if not isinstance(para, Paragraph):

            continue

        if not '\x0b' in para.Text:

            continue

        range(section.Body.Paragraphs.Count)

        return True

    return False





def Tao_group_tu_indices(doc, indices):

    a = 0

    groups = []

    items = []

    section = doc.Sections[0]

    for idx in range(indices[0], indices[-1]):

        item = section.Body.ChildObjects[idx]

        items.append(item)

    for k in range(len(indices) - 1):

        start = indices[k] - indices[0]

        end = indices[k + 1] - indices[0]

        group = items[start:end]

        groups.append(group)

        a += 1

    return groups





def tron_group_tu_indices(doc, indices):

    if len(indices) > 2:

        groups = Tao_group_tu_indices(doc, indices)

        random.shuffle(groups)

        body = doc.Sections[0].Body

        start = indices[0]

        end = indices[-1]

        for i in range(end - 1, start - 1, -1):

            body.ChildObjects.RemoveAt(i)

        insert_index = start

        for group in groups:

            for item in group:

                body.ChildObjects.Insert(insert_index, item.Clone())

                insert_index += 1

        return None





def tim_chum_tu_a_den_b(doc, from_a, to_b):

    section = doc.Sections[0]

    text_start = '<SCHUM@>'

    text_end = '<ECHUM@>'

    indices_S = []

    indices_E = []

    for j in range(from_a, to_b):

        child = section.Body.ChildObjects[j]

        if not isinstance(child, Paragraph):

            continue

        text = child.Text.strip()

        if text.startswith(text_start):

            indices_S.append(j)

            continue

        if not text.startswith(text_end):

            continue

        indices_E.append(j)

    return (indices_S, indices_E)





def chon_mot_group_tu_indices(doc, indices):

    if len(indices) > 2:

        groups = Tao_group_tu_indices(doc, indices)

        selected_group = random.choice(groups)

        body = doc.Sections[0].Body

        start = indices[0]

        end = indices[-1]

        for i in range(end - 1, start - 1, -1):

            body.ChildObjects.RemoveAt(i)

        insert_index = start

        for item in selected_group:

            body.ChildObjects.Insert(insert_index, item.Clone())

            insert_index += 1

        return None





def chon_cau_CHUMCHON(doc):

    (Schum, Echum) = mspr_f.find_phan(doc, '<SCHUM@><CHON>', '<ECHUM@><CHON>')

    if len(Schum) != len(Echum):

        messagebox.showinfo('Thông báo', 'Số kí hiệu <SCHUM@><CHON> và <ECHUM@><CHON> không bằng nhau, hãy chỉnh sửa phù hợp')

        return False

    if len(Schum) != 0:

        for start, end in zip(reversed(Schum), reversed(Echum)):

            indices_cau = tim_cau_tu_a_den_b_tron_cau(doc, start + 1, end)

            chon_mot_group_tu_indices(doc, indices_cau)

    return True





def tim_cau_tu_a_den_b_tron_phuong_an(doc, from_a, to_b):

    section = doc.Sections[0]

    pattern1 = '^Câu [0-9]{1,}[.:]'

    pattern_hum = '<(?:SCHUM@|ECHUM@)>'

    indices_cau = []

    for j in range(from_a, to_b):

        child = section.Body.ChildObjects[j]

        if not isinstance(child, Paragraph):

            continue

        if not re.match(pattern1, child.Text) and re.search(pattern_hum, child.Text, re.IGNORECASE):

            continue

        indices_cau.append(j)

    indices_cau.append(to_b)

    return indices_cau





def tron_phuong_an_phan_I(doc):

    (indices_S, indices_E) = mspr_f.find_phan(doc, 'S1@', 'E1@')

    if len(indices_S) != len(indices_E):

        messagebox.showinfo('Thông báo', 'Số kí hiệu S1@ và E2@ không bằng nhau, hãy chỉnh sửa phù hợp')

        return None

    if len(indices_S) != 0:

        for start, end in zip(indices_S, indices_E):

            cau = tim_cau_tu_a_den_b_tron_phuong_an(doc, start, end)

            for k in range(len(cau) - 1):

                phuong_an = mspr_f.tim_phuong_an_trong_cau_phan_I(doc, cau[k], cau[k + 1])

                tron_group_tu_indices(doc, phuong_an)

        return None





def tron_phuong_an_phan_II(doc):

    (indices_S, indices_E) = mspr_f.find_phan(doc, 'S2@', 'E2@')

    if len(indices_S) != len(indices_E):

        messagebox.showinfo('Thông báo', 'Số kí hiệu S1@ và E2@ không bằng nhau, hãy chỉnh sửa phù hợp')

        return None

    if len(indices_S) != 0:

        for start, end in zip(indices_S, indices_E):

            cau = tim_cau_tu_a_den_b_tron_phuong_an(doc, start, end)

            for k in range(len(cau) - 1):

                phuong_an = mspr_f.tim_phuong_an_trong_cau_phan_II(doc, cau[k], cau[k + 1])

                tron_group_tu_indices(doc, phuong_an)

        return None





def tim_cau_tu_a_den_b_tron_cau(doc, from_a, to_b):

    section = doc.Sections[0]

    pattern_cau = '^Câu [0-9]{1,}[.:]'

    pattern_chum_start = '<SCHUM@>'

    pattern_chum_end = '<ECHUM@>'

    inside_chum = False

    indices_cau = []

    for j in range(from_a, to_b):

        child = section.Body.ChildObjects[j]

        if not isinstance(child, Paragraph):

            continue

        text = child.Text.strip()

        if re.search(pattern_chum_start, text, re.IGNORECASE):

            indices_cau.append(j)

            inside_chum = True

            continue

        if re.search(pattern_chum_end, text, re.IGNORECASE):

            inside_chum = False

            continue

        if inside_chum:

            continue

        if not re.match(pattern_cau, text):

            continue

        indices_cau.append(j)

    indices_cau.append(to_b)

    return indices_cau





def tron_cau_trong_chum(doc, from_a, to_b):

    (Schum, Echum) = tim_chum_tu_a_den_b(doc, from_a, to_b)

    if len(Schum) != len(Echum):

        messagebox.showinfo('Thông báo', 'Số kí hiệu <SCHUM@> và <ECHUM@> không bằng nhau, hãy chỉnh sửa phù hợp')

        return None

    if len(Schum) != 0:

        for start, end in zip(Schum, Echum):

            indices_cau = tim_cau_tu_a_den_b_tron_cau(doc, start + 1, end)

            tron_group_tu_indices(doc, indices_cau)

        return None





def tron_cau_trong_vung(doc, from_a, to_b):

    indices_cau = tim_cau_tu_a_den_b_tron_cau(doc, from_a, to_b)

    tron_group_tu_indices(doc, indices_cau)





def tron_cau(doc, text_start, text_end):

    (indices_S, indices_E) = mspr_f.find_phan(doc, text_start, text_end)

    if len(indices_S) != len(indices_E):

        messagebox.showinfo('Thông báo', f'''Số kí hiệu {text_start} và {text_end} không bằng nhau, hãy chỉnh sửa phù hợp''')

        return None

    if len(indices_S) != 0:

        for start, end in zip(indices_S, indices_E):

            tron_cau_trong_chum(doc, start, end)

            tron_cau_trong_vung(doc, start, end)

        return None





def tron_cau_phan_I(doc):

    tron_cau(doc, 'S1@', 'E1@')





def tron_cau_phan_II(doc):

    tron_cau(doc, 'S2@', 'E2@')





def tron_cau_phan_III(doc):

    tron_cau(doc, 'S3@', 'E3@')





def tron_cau_phan_IV(doc):

    tron_cau(doc, 'S4@', 'E4@')





def dem_so_cau(doc):

    cau_p1 = mspr_f.tim_cau_trong_phan(doc, 'S1@', 'E1@')

    so_cau_p1 = 0

    if len(cau_p1) > 0:

        for i in range(len(cau_p1)):

            so_cau_p1 += len(cau_p1[i]) - 1

    cau_p2 = mspr_f.tim_cau_trong_phan(doc, 'S2@', 'E2@')

    so_cau_p2 = 0

    if len(cau_p2) > 0:

        for i in range(len(cau_p2)):

            so_cau_p2 += len(cau_p2[i]) - 1

    cau_p3 = mspr_f.tim_cau_trong_phan(doc, 'S3@', 'E3@')

    so_cau_p3 = 0

    if len(cau_p3) > 0:

        for i in range(len(cau_p3)):

            so_cau_p3 += len(cau_p3[i]) - 1

    return (so_cau_p1, so_cau_p2, so_cau_p3)





def lay_dap_an_vao_excel_TN(doc, ws, column = (1,)):

    dong = 0

    section = doc.Sections[0]

    labels = [

        'A.',

        'B.',

        'C.',

        'D.']

    labels2 = [

        'A',

        'B',

        'C',

        'D']

    all_indices_cau = mspr_f.tim_cau_trong_phan(doc, 'S1@', 'E1@')

    if len(all_indices_cau) > 0:

        for idx in range(len(all_indices_cau)):

            cau = all_indices_cau[idx]

            for k in range(len(cau) - 1):

                phuong_an = mspr_f.tim_phuong_an_trong_cau_phan_I(doc, cau[k], cau[k + 1])

                encounter_count = 0

                for h in range(phuong_an[0], phuong_an[-1]):

                    child = section.Body.ChildObjects[h]

                    if not isinstance(child, Paragraph):

                        continue

                    if not child.Text.startswith(tuple(labels)):

                        continue

                    first_child = child.ChildObjects[0]

                    current_label = labels2[encounter_count % len(labels2)]

                    encounter_count = encounter_count + 1

                    if not first_child.CharacterFormat.UnderlineStyle == UnderlineStyle.Single and chspr.is_red_color_spr(first_child.CharacterFormat.TextColor):

                        continue

                    answer = current_label

                    ws.cell(row = dong + 4, column = column, value = answer)

                    range(phuong_an[0], phuong_an[-1])

            dong = dong + 1

        continue

        return None





def lay_dap_an_vao_excel_DS_TNmaker(doc, ws, column = (1,)):

    (so_cau_p1, so_cau_p2, so_cau_p3) = dem_so_cau(doc)

    dong = so_cau_p1

    section = doc.Sections[0]

    labels = [

        'a)',

        'b)',

        'c)',

        'd)']

    all_indices_cau = mspr_f.tim_cau_trong_phan(doc, 'S2@', 'E2@')

# WARNING: Decompyle incomplete





def lay_dap_an_vao_excel_DS_Testpro(doc, ws, column = (1,)):

    (so_cau_p1, so_cau_p2, so_cau_p3) = dem_so_cau(doc)

    dong = so_cau_p1

    section = doc.Sections[0]

    labels = [

        'a)',

        'b)',

        'c)',

        'd)']

    all_indices_cau = mspr_f.tim_cau_trong_phan(doc, 'S2@', 'E2@')

# WARNING: Decompyle incomplete





def lay_dap_an_vao_excel_TLN(doc, ws, column = (1,)):

    (so_cau_p1, so_cau_p2, so_cau_p3) = dem_so_cau(doc)

    dong = so_cau_p1 + so_cau_p2

    section = doc.Sections[0]

    all_indices_cau = mspr_f.tim_cau_trong_phan(doc, 'S3@', 'E3@')

    if len(all_indices_cau) > 0:

        for idx in range(len(all_indices_cau)):

            cau = all_indices_cau[idx]

            answer = 'NO'

            for k in range(len(cau) - 1):

                for h in range(cau[k], cau[k + 1]):

                    child = section.Body.ChildObjects[h]

                    if not isinstance(child, Paragraph):

                        continue

                    if not child.Text.startswith('ĐS:'):

                        continue

                    answer = child.Text.replace('ĐS:', '').strip()

                ws.cell(row = dong + 4, column = column, value = answer)

                dong = dong + 1

        return None





def tron_and_xuat_de(check1_1_var, check1_2_var, check2_1_var, check2_2_var, check3_1_var, check4_1_var, check5_1_var, check6_1_var, check7_1_var, entry_input, entry_lantron, entry_newname):

    pass

# WARNING: Decompyle incomplete



