from spire.doc import Section, Paragraph, Table, TextRange, Shape, FileFormat, HorizontalAlignment, Regex

from spire.doc import BookmarkStart, BookmarkEnd

from spire.doc import UnderlineStyle

from spire.doc import Document as SpireDocument

from spire.doc import Color

import os

import re

import random

import tkinter as tk

from tkinter import ttk, Label, filedialog, Entry, Button, Frame, Listbox, Scrollbar, messagebox

from tkinter import simpledialog

from openpyxl import Workbook

from openpyxl.styles import Font

from openpyxl.styles import Alignment

from openpyxl.styles import Border, Side

from openpyxl.styles import PatternFill

from openpyxl.styles import GradientFill

from openpyxl.utils import get_column_letter

from openpyxl import load_workbook

from formatting import chuan_hoa_spr as chspr

from tools import tool_by_docx as docxtool

from mixing import mix_spr_func as mspr_f



def delete_only_bookmark_tags(doc):

    excluded_names = {

        'MDH',

        'num_page'}

    names_to_keep = set()

    bookmark_ends_to_keep = set()

    for i in range(doc.Bookmarks.Count):

        bookmark = doc.Bookmarks[i]

        if not bookmark.Name in excluded_names:

            continue

        names_to_keep.add(bookmark.Name)

        bookmark_ends_to_keep.add(bookmark.BookmarkEnd)

    for i in range(doc.Sections.Count):

        section = doc.Sections[i]

        for j in range(section.Paragraphs.Count):

            para = section.Paragraphs[j]

            index = para.ChildObjects.Count - 1

            if not index >= 0:

                continue

            child = para.ChildObjects[index]

            if isinstance(child, BookmarkStart) or child.Name not in names_to_keep:

                para.ChildObjects.RemoveAt(index)

            elif isinstance(child, BookmarkEnd) and child not in bookmark_ends_to_keep:

                para.ChildObjects.RemoveAt(index)

            index -= 1

            if index >= 0:

                continue

    continue





def delete_bookmarks(doc):

    excluded_names = {

        'MDH',

        'num_page'}

    names_to_keep = set()

    bookmark_ends_to_keep = set()

    for i in range(doc.Bookmarks.Count):

        bookmark = doc.Bookmarks[i]

        if not bookmark.Name in excluded_names:

            continue

        names_to_keep.add(bookmark.Name)

        bookmark_ends_to_keep.add(bookmark.BookmarkEnd)

    for i in range(doc.Sections.Count):

        section = doc.Sections[i]

        for j in range(section.Paragraphs.Count):

            para = section.Paragraphs[j]

            index = para.ChildObjects.Count - 1

            if not index >= 0:

                continue

            child = para.ChildObjects[index]

            if isinstance(child, BookmarkStart) or child.Name not in names_to_keep:

                para.ChildObjects.RemoveAt(index)

            elif isinstance(child, BookmarkEnd) and child not in bookmark_ends_to_keep:

                para.ChildObjects.RemoveAt(index)

            index -= 1

            if index >= 0:

                continue

    continue





def delete_run0_empty(doc, indices):

    section = doc.Sections[0]

    for p in indices:

        para = section.Body.Paragraphs[p]

        if not para != None:

            continue

        if not para.ChildObjects.Count > 0:

            continue

        first = para.ChildObjects[0]

        if not isinstance(first, TextRange):

            continue

        if not first.Text.strip() == '':

            continue

        para.ChildObjects.Remove(first)





def tim_chi_so_cau_dap_an_phan_chum(doc):

    indices = []

    section = doc.Sections[0]

    pattern1 = '^Câu [0-9]{1,}[.:]'

    pattern2 = '^Question [0-9]{1,}[.:]'

    labels = [

        'A.',

        'B.',

        'C.',

        'D.',

        'a)',

        'b)',

        'c)',

        'd)',

        'S1@',

        'E1@',

        'S2@',

        'E2@',

        'S3@',

        'E3@',

        'S4@',

        'E4@',

        '<S',

        '<E',

        'ĐS:']

    for p in range(section.Body.Paragraphs.Count):

        para = section.Body.Paragraphs[p]

        text = para.Text.strip()

        if not re.match(pattern1, text) and re.match(pattern2, text) and text.startswith(tuple(labels)):

            continue

        indices.append(p)

    return indices





def xoa_run_0_all(doc):

    indices = tim_chi_so_cau_dap_an_phan_chum(doc)

    delete_run0_empty(doc, indices)





def replace_mu_11(doc):

    for i in range(doc.Sections.Count):

        section = doc.Sections[i]

        for p in range(section.Body.Paragraphs.Count):

            para = section.Body.Paragraphs[p]

            if not isinstance(para, Paragraph):

                continue

            para.Replace('\x0b', '\n', False, False)





def check_mu_11(doc):

    section = doc.Sections[0]

    for p in range(section.Body.Paragraphs.Count):

        para = section.Body.Paragraphs[p]

        if not isinstance(para, Paragraph):

            continue

        if not '\x0b' in para.Text:

            continue

        range(section.Body.Paragraphs.Count)

        return True

    return False





def find_nhom_all(doc, text_start, text_end):

    (indices_S, indices_E) = mspr_f.find_phan(doc, text_start, text_end)

    return (indices_S, indices_E)





def Tao_group_tu_indices(doc, indices):

    groups = []

    items = []

    section = doc.Sections[0]

    for idx in range(indices[0], indices[-1]):

        item = section.Body.ChildObjects[idx]

        items.append(item)

    for k in range(len(indices) - 1):

        start = indices[k] - indices[0]

        end = indices[k + 1] - indices[0]

        group = items[start:end]

        groups.append(group)

    return groups





def tron_group_tu_indices(doc, indices):

    if len(indices) > 2:

        groups = Tao_group_tu_indices(doc, indices)

        random.shuffle(groups)

        body = doc.Sections[0].Body

        start = indices[0]

        end = indices[-1]

        for i in range(end - 1, start - 1, -1):

            body.ChildObjects.RemoveAt(i)

        insert_index = start

        for group in groups:

            for item in group:

                body.ChildObjects.Insert(insert_index, item.Clone())

                insert_index += 1

        return None





def tim_nhom_tu_a_den_b_tron_nhom(doc, from_a, to_b):

    section = doc.Sections[0]

    text_start = '<SNHOM@>'

    indices_S = []

    for j in range(from_a, to_b):

        child = section.Body.ChildObjects[j]

        if not isinstance(child, Paragraph):

            continue

        text = child.Text.strip()

        if not text.startswith(text_start):

            continue

        indices_S.append(j)

    indices_S.append(to_b)

    return indices_S





def tim_cau_tu_a_den_b(doc, from_a, to_b):

    section = doc.Sections[0]

    pattern1 = '^Question [0-9]{1,}[.:]'

    pattern2 = '^Câu [0-9]{1,}[.:]'

    indices_cau = []

    for j in range(from_a, to_b):

        child = section.Body.ChildObjects[j]

        if not isinstance(child, Paragraph):

            continue

        if not re.match(pattern1, child.Text) and re.match(pattern2, child.Text):

            continue

        indices_cau.append(j)

    indices_cau.append(to_b)

    return indices_cau





def tron_phuong_an_trong_vung(doc, snhom, enhom):

    (indices_S, indices_E) = find_nhom_all(doc, snhom, enhom)

    if len(indices_S) != len(indices_E):

        messagebox.showinfo('Thông báo', f'''Số kí hiệu {snhom} và {enhom} không bằng nhau, hãy chỉnh sửa phù hợp''')

        return None

    if len(indices_S) != 0:

        for start, end in zip(indices_S, indices_E):

            indices_cau = tim_cau_tu_a_den_b(doc, start, end)

            for k in range(len(indices_cau) - 1):

                phuong_an = mspr_f.tim_phuong_an_trong_cau_phan_I(doc, indices_cau[k], indices_cau[k + 1])

                tron_group_tu_indices(doc, phuong_an)

        return None





def tron_phuong_an(doc):

    tron_phuong_an_trong_vung(doc, '<SNHOM@><D', '<ENHOM@><D')

    tron_phuong_an_trong_vung(doc, '<SNHOM@><CD>', '<ENHOM@><CD>')





def tron_cau_trong_vung(doc, text_start, text_end):

    (indices_S, indices_E) = find_nhom_all(doc, text_start, text_end)

    if len(indices_S) != len(indices_E):

        messagebox.showinfo('Thông báo', f'''Số kí hiệu {text_start} và {text_end} không bằng nhau, hãy chỉnh sửa phù hợp''')

        return None

    if len(indices_S) != 0:

        for start, end in zip(indices_S, indices_E):

            indices_cau = tim_cau_tu_a_den_b(doc, start, end)

            tron_group_tu_indices(doc, indices_cau)

        return None





def tron_cau(doc):

    tron_cau_trong_vung(doc, '<SNHOM@><C', '<ENHOM@><C')

    tron_cau_trong_vung(doc, '<SNHOM@><DC>', '<ENHOM@><DC>')





def tron_nhom(doc):

    (indices_S, indices_E) = mspr_f.find_phan_EN(doc)

    if len(indices_S) != len(indices_E):

        messagebox.showinfo('Thông báo', 'Số kí hiệu S1@ và E2@ không bằng nhau, hãy chỉnh sửa phù hợp')

        return None

    if len(indices_S) != 0:

        for start, end in zip(indices_S, indices_E):

            indices_nhom = tim_nhom_tu_a_den_b_tron_nhom(doc, start, end)

            tron_group_tu_indices(doc, indices_nhom)

        return None





def tim_cau_trong_phan(doc):

    all_indices_cau = mspr_f.tim_cau_trong_phan_EN(doc)

    return all_indices_cau





def dem_so_cau(doc):

    cau_p1 = tim_cau_trong_phan(doc)

    so_cau_p1 = 0

    if len(cau_p1) > 0:

        for i in range(len(cau_p1)):

            so_cau_p1 += len(cau_p1[i]) - 1

    return so_cau_p1





def lay_dap_an_vao_excel_TN(doc, ws, column = (1,)):

    dong = 0

    section = doc.Sections[0]

    labels = [

        'A.',

        'B.',

        'C.',

        'D.']

    labels2 = [

        'A',

        'B',

        'C',

        'D']

    all_indices_cau = tim_cau_trong_phan(doc)

    if len(all_indices_cau) > 0:

        for idx in range(len(all_indices_cau)):

            cau = all_indices_cau[idx]

            for k in range(len(cau) - 1):

                phuong_an = mspr_f.tim_phuong_an_trong_cau_phan_I(doc, cau[k], cau[k + 1])

                encounter_count = 0

                for h in range(phuong_an[0], phuong_an[-1]):

                    child = section.Body.ChildObjects[h]

                    if not isinstance(child, Paragraph):

                        continue

                    if not child.Text.startswith(tuple(labels)):

                        continue

                    first_child = child.ChildObjects[0]

                    current_label = labels2[encounter_count % len(labels2)]

                    encounter_count = encounter_count + 1

                    if not first_child.CharacterFormat.UnderlineStyle == UnderlineStyle.Single and chspr.is_red_color_spr(first_child.CharacterFormat.TextColor):

                        continue

                    answer = current_label

                    ws.cell(row = dong + 4, column = column, value = answer)

                    range(phuong_an[0], phuong_an[-1])

            dong = dong + 1

        continue

        return None





def tron_and_xuat_de_NN(entry_input, entry_lantron, entry_newname, check6_1_var, check7_1_var, check8_1_var):

    newname_first = entry_lantron.get().strip()

    input_file_get = entry_input.get()

    input_file = os.path.normpath(input_file_get)

    if not os.path.exists(input_file):

        messagebox.showinfo('Thông báo', 'Hãy tắt hết các file word, và chọn lại file gốc để trộn.')

        return None

    page_kind = 'A4'

    bo_loi_giai = 'YES'

    langue = 'english'

    if check6_1_var.get():

        page_kind = '2C'

    if check7_1_var.get():

        bo_loi_giai = 'NO'

    if check8_1_var.get():

        langue = 'japan'

    new_names = entry_newname.get().strip().split(',')

    if not new_names:

        messagebox.showwarning('Warning', 'Please enter new name(s).')

        return None

    i = 0

# WARNING: Decompyle incomplete



