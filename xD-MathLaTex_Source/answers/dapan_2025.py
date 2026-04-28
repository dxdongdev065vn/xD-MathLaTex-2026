import os

import datetime

import win32com.client as win32com

from win32com.client import client as win32

import win32api

import re

import sys

import shutil

import pythoncom

import tempfile

import wmi

import docx

import random

from docx import Document

from docx.shared import Pt

from docx.enum.text import WD_PARAGRAPH_ALIGNMENT

from docx.shared import RGBColor

from openpyxl import Workbook

from openpyxl.styles import Font

from openpyxl.styles import Alignment

from openpyxl.styles import Border, Side

from openpyxl.styles import PatternFill

from openpyxl.styles import GradientFill

from openpyxl.utils import get_column_letter

from openpyxl import load_workbook

import tkinter as tk

from tkinter import ttk, Label, filedialog, Entry, Button, Frame, Listbox, Scrollbar, messagebox

from tkinter import simpledialog

import webbrowser

from core import functions as vbf



def Convert_da_to_table_p1(word, socot):

    doc = word.ActiveDocument

    selection = word.Selection

    selection.HomeKey(Unit = 6)

    find = selection.Find

    find.ClearFormatting()

    find.Text = '(1S\\@)(*)(1E\\@)'

    find.Replacement.Text = ''

    find.Forward = True

    find.Wrap = 1

    find.Format = False

    find.MatchCase = False

    find.MatchWholeWord = False

    find.MatchAllWordForms = False

    find.MatchSoundsLike = False

    find.MatchWildcards = True

    if find.Execute():

        selection.Find.Execute()

        myrange = selection.Range

        myrange.SetRange(Start = myrange.Start + 4, End = myrange.End - 3)

        myrange.Select()

        if myrange.Text.strip():

            num_rows = myrange.Paragraphs.Count

            myrange.Select()

            selection.ConvertToTable(Separator = '_', NumColumns = socot, NumRows = 1, AutoFitBehavior = 1)

            table = selection.Tables(1)

            vbf.autofit_one_table(word, table)

            table.Borders.Enable = True

            table.Range.Font.Size = 12

            return None

        return None

    return None

# WARNING: Decompyle incomplete





def Convert_da_to_table_p2(word, socot):

    doc = word.ActiveDocument

    selection = word.Selection

    selection.HomeKey(Unit = 6)

    find = selection.Find

    find.ClearFormatting()

    find.Text = '(2S\\@)(*)(2E\\@)'

    find.Replacement.Text = ''

    find.Forward = True

    find.Wrap = 1

    find.Format = False

    find.MatchCase = False

    find.MatchWholeWord = False

    find.MatchAllWordForms = False

    find.MatchSoundsLike = False

    find.MatchWildcards = True

    if find.Execute():

        selection.Find.Execute()

        myrange = selection.Range

        myrange.SetRange(Start = myrange.Start + 4, End = myrange.End - 3)

        myrange.Select()

        if myrange.Text.strip():

            num_rows = myrange.Paragraphs.Count

            myrange.Select()

            selection.ConvertToTable(Separator = '_', NumColumns = socot, NumRows = 1, AutoFitBehavior = 1)

            table = selection.Tables(1)

            vbf.autofit_one_table(word, table)

            table.Borders.Enable = True

            table.Range.Font.Size = 12

            return None

        return None

    return None

# WARNING: Decompyle incomplete





def Convert_da_to_table_p3(word, socot):

    doc = word.ActiveDocument

    selection = word.Selection

    selection.HomeKey(Unit = 6)

    find = selection.Find

    find.ClearFormatting()

    find.Text = '(3S\\@)(*)(3E\\@)'

    find.Replacement.Text = ''

    find.Forward = True

    find.Wrap = 1

    find.Format = False

    find.MatchCase = False

    find.MatchWholeWord = False

    find.MatchAllWordForms = False

    find.MatchSoundsLike = False

    find.MatchWildcards = True

    if find.Execute():

        selection.Find.Execute()

        myrange = selection.Range

        myrange.SetRange(Start = myrange.Start + 4, End = myrange.End - 3)

        myrange.Select()

        if myrange.Text.strip():

            num_rows = myrange.Paragraphs.Count

            myrange.Select()

            selection.ConvertToTable(Separator = '_', NumColumns = socot, NumRows = 1, AutoFitBehavior = 1)

            table = selection.Tables(1)

            vbf.autofit_one_table(word, table)

            table.Borders.Enable = True

            table.Range.Font.Size = 12

            return None

        return None

    return None

# WARNING: Decompyle incomplete





def DapAn2025_En(word):

    doc = word.ActiveDocument

    vbf.Convert_Auto_To_Text(word)

    txtDA = ''

    txt_ds = ''

    txt_tln = ''

    sc_a = 0

    sc_ds = 0

    sc_tln = 0

    myrange = doc.Range()

    myrange_find = myrange.Duplicate

    find = myrange_find.Find

    find.ClearFormatting()

    find.Text = '(Question [0-9]{1,2}[.:])'

    find.MatchWildcards = True

# WARNING: Decompyle incomplete





def DapAn2025(word):

    doc = word.ActiveDocument

    vbf.Convert_Auto_To_Text(word)

    txtDA = ''

    txt_ds = ''

    txt_tln = ''

    sc_a = 0

    sc_ds = 0

    sc_tln = 0

    myrange = doc.Range()

    myrange_find = myrange.Duplicate

    find = myrange_find.Find

    find.ClearFormatting()

    find.Text = '(Câu [0-9]{1,2}[.:])'

    find.MatchWildcards = True

    if find.Execute():

        vbf.add_blank_line_at_Home(word)

        vbf.add_blank_line_after_table(word)

        vbf.them_cau_acong_cuoi(word)

        for i in range(1, 500):

            d_a = ''

            da_dsa = ''

            da_dsb = ''

            da_dsc = ''

            da_dsd = ''

            myrange_cau = myrange.Duplicate

            find = myrange_cau.Find

            find.ClearFormatting()

            find.Text = '(Câu [0-9]{1,2}[.:])(*)(Câu [0-9]{1,2}[.:])'

            find.MatchWildcards = True

            if find.Execute():

                myrange = doc.Range(myrange_cau.End - 9, doc.Range().End)

                myrange_find = myrange_cau.Duplicate

                myrange_find.Select()

                find = myrange_find.Find

                find.ClearFormatting()

                find.Text = '([^13^9])([Dd])([.\\)])'

                find.MatchWildcards = True

                if find.Execute():

                    myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                    text = myrange_find.Text.strip()

                    if text == 'D':

                        myrange_find = myrange_cau.Duplicate

                        find = myrange_find.Find

                        find.ClearFormatting()

                        find.Text = '([^13^9])(A.)'

                        find.MatchWildcards = True

                        if find.Execute():

                            myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                            if myrange_find.Font.Underline == True or myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                                text = myrange_find.Text.strip()

                                d_a += text

                        myrange_find = myrange_cau.Duplicate

                        find = myrange_find.Find

                        find.ClearFormatting()

                        find.Text = '([^13^9])(B.)'

                        find.MatchWildcards = True

                        if find.Execute():

                            myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                            if myrange_find.Font.Underline == True or myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                                text = myrange_find.Text.strip()

                                d_a += text

                        myrange_find = myrange_cau.Duplicate

                        find = myrange_find.Find

                        find.ClearFormatting()

                        find.Text = '([^13^9])(C.)'

                        find.MatchWildcards = True

                        if find.Execute():

                            myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                            if myrange_find.Font.Underline == True or myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                                text = myrange_find.Text.strip()

                                d_a += text

                        myrange_find = myrange_cau.Duplicate

                        find = myrange_find.Find

                        find.ClearFormatting()

                        find.Text = '([^13^9])(D.)'

                        find.MatchWildcards = True

                        if find.Execute():

                            myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                            if myrange_find.Font.Underline == True or myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                                text = myrange_find.Text.strip()

                                d_a += text

                        if d_a == '':

                            myrange_find = myrange_cau.Duplicate

                            find = myrange_find.Find

                            find.ClearFormatting()

                            find.Text = '(Chọn )([A-D])'

                            find.MatchWildcards = True

                            if find.Execute():

                                myrange_find.SetRange(Start = myrange_find.Start + 5, End = myrange_find.End)

                                text = myrange_find.Text.strip()

                                d_a += text

                        sc_a += 1

                        txtDA = txtDA + f'''{sc_a}.''' + d_a + '_'

                        continue

                    dap_an_DS_cau = ''

                    keywords = [

                        '[^13^9]ĐS:',

                        '[^13^9]Đáp số:',

                        '[^13^9]ĐA:',

                        '[^13^9]Đáp án:',

                        '[^13^9]TL:',

                        '[^13^9]Trả lời:',

                        '[^13^9]ĐS[ ]@:',

                        '[^13^9]Đáp số[ ]@:',

                        '[^13^9]ĐA[ ]@:',

                        '[^13^9]Đáp án[ ]@:',

                        '[^13^9]TL[ ]@:',

                        '[^13^9]Trả lời[ ]@:']

                    for keyword in keywords:

                        myrange_find2 = myrange_cau.Duplicate

                        find = myrange_find2.Find

                        find.ClearFormatting()

                        find.Text = keyword

                        find.MatchWildcards = True

                        find.MatchCase = False

                        if not find.Execute():

                            continue

                        sc_ds += 1

                        paragraph = myrange_find2.Paragraphs(myrange_find2.Paragraphs.Count)

                        myrange_find2.Start = myrange_find2.End

                        myrange_find2.End = paragraph.Range.End - 1

                        dap_an_DS_cau = myrange_find2.Text.strip()

                        txt_ds += f'''{sc_ds}:''' + dap_an_DS_cau + '_'

                        keywords

                    if not dap_an_DS_cau == '':

                        continue

                    da_dsa = 'S'

                    myrange_find = myrange_cau.Duplicate

                    find = myrange_find.Find

                    find.ClearFormatting()

                    find.Text = '([^13^9])(a[\\)])'

                    find.MatchWildcards = True

                    if find.Execute():

                        myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                        if myrange_find.Font.Underline == True or myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                            da_dsa = 'Đ'

                        else:

                            myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_cau.End)

                            find = myrange_find.Find

                            find.ClearFormatting()

                            find.Text = '([^13^9])(a[\\)])'

                            find.MatchWildcards = True

                            if find.Execute() and myrange_find.Paragraphs.Count >= 2:

                                para_text = myrange_find.Paragraphs(2).Range.Text

                                if 'đúng' in para_text.lower():

                                    da_dsa = 'Đ'

                    da_dsb = 'S'

                    myrange_find = myrange_cau.Duplicate

                    find = myrange_find.Find

                    find.ClearFormatting()

                    find.Text = '([^13^9])(b[\\)])'

                    find.MatchWildcards = True

                    if find.Execute():

                        myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                        if myrange_find.Font.Underline == True or myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                            da_dsb = 'Đ'

                        else:

                            myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_cau.End)

                            find = myrange_find.Find

                            find.ClearFormatting()

                            find.Text = '([^13^9])(b[\\)])'

                            find.MatchWildcards = True

                            if find.Execute() and myrange_find.Paragraphs.Count >= 2:

                                para_text = myrange_find.Paragraphs(2).Range.Text

                                if 'đúng' in para_text.lower():

                                    da_dsa = 'Đ'

                    da_dsc = 'S'

                    myrange_find = myrange_cau.Duplicate

                    find = myrange_find.Find

                    find.ClearFormatting()

                    find.Text = '([^13^9])(c[\\)])'

                    find.MatchWildcards = True

                    if find.Execute():

                        myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                        if myrange_find.Font.Underline == True or myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                            da_dsc = 'Đ'

                        else:

                            myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_cau.End)

                            find = myrange_find.Find

                            find.ClearFormatting()

                            find.Text = '([^13^9])(c[\\)])'

                            find.MatchWildcards = True

                            if find.Execute() and myrange_find.Paragraphs.Count >= 2:

                                para_text = myrange_find.Paragraphs(2).Range.Text

                                if 'đúng' in para_text.lower():

                                    da_dsa = 'Đ'

                    da_dsd = 'S'

                    myrange_find = myrange_cau.Duplicate

                    find = myrange_find.Find

                    find.ClearFormatting()

                    find.Text = '([^13^9])(d[\\)])'

                    find.MatchWildcards = True

                    if find.Execute():

                        myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                        if myrange_find.Font.Underline == True or myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                            da_dsd = 'Đ'

                        else:

                            myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_cau.End)

                            find = myrange_find.Find

                            find.ClearFormatting()

                            find.Text = '([^13^9])(d[\\)])'

                            find.MatchWildcards = True

                            if find.Execute() and myrange_find.Paragraphs.Count >= 2:

                                para_text = myrange_find.Paragraphs(2).Range.Text

                                if 'đúng' in para_text.lower():

                                    da_dsa = 'Đ'

                    sc_ds += 1

                    txt_ds = txt_ds + f'''{sc_ds}:''' + da_dsa + da_dsb + da_dsc + da_dsd + '_'

                    continue

                myrange = doc.Range(myrange_cau.End - 8, doc.Range().End)

                keywords = [

                    '[^13^9]ĐS:',

                    '[^13^9]Đáp số:',

                    '[^13^9]ĐA:',

                    '[^13^9]Đáp án:',

                    '[^13^9]TL:',

                    '[^13^9]Trả lời:',

                    '[^13^9]ĐS[ ]@:',

                    '[^13^9]Đáp số[ ]@:',

                    '[^13^9]ĐA[ ]@:',

                    '[^13^9]Đáp án[ ]@:',

                    '[^13^9]TL[ ]@:',

                    '[^13^9]Trả lời[ ]@:']

                for keyword in keywords:

                    myrange_find2 = myrange_cau.Duplicate

                    find = myrange_find2.Find

                    find.ClearFormatting()

                    find.Text = keyword

                    find.MatchWildcards = True

                    find.MatchCase = False

                    if not find.Execute():

                        continue

                    sc_tln += 1

                    paragraph = myrange_find2.Paragraphs(myrange_find2.Paragraphs.Count)

                    myrange_find2.Start = myrange_find2.End

                    myrange_find2.End = paragraph.Range.End - 1

                    txt_cau = myrange_find2.Text.strip()

                    contains_mathtype = (lambda .0: pass# WARNING: Decompyle incomplete

)(myrange_find2.InlineShapes())

                    contains_image = (lambda .0: pass# WARNING: Decompyle incomplete

)(myrange_find2.InlineShapes())

                    if contains_mathtype:

                        txt_cau = 'Mathtype'

                    if contains_image:

                        txt_cau = 'Picture'

                    txt_tln += f'''{sc_tln}:''' + txt_cau + '_'

                    any

                continue

        txtDA = txtDA.strip('_')

        txt_ds = txt_ds.strip('_')

        txt_tln = txt_tln.strip('_')

        paragraphs = doc.Paragraphs

        pattern = re.compile('^phần', re.IGNORECASE)

        for i in range(1, 10):

            para = paragraphs(i)

            text = para.Range.Text.strip()

            if pattern.match(text):

                rng = para.Range

                rng.Collapse(Direction = 1)

                rng.Select()

                range(1, 10)

            else:

                word.Selection.HomeKey(Unit = 6)

        word.Selection.InsertBefore('\n')

        vbf.Page_Left_Right00(word)

        word.Selection.ParagraphFormat.Alignment = 1

        word.Selection.Font.Bold = True

        word.Selection.Font.Italic = False

        word.Selection.Font.Name = 'Times New Roman'

        word.Selection.Font.Size = 12

        if txtDA != '':

            word.Selection.Font.Color = win32api.RGB(0, 32, 96)

            word.Selection.TypeText('BẢNG ĐÁP ÁN TN')

            word.Selection.TypeParagraph()

            word.Selection.Font.Color = win32api.RGB(0, 128, 0)

            word.Selection.TypeText('1S@:')

            word.Selection.TypeText(txtDA)

            word.Selection.TypeText('1E@:')

            word.Selection.TypeParagraph()

        if txt_ds != '':

            word.Selection.Font.Color = win32api.RGB(0, 32, 96)

            word.Selection.TypeText('BẢNG ĐÁP ÁN ĐÚNG SAI')

            word.Selection.TypeParagraph()

            word.Selection.Font.Color = win32api.RGB(0, 128, 0)

            word.Selection.TypeText('2S@:')

            word.Selection.TypeText(txt_ds)

            word.Selection.TypeText('2E@:')

            word.Selection.TypeParagraph()

        if txt_tln != '':

            word.Selection.Font.Color = win32api.RGB(0, 32, 96)

            word.Selection.TypeText('BẢNG ĐÁP ÁN TRẢ LỜI NGẮN')

            word.Selection.TypeParagraph()

            word.Selection.Font.Color = win32api.RGB(0, 128, 0)

            word.Selection.TypeText('3S@:')

            word.Selection.TypeText(txt_tln)

            word.Selection.TypeText('3E@:')

            word.Selection.TypeParagraph()

        if sc_a <= 12:

            socot = sc_a

        else:

            socot = 10

        if sc_ds <= 6:

            socot_ds = sc_ds

        else:

            socot_ds = 6

        if sc_tln <= 8:

            socot_tln = sc_tln

        else:

            socot_tln = 8

        Convert_da_to_table_p1(word, socot)

        Convert_da_to_table_p2(word, socot_ds)

        Convert_da_to_table_p3(word, socot_tln)

        vbf.xoa_dong_acong(word)

        vbf.xoa_dong_trang(word)

        word.Selection.HomeKey(Unit = 6)

    DapAn2025_En(word)

    return None

# WARNING: Decompyle incomplete





def DapAn2025_TNmaker(word):

    doc = word.ActiveDocument

    vbf.Convert_Auto_To_Text(word)

    txtDA = ''

    txt_ds = ''

    txt_tln = ''

    sc_a = 0

    sc_ds = 0

    sc_tln = 0

    myrange = doc.Range()

    myrange_find = myrange.Duplicate

    find = myrange_find.Find

    find.ClearFormatting()

    find.Text = '(Câu [0-9]{1,2}[.:])'

    find.MatchWildcards = True

    if find.Execute():

        vbf.add_blank_line_at_Home(word)

        vbf.add_blank_line_after_table(word)

        vbf.them_cau_acong_cuoi(word)

        for i in range(1, 500):

            d_a = ''

            da_dsa = ''

            da_dsb = ''

            da_dsc = ''

            da_dsd = ''

            myrange_cau = myrange.Duplicate

            find = myrange_cau.Find

            find.ClearFormatting()

            find.Text = '(Câu [0-9]{1,2}[.:])(*)(Câu [0-9]{1,2}[.:])'

            find.MatchWildcards = True

            if find.Execute():

                myrange = doc.Range(myrange_cau.End - 9, doc.Range().End)

                myrange_find = myrange_cau.Duplicate

                myrange_find.Select()

                find = myrange_find.Find

                find.ClearFormatting()

                find.Text = '([^13^9])([Dd])([.\\)])'

                find.MatchWildcards = True

                if find.Execute():

                    myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                    text = myrange_find.Text.strip()

                    if text == 'D':

                        myrange_find = myrange_cau.Duplicate

                        find = myrange_find.Find

                        find.ClearFormatting()

                        find.Text = '([^13^9])(A.)'

                        find.MatchWildcards = True

                        if find.Execute():

                            myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                            if myrange_find.Font.Underline == True or myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                                text = myrange_find.Text.strip()

                                d_a += text

                        myrange_find = myrange_cau.Duplicate

                        find = myrange_find.Find

                        find.ClearFormatting()

                        find.Text = '([^13^9])(B.)'

                        find.MatchWildcards = True

                        if find.Execute():

                            myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                            if myrange_find.Font.Underline == True or myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                                text = myrange_find.Text.strip()

                                d_a += text

                        myrange_find = myrange_cau.Duplicate

                        find = myrange_find.Find

                        find.ClearFormatting()

                        find.Text = '([^13^9])(C.)'

                        find.MatchWildcards = True

                        if find.Execute():

                            myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                            if myrange_find.Font.Underline == True or myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                                text = myrange_find.Text.strip()

                                d_a += text

                        myrange_find = myrange_cau.Duplicate

                        find = myrange_find.Find

                        find.ClearFormatting()

                        find.Text = '([^13^9])(D.)'

                        find.MatchWildcards = True

                        if find.Execute():

                            myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                            if myrange_find.Font.Underline == True or myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                                text = myrange_find.Text.strip()

                                d_a += text

                        if d_a == '':

                            myrange_find = myrange_cau.Duplicate

                            find = myrange_find.Find

                            find.ClearFormatting()

                            find.Text = '(Chọn )([A-D])'

                            find.MatchWildcards = True

                            if find.Execute():

                                myrange_find.SetRange(Start = myrange_find.Start + 5, End = myrange_find.End)

                                text = myrange_find.Text.strip()

                                d_a += text

                        sc_a += 1

                        txtDA = txtDA + d_a + '_'

                        continue

                    dap_an_DS_cau = ''

                    keywords = [

                        '[^13^9]ĐS:',

                        '[^13^9]Đáp số:',

                        '[^13^9]ĐA:',

                        '[^13^9]Đáp án:',

                        '[^13^9]TL:',

                        '[^13^9]Trả lời:']

                    for keyword in keywords:

                        myrange_find2 = myrange_cau.Duplicate

                        find = myrange_find2.Find

                        find.ClearFormatting()

                        find.Text = keyword

                        find.MatchWildcards = True

                        find.MatchCase = False

                        if not find.Execute():

                            continue

                        sc_ds += 1

                        paragraph = myrange_find2.Paragraphs(myrange_find2.Paragraphs.Count)

                        myrange_find2.Start = myrange_find2.End

                        myrange_find2.End = paragraph.Range.End - 1

                        dap_an_DS_cau = myrange_find2.Text.strip()

                        txt_ds += dap_an_DS_cau + '_'

                        keywords

                    if not dap_an_DS_cau == '':

                        continue

                    myrange_find = myrange_cau.Duplicate

                    find = myrange_find.Find

                    find.ClearFormatting()

                    find.Text = '([^13^9])(a[\\)])'

                    find.MatchWildcards = True

                    if find.Execute():

                        myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                        if myrange_find.Font.Underline == True or myrange.Font.Color == win32api.RGB(255, 0, 0):

                            da_dsa = 'Đ'

                        else:

                            da_dsa = 'S'

                    myrange_find = myrange_cau.Duplicate

                    find = myrange_find.Find

                    find.ClearFormatting()

                    find.Text = '([^13^9])(b[\\)])'

                    find.MatchWildcards = True

                    if find.Execute():

                        myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                        if myrange_find.Font.Underline == True or myrange.Font.Color == win32api.RGB(255, 0, 0):

                            da_dsb = 'Đ'

                        else:

                            da_dsb = 'S'

                    myrange_find = myrange_cau.Duplicate

                    find = myrange_find.Find

                    find.ClearFormatting()

                    find.Text = '([^13^9])(c[\\)])'

                    find.MatchWildcards = True

                    if find.Execute():

                        myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                        if myrange_find.Font.Underline == True or myrange.Font.Color == win32api.RGB(255, 0, 0):

                            da_dsc = 'Đ'

                        else:

                            da_dsc = 'S'

                    myrange_find = myrange_cau.Duplicate

                    find = myrange_find.Find

                    find.ClearFormatting()

                    find.Text = '([^13^9])(d[\\)])'

                    find.MatchWildcards = True

                    if find.Execute():

                        myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                        if myrange_find.Font.Underline == True or myrange.Font.Color == win32api.RGB(255, 0, 0):

                            da_dsd = 'Đ'

                        else:

                            da_dsd = 'S'

                    if not da_dsa != '':

                        continue

                    if not da_dsb != '':

                        continue

                    if not da_dsc != '':

                        continue

                    if not da_dsd != '':

                        continue

                    sc_ds += 1

                    txt_ds = txt_ds + da_dsa + da_dsb + da_dsc + da_dsd + '_'

                    continue

                myrange = doc.Range(myrange_cau.End - 8, doc.Range().End)

                keywords = [

                    '[^13^9]ĐS:',

                    '[^13^9]Đáp số:',

                    '[^13^9]ĐA:',

                    '[^13^9]Đáp án:',

                    '[^13^9]TL:',

                    '[^13^9]Trả lời:']

                for keyword in keywords:

                    myrange_find2 = myrange_cau.Duplicate

                    find = myrange_find2.Find

                    find.ClearFormatting()

                    find.Text = keyword

                    find.MatchWildcards = True

                    find.MatchCase = False

                    if not find.Execute():

                        continue

                    sc_tln += 1

                    paragraph = myrange_find2.Paragraphs(myrange_find2.Paragraphs.Count)

                    myrange_find2.Start = myrange_find2.End

                    myrange_find2.End = paragraph.Range.End - 1

                    txt_cau = myrange_find2.Text.strip()

                    contains_mathtype = (lambda .0: pass# WARNING: Decompyle incomplete

)(myrange_find2.InlineShapes())

                    contains_image = (lambda .0: pass# WARNING: Decompyle incomplete

)(myrange_find2.InlineShapes())

                    if contains_mathtype:

                        txt_cau = 'Mathtype'

                    if contains_image:

                        txt_cau = 'Picture'

                    txt_tln = txt_tln + txt_cau + '_'

                    any

                continue

        txtDA = txtDA.strip('_')

        txtDA_split = txtDA.split('_')

        txt_ds = txt_ds.strip('_')

        txt_ds_split = txt_ds.split('_')

        txt_tln = txt_tln.strip('_')

        txt_tln_split = txt_tln.split('_')

        word.Selection.EndKey(Unit = 6)

        word.Selection.TypeParagraph()

        vbf.Page_Left_Right00(word)

        word.Selection.ParagraphFormat.Alignment = 1

        word.Selection.Font.Bold = True

        word.Selection.Font.Italic = False

        word.Selection.Font.Name = 'Times New Roman'

        word.Selection.Font.Size = 12

        word.Selection.Font.Color = win32api.RGB(0, 32, 96)

        word.Selection.TypeText('BẢNG ĐÁP ÁN TNMAKER')

        word.Selection.TypeParagraph()

        so_dong = sc_a + sc_ds + sc_tln

        range_end = doc.Content

        range_end.Collapse(Direction = 0)

        table = doc.Tables.Add(range_end, so_dong, 2)

        table.Borders.Enable = True

        table.Columns(1).Width = 50

        table.Columns(2).Width = 100

        for i in range(0, sc_a):

            table.Cell(i + 1, 1).Range.Text = str(i + 1)

            table.Cell(i + 1, 2).Range.Text = txtDA_split[i]

        for i in range(sc_a, sc_a + sc_ds):

            table.Cell(i + 1, 1).Range.Text = str(i + 1)

            table.Cell(i + 1, 2).Range.Text = txt_ds_split[i - sc_a]

        for i in range(sc_a + sc_ds, sc_a + sc_ds + sc_tln):

            table.Cell(i + 1, 1).Range.Text = str(i + 1)

            table.Cell(i + 1, 2).Range.Text = txt_tln_split[i - sc_a - sc_ds]

        vbf.autofit_one_table(word, table)

        table.ParagraphFormat.Alignment = 1

        vbf.xoa_dong_acong(word)

        vbf.xoa_dong_trang(word)

        word.Selection.EndKey(Unit = 6)

    DapAn2025_TNmaker_En(word)

    return None

# WARNING: Decompyle incomplete





def DapAn2025_TNmaker_En(word):

    doc = word.ActiveDocument

    vbf.Convert_Auto_To_Text(word)

    txtDA = ''

    txt_ds = ''

    txt_tln = ''

    sc_a = 0

    sc_ds = 0

    sc_tln = 0

    myrange = doc.Range()

    myrange_find = myrange.Duplicate

    find = myrange_find.Find

    find.ClearFormatting()

    find.Text = '(Question [0-9]{1,2}[.:])'

    find.MatchWildcards = True

    if find.Execute():

        vbf.add_blank_line_at_Home(word)

        vbf.add_blank_line_after_table(word)

        vbf.them_cau_acong_cuoi_En(word)

        for i in range(1, 500):

            d_a = ''

            da_dsa = ''

            da_dsb = ''

            da_dsc = ''

            da_dsd = ''

            myrange_cau = myrange.Duplicate

            find = myrange_cau.Find

            find.ClearFormatting()

            find.Text = '(Question [0-9]{1,2}[.:])(*)(Question [0-9]{1,2}[.:])'

            find.MatchWildcards = True

            if find.Execute():

                myrange = doc.Range(myrange_cau.End - 15, doc.Range().End)

                myrange_find = myrange_cau.Duplicate

                myrange_find.Select()

                find = myrange_find.Find

                find.ClearFormatting()

                find.Text = '([^13^9])([Dd])([.\\)])'

                find.MatchWildcards = True

                if find.Execute():

                    myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                    text = myrange_find.Text.strip()

                    if text == 'D':

                        myrange_find = myrange_cau.Duplicate

                        find = myrange_find.Find

                        find.ClearFormatting()

                        find.Text = '([^13^9])(A.)'

                        find.MatchWildcards = True

                        if find.Execute():

                            myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                            if myrange_find.Font.Underline == True or myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                                text = myrange_find.Text.strip()

                                d_a += text

                        myrange_find = myrange_cau.Duplicate

                        find = myrange_find.Find

                        find.ClearFormatting()

                        find.Text = '([^13^9])(B.)'

                        find.MatchWildcards = True

                        if find.Execute():

                            myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                            if myrange_find.Font.Underline == True or myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                                text = myrange_find.Text.strip()

                                d_a += text

                        myrange_find = myrange_cau.Duplicate

                        find = myrange_find.Find

                        find.ClearFormatting()

                        find.Text = '([^13^9])(C.)'

                        find.MatchWildcards = True

                        if find.Execute():

                            myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                            if myrange_find.Font.Underline == True or myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                                text = myrange_find.Text.strip()

                                d_a += text

                        myrange_find = myrange_cau.Duplicate

                        find = myrange_find.Find

                        find.ClearFormatting()

                        find.Text = '([^13^9])(D.)'

                        find.MatchWildcards = True

                        if find.Execute():

                            myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                            if myrange_find.Font.Underline == True or myrange_find.Font.Color == win32api.RGB(255, 0, 0):

                                text = myrange_find.Text.strip()

                                d_a += text

                        if d_a == '':

                            myrange_find = myrange_cau.Duplicate

                            find = myrange_find.Find

                            find.ClearFormatting()

                            find.Text = '(Chọn )([A-D])'

                            find.MatchWildcards = True

                            if find.Execute():

                                myrange_find.SetRange(Start = myrange_find.Start + 5, End = myrange_find.End)

                                text = myrange_find.Text.strip()

                                d_a += text

                        sc_a += 1

                        txtDA = txtDA + d_a + '_'

                        continue

                    dap_an_DS_cau = ''

                    keywords = [

                        '[^13^9]Key:',

                        '[^13^9]Answer:']

                    for keyword in keywords:

                        myrange_find2 = myrange_cau.Duplicate

                        find = myrange_find2.Find

                        find.ClearFormatting()

                        find.Text = keyword

                        find.MatchWildcards = True

                        find.MatchCase = False

                        if not find.Execute():

                            continue

                        sc_ds += 1

                        paragraph = myrange_find2.Paragraphs(myrange_find2.Paragraphs.Count)

                        myrange_find2.Start = myrange_find2.End

                        myrange_find2.End = paragraph.Range.End - 1

                        dap_an_DS_cau = myrange_find2.Text.strip()

                        txt_ds += dap_an_DS_cau + '_'

                        keywords

                    if not dap_an_DS_cau == '':

                        continue

                    myrange_find = myrange_cau.Duplicate

                    find = myrange_find.Find

                    find.ClearFormatting()

                    find.Text = '([^13^9])(a[\\)])'

                    find.MatchWildcards = True

                    if find.Execute():

                        myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                        if myrange_find.Font.Underline == True or myrange.Font.Color == win32api.RGB(255, 0, 0):

                            da_dsa = 'Đ'

                        else:

                            da_dsa = 'S'

                    myrange_find = myrange_cau.Duplicate

                    find = myrange_find.Find

                    find.ClearFormatting()

                    find.Text = '([^13^9])(b[\\)])'

                    find.MatchWildcards = True

                    if find.Execute():

                        myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                        if myrange_find.Font.Underline == True or myrange.Font.Color == win32api.RGB(255, 0, 0):

                            da_dsb = 'Đ'

                        else:

                            da_dsb = 'S'

                    myrange_find = myrange_cau.Duplicate

                    find = myrange_find.Find

                    find.ClearFormatting()

                    find.Text = '([^13^9])(c[\\)])'

                    find.MatchWildcards = True

                    if find.Execute():

                        myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                        if myrange_find.Font.Underline == True or myrange.Font.Color == win32api.RGB(255, 0, 0):

                            da_dsc = 'Đ'

                        else:

                            da_dsc = 'S'

                    myrange_find = myrange_cau.Duplicate

                    find = myrange_find.Find

                    find.ClearFormatting()

                    find.Text = '([^13^9])(d[\\)])'

                    find.MatchWildcards = True

                    if find.Execute():

                        myrange_find.SetRange(Start = myrange_find.Start + 1, End = myrange_find.End - 1)

                        if myrange_find.Font.Underline == True or myrange.Font.Color == win32api.RGB(255, 0, 0):

                            da_dsd = 'Đ'

                        else:

                            da_dsd = 'S'

                    if not da_dsa != '':

                        continue

                    if not da_dsb != '':

                        continue

                    if not da_dsc != '':

                        continue

                    if not da_dsd != '':

                        continue

                    sc_ds += 1

                    txt_ds = txt_ds + da_dsa + da_dsb + da_dsc + da_dsd + '_'

                    continue

                myrange = doc.Range(myrange_cau.End - 8, doc.Range().End)

                keywords = [

                    '[^13^9]Key:',

                    '[^13^9]Answer:']

                for keyword in keywords:

                    myrange_find2 = myrange_cau.Duplicate

                    find = myrange_find2.Find

                    find.ClearFormatting()

                    find.Text = keyword

                    find.MatchWildcards = True

                    find.MatchCase = False

                    if not find.Execute():

                        continue

                    sc_tln += 1

                    paragraph = myrange_find2.Paragraphs(myrange_find2.Paragraphs.Count)

                    myrange_find2.Start = myrange_find2.End

                    myrange_find2.End = paragraph.Range.End - 1

                    txt_cau = myrange_find2.Text.strip()

                    contains_mathtype = (lambda .0: pass# WARNING: Decompyle incomplete

)(myrange_find2.InlineShapes())

                    contains_image = (lambda .0: pass# WARNING: Decompyle incomplete

)(myrange_find2.InlineShapes())

                    if contains_mathtype:

                        txt_cau = 'Mathtype'

                    if contains_image:

                        txt_cau = 'Picture'

                    txt_tln = txt_tln + txt_cau + '_'

                    any

                continue

        txtDA = txtDA.strip('_')

        txtDA_split = txtDA.split('_')

        txt_ds = txt_ds.strip('_')

        txt_ds_split = txt_ds.split('_')

        txt_tln = txt_tln.strip('_')

        txt_tln_split = txt_tln.split('_')

        word.Selection.EndKey(Unit = 6)

        word.Selection.TypeParagraph()

        vbf.Page_Left_Right00(word)

        word.Selection.ParagraphFormat.Alignment = 1

        word.Selection.Font.Bold = True

        word.Selection.Font.Italic = False

        word.Selection.Font.Name = 'Times New Roman'

        word.Selection.Font.Size = 12

        word.Selection.Font.Color = win32api.RGB(0, 32, 96)

        word.Selection.TypeText('BẢNG ĐÁP ÁN TNMAKER')

        word.Selection.TypeParagraph()

        so_dong = sc_a + sc_ds + sc_tln

        range_end = doc.Content

        range_end.Collapse(Direction = 0)

        table = doc.Tables.Add(range_end, so_dong, 2)

        table.Borders.Enable = True

        table.Columns(1).Width = 50

        table.Columns(2).Width = 100

        for i in range(0, sc_a):

            table.Cell(i + 1, 1).Range.Text = str(i + 1)

            table.Cell(i + 1, 2).Range.Text = txtDA_split[i]

        for i in range(sc_a, sc_a + sc_ds):

            table.Cell(i + 1, 1).Range.Text = str(i + 1)

            table.Cell(i + 1, 2).Range.Text = txt_ds_split[i - sc_a]

        for i in range(sc_a + sc_ds, sc_a + sc_ds + sc_tln):

            table.Cell(i + 1, 1).Range.Text = str(i + 1)

            table.Cell(i + 1, 2).Range.Text = txt_tln_split[i - sc_a - sc_ds]

        vbf.autofit_one_table(word, table)

        table.ParagraphFormat.Alignment = 1

        vbf.xoa_dong_acong(word)

        vbf.xoa_dong_trang(word)

        word.Selection.EndKey(Unit = 6)

        return None

    return None

# WARNING: Decompyle incomplete



