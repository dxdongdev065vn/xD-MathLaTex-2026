import os
import tkinter as tk
from tkinter import ttk, Label, filedialog, Entry, Button, Frame, Listbox, Scrollbar, messagebox
from tkinter import simpledialog
import spire.doc as spire
from spire.doc import Section, Paragraph, Table, TextRange, Shape, FileFormat, HorizontalAlignment, Regex
from spire.doc import BookmarkStart, BookmarkEnd
from spire.doc import UnderlineStyle
from spire.doc import Document as SpireDocument
from spire.doc import UnderlineStyle
from spire.doc import Color
from spire.doc import DocumentObjectType
from spire.doc import OfficeMath
import re
import random
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import GradientFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from formatting import chuan_hoa_spr as chspr
from tools import tool_by_docx as docxtool

def open_doc_off_spire(doc_path):
    doc = SpireDocument()
    doc.LoadFromFile(doc_path)
    return doc


def new_doc_off_spire():
    doc = SpireDocument()
    doc.AddSection()
    return doc


def InchesToPoints(inches):
    return inches * 72


def LinesToPoints(lines):
    return lines * 12


def remove_last_element(lst):
    return lst[:-1]


def copy_any_element(input_path, output_path):
    input_doc = open_doc_off_spire(input_path)
    if os.path.exists(output_path):
        output_doc = open_doc_off_spire(output_path)
    else:
        output_doc = SpireDocument()
        output_doc.AddSection()
    input_section = input_doc.Sections[0]
    output_section = output_doc.Sections[0]
    elements = input_section.Body.ChildObjects
    for idx in range(1, 8):
        element = elements.get_Item(idx)
        output_section.Body.ChildObjects.Add(element.Clone())
    output_doc.SaveToFile(output_path, FileFormat.Docx)


def browse_folder_entry(entry_folder_path):
    folder_path = filedialog.askdirectory()
    if not folder_path:
        return None
    folder_path = os.path.normpath(folder_path)
    entry_folder_path.delete(0, tk.END)
    entry_folder_path.insert(tk.END, folder_path)


def browse_file(listbox_tach_de):
    files = filedialog.askopenfilenames(filetypes = [
        ('Word Files', '*.docx;*.doc')])
    if not files:
        return None
    for file in files:
        input_file = os.path.normpath(file)
        listbox_tach_de.insert(tk.END, input_file)


def clear_files(listbox_tach_de):
    listbox_tach_de.delete(0, tk.END)


def Tim_kiem_cau_hoi(root):
    pass
# WARNING: Decompyle incomplete


def Thay_the_spr(root):
    pass
# WARNING: Decompyle incomplete

if __name__ == '__main__':
    root = tk.Tk()
    Thay_the_spr(root)
    return None
