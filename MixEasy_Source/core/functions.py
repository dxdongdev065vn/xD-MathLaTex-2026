import os

import datetime

import time

import win32com.client as win32com

from win32com.client import client as win32

from win32com.client import constants

import win32api

import re

import sys

import shutil

import pythoncom

import subprocess

import tempfile

import wmi

import docx

import random

import webbrowser

from docx import Document

from docx.shared import Pt

from docx.enum.text import WD_PARAGRAPH_ALIGNMENT

from docx.shared import RGBColor

from openpyxl import Workbook

from openpyxl.styles import Font

from openpyxl.styles import Alignment

from openpyxl.styles import Border, Side

from openpyxl.styles import PatternFill

from openpyxl.styles import GradientFill

from openpyxl.utils import get_column_letter

from openpyxl import load_workbook

import tkinter as tk

from tkinter import ttk, Label, filedialog, Entry, Button, Frame, Listbox, Scrollbar, messagebox

from tkinter import simpledialog

from tkinter import Toplevel

import pythoncom

from urllib.parse import unquote



def reset_com():

    pythoncom.CoUninitialize()

    pythoncom.CoInitialize()





def xoa_duong_dan_onedrive_pc7(doc_path):

    bad_part = '\\Python\\VBA\\PC7\\https:\\tcdcnh-my.sharepoint.com\\personal\\hungnn_giahoi_cee_edu_vn\\Documents'

    if bad_part in doc_path:

        doc_path = doc_path.replace(bad_part, '')

    return doc_path





def chuan_hoa_duong_dan(input_path):

    input_path = unquote(input_path)

    input_path = os.path.normpath(os.path.abspath(input_path))

    input_path = xoa_duong_dan_onedrive_pc7(input_path)

    return input_path





def open_folder(entry_input):

    '''

    Mở thư mục bằng File Explorer trên Windows

    '''

    file_path = entry_input.get()

    file_path = os.path.normpath(file_path)

    thu_muc = os.path.dirname(file_path)

    if os.path.exists(thu_muc):

        os.startfile(thu_muc)

        return None

    print('Đường dẫn không tồn tại!')

    messagebox.showerror('Lỗi', 'đường dẫn không tồn tại!')





def cleanup_pyinstaller_temp():

    temp_dir = tempfile.gettempdir()

    for item in os.listdir(temp_dir):

        item_path = os.path.join(temp_dir, item)

        if not item.startswith('_MEI'):

            continue

        if not os.path.isdir(item_path):

            continue

        shutil.rmtree(item_path)

    return None

# WARNING: Decompyle incomplete





def rebuild_win32com_cache():

    

    gencache

    False = import win32com.client.gencache, client

    gencache.Rebuild()

    print('✔ Rebuilt win32com gencache')

    return None

# WARNING: Decompyle incomplete





def clear_win32com_cache():

    gencache_path = os.path.join(os.environ['LOCALAPPDATA'], 'Temp', 'gen_py')

    if os.path.exists(gencache_path):

        shutil.rmtree(gencache_path)

        return None

    return None

# WARNING: Decompyle incomplete





def clear_win32com_cache_2():

    folder = win32com.client.gencache.GetGeneratePath()

    if os.path.exists(folder):

        shutil.rmtree(folder)

        os.mkdir(folder)

        return None

    return None

# WARNING: Decompyle incomplete





def clear_win32com_cache_all():

    cleanup_pyinstaller_temp()

    clear_win32com_cache()

    clear_win32com_cache_2()

    return None

# WARNING: Decompyle incomplete





def khoi_tao_word_goc():

    '''Khởi tạo Word COM, nếu gặp lỗi sẽ Quit và khởi động lại'''

    word = win32com.client.gencache.EnsureDispatch('Word.Application')

    return word

# WARNING: Decompyle incomplete





def khoi_tao_word():

    messages_loi = []

    word = win32com.client.gencache.EnsureDispatch('Word.Application')

    print('✔ 1. Đã tạo Word EnsureDispatch')

    return word

# WARNING: Decompyle incomplete





def khoi_tao_word_2():

    print('Khởi tạo lại word 2')

    word = win32com.client.gencache.EnsureDispatch('Word.Application')

    print('✔ 1. Đã tạo Word EnsureDispatch')

    return word

# WARNING: Decompyle incomplete





def khoi_tao_powerpoint_goc():

    '''Khởi tạo PowerPoint COM, nếu gặp lỗi sẽ Quit và khởi động lại'''

    powerpoint = win32com.client.gencache.EnsureDispatch('PowerPoint.Application')

    return powerpoint

# WARNING: Decompyle incomplete





def khoi_tao_powerpoint():

    powerpoint = win32com.client.gencache.EnsureDispatch('PowerPoint.Application')

    print('✔ 2. Khởi tạo PowerPoint thành công bằng EnsureDispatch.')

    return powerpoint

# WARNING: Decompyle incomplete





def khoi_tao_powerpoint_2():

    powerpoint = win32com.client.gencache.EnsureDispatch('PowerPoint.Application')

    print('✔ 2. Khởi tạo PowerPoint thành công bằng EnsureDispatch.')

    return powerpoint

# WARNING: Decompyle incomplete





def Refresh_word():

    rebuild_win32com_cache()

    word = khoi_tao_word()





def Refresh_powerpoint():

    rebuild_win32com_cache()

    powerpoint = khoi_tao_powerpoint()





def restart_program():

    '''Restart sạch sẽ cho cả .py và .exe onefile'''

    time.sleep(0.5)

    env = dict(os.environ)

    env['PYINSTALLER_RESET_ENVIRONMENT'] = '1'

    if getattr(sys, 'frozen', False):

        subprocess.Popen([

            sys.executable], env = env)

    else:

        subprocess.Popen([

            sys.executable] + sys.argv, env = env)

    os._exit(0)

    return None

# WARNING: Decompyle incomplete





def restart_my_program2():

    result = messagebox.askokcancel('Thông báo', 'Bạn phải tắt hết word, powerpoint mới được bấm nút này\nNút này có chức năng sửa chữa lỗi về hoạt động, sau đó chương trình sẽ tự tắt và tự mở lại chương trình')

    if result:

        clear_win32com_cache_all()

        word = khoi_tao_word()

        word.Quit()

        powerpoint = khoi_tao_powerpoint()

        powerpoint.Quit()

        exe_path = sys.executable

        os.startfile(exe_path)

        sys.exit()

        return None

    return None

# WARNING: Decompyle incomplete





def restart_my_program():

    '''Hàm để tự tắt và khởi động lại chương trình.'''

    print('chuẩn bị tắt và khởi động')

    exe_path = sys.executable

    os.startfile(exe_path)

    sys.exit()

    return None

# WARNING: Decompyle incomplete





def Exit_my_program():

    result = messagebox.askokcancel('Thông báo', 'Bạn phải tắt hết word, powerpoint mới được bấm nút này\nNút này có chức năng sửa chữa lỗi về hoạt động, sau đó chương trình sẽ tự tắt và bạn khởi động lại để dùng')

    if result:

        clear_win32com_cache_all()

        word = khoi_tao_word()

        word.Quit()

        powerpoint = khoi_tao_powerpoint()

        powerpoint.Quit()

        sys.exit(0)

        return None





def LinesToPoints(lines):

    return lines * 12





def InchesToPoints(inches):

    return inches * 72





def CentimetersToPoints(cm):

    return cm * 28.3465





def Select_All(word):

    doc = word.ActiveDocument

    doc.Content.Select()

    return None

# WARNING: Decompyle incomplete





def mo_rong_vung_chon(doc, myrange):

    doc_length = doc.Content.End

    start_old = myrange.Start

    end_old = myrange.End

    new_start = start_old - 1 if start_old > 0 else start_old

    new_end = end_old + 1 if end_old + 1 <= doc_length else end_old

    myrange.SetRange(Start = new_start, End = new_end)

    return myrange

# WARNING: Decompyle incomplete





def them_cau_acong_cuoi(word):

    doc = word.ActiveDocument

    selection = word.Selection

    selection.EndKey(Unit = 6)

    word.Selection.TypeParagraph()

    word.Selection.TypeText('Câu 00. @:')

    word.Selection.TypeParagraph()

    return None

# WARNING: Decompyle incomplete





def them_cau_acong_cuoi_En(word):

    doc = word.ActiveDocument

    selection = word.Selection

    selection.EndKey(Unit = 6)

    word.Selection.TypeParagraph()

    word.Selection.TypeText('Question 00. @:')

    word.Selection.TypeParagraph()

    return None

# WARNING: Decompyle incomplete





def them_acong_dau(word):

    doc = word.ActiveDocument

    selection = word.Selection

    selection.HomeKey(Unit = 6)

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '(Câu [0-9]{1,2})'

    find.Replacement.Text = 'Câu 00. @:^13\\1'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Format = True

    find.Execute(Replace = 1)

    return None

# WARNING: Decompyle incomplete





def ThemcauaoPY(word):

    doc = word.ActiveDocument

    rng = doc.Content

    find = rng.Find

    find.Text = '^13Câu'

    find.MatchWildcards = True

    find.Forward = True

    if find.Execute():

        if find.Found:

            rng.EndOf(Unit = 4)

            find.Execute()

            if find.Found:

                continue

        rng.InsertParagraphAfter()

        rng.Collapse(Direction = 0)

        rng.Text = 'Câu @@@ @@@:\n'

        return None

    return None

# WARNING: Decompyle incomplete





def thay_the_replace_Mix(doc, word, text, text_replace):

    rng = doc.Content

    find = rng.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = text

    find.Replacement.Text = text_replace

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = False

    find.Execute(Replace = 2)





def thay_the_replace_12_xanh(word, text, text_replace):

    doc = word.ActiveDocument

    rng = doc.Content

    find = rng.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = text

    find.Replacement.Text = text_replace

    find.Replacement.Font.Bold = True

    find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def thay_the_replace(word, text, text_replace):

    doc = word.ActiveDocument

    rng = doc.Content

    find = rng.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = text

    find.Replacement.Text = text_replace

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def thay_the_replace_no_format(word, text, text_replace):

    doc = word.ActiveDocument

    rng = doc.Content

    find = rng.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = False

    find.Replacement.Font.Underline = False

    find.Replacement.Font.Color = win32api.RGB(0, 0, 0)

    find.Text = text

    find.Replacement.Text = text_replace

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def thay_the_replace_false(word, text, text_replace):

    doc = word.ActiveDocument

    rng = doc.Content

    find = rng.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = text

    find.Replacement.Text = text_replace

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = False

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def thay_the_replace_MatchCase_false(word, text, text_replace):

    doc = word.ActiveDocument

    rng = doc.Content

    find = rng.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = text

    find.Replacement.Text = text_replace

    find.Forward = True

    find.MatchCase = False

    find.MatchWholeWord = False

    find.MatchWildcards = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = False

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def thay_the_replace_1(word, text, text_replace):

    doc = word.ActiveDocument

    rng = doc.Content

    find = rng.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = text

    find.Replacement.Text = text_replace

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = False

    find.Execute(Replace = 1)

    return None

# WARNING: Decompyle incomplete





def thay_the_replace_stop(word, text, text_replace):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        rng = word.Selection.Range

        find = rng.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = text

        find.Replacement.Text = text_replace

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        find.Execute(Replace = 2)

        return None

    return None

# WARNING: Decompyle incomplete





def thay_the_replace_stop_12_xanh(word, text, text_replace):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        rng = word.Selection.Range

        find = rng.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = text

        find.Replacement.Text = text_replace

        find.Replacement.Font.Bold = True

        find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = True

        find.Execute(Replace = 2)

        return None

    return None

# WARNING: Decompyle incomplete





def thay_the_replace_stop_no_format(word, text, text_replace):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        rng = word.Selection.Range

        find = rng.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = text

        find.Replacement.Text = text_replace

        find.Replacement.Font.Bold = False

        find.Replacement.Font.Underline = False

        find.Replacement.Font.Color = win32api.RGB(0, 0, 0)

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = True

        find.Execute(Replace = 2)

        return None

    return None

# WARNING: Decompyle incomplete





def thay_the_replace_stop_false(word, text, text_replace):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        rng = word.Selection.Range

        find = rng.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = text

        find.Replacement.Text = text_replace

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = False

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        find.Execute(Replace = 2)

        return None

    return None

# WARNING: Decompyle incomplete





def thay_the_replace_stop_1(word, text, text_replace):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        rng = word.Selection.Range

        find = rng.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = text

        find.Replacement.Text = text_replace

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        find.Execute(Replace = 1)

        return None

    return None

# WARNING: Decompyle incomplete





def Convert_Auto_To_Text(word):

    doc = word.ActiveDocument

    myrange = doc.Range()

    myrange.ListFormat.ConvertNumbersToText()

    return None

# WARNING: Decompyle incomplete





def convert_table_to_text(word):

    doc = word.ActiveDocument

    selection = word.Selection

    myrange = word.Selection.Range

    table_count = myrange.Tables.Count

    if table_count > 0:

        for i in range(table_count, 0, -1):

            table = myrange.Tables(i)

            table.ConvertToText(Separator = ' ', NestedTables = False)

        return None

    messagebox.showinfo('Thông báo', 'Chưa có Bảng được chọn')

    return None

# WARNING: Decompyle incomplete





def convert_table_to_text_tab(word):

    doc = word.ActiveDocument

    selection = word.Selection

    myrange = word.Selection.Range

    table_count = myrange.Tables.Count

    if table_count > 0:

        for i in range(table_count, 0, -1):

            table = myrange.Tables(i)

            table.ConvertToText(Separator = 1, NestedTables = False)

        return None

    messagebox.showinfo('Thông báo', 'Chưa có Bảng được chọn')

    return None

# WARNING: Decompyle incomplete





def convert_table_to_text_All(word):

    doc = word.ActiveDocument

    selection = word.Selection

    myrange = word.Selection.Range

    table_count = myrange.Tables.Count

    if table_count > 0:

        for i in range(table_count, 0, -1):

            table = myrange.Tables(i)

            table.ConvertToText(Separator = ' ', NestedTables = True)

        return None

    messagebox.showinfo('Thông báo', 'Chưa có Bảng được chọn')

    return None

# WARNING: Decompyle incomplete





def convert_table_to_text_All_tab(word):

    doc = word.ActiveDocument

    selection = word.Selection

    myrange = word.Selection.Range

    table_count = myrange.Tables.Count

    if table_count > 0:

        for i in range(table_count, 0, -1):

            table = myrange.Tables(i)

            table.ConvertToText(Separator = win32com.client.constants.wdSeparateByTabs, NestedTables = True)

        return None

    messagebox.showinfo('Thông báo', 'Chưa có Bảng được chọn')

    return None

# WARNING: Decompyle incomplete





def Table_to_text_tool(root, word):

    pass

# WARNING: Decompyle incomplete





def xoa_cau_00(word):

    doc = word.ActiveDocument

    for paragraph in doc.Paragraphs:

        if not 'Câu 00' in paragraph.Range.Text:

            continue

        paragraph.Range.Delete()

    for paragraph in doc.Paragraphs:

        if not 'Question 00' in paragraph.Range.Text:

            continue

        paragraph.Range.Delete()

    return None

# WARNING: Decompyle incomplete





def xoa_dong_acong(word):

    doc = word.ActiveDocument

    for paragraph in doc.Paragraphs:

        if not '@:' in paragraph.Range.Text:

            continue

        paragraph.Range.Delete()

    return None

# WARNING: Decompyle incomplete





def xoa_dong_dap_so(word):

    doc = word.ActiveDocument

    for paragraph in doc.Paragraphs:

        if not 'ĐS:' in paragraph.Range.Text:

            continue

        paragraph.Range.Delete()

    return None

# WARNING: Decompyle incomplete





def xoa_dong_tra_loi(doc, word):

    for paragraph in doc.Paragraphs:

        if not 'Trả lời:' in paragraph.Range.Text:

            continue

        paragraph.Range.Delete()





def xoa_multi_space(word):

    doc = word.ActiveDocument

    rng = doc.Content

    find = rng.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '[ ]{2,}'

    find.Replacement.Text = ' '

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = False

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def xoa_to_home(word):

    doc = word.ActiveDocument

    selection = word.Selection

    delrange = selection.Range

    delrange.Start = doc.Range().Start

    delrange.End = delrange.End

    delrange.Delete()

    return None

# WARNING: Decompyle incomplete





def xoa_to_end(word):

    doc = word.ActiveDocument

    selection = word.Selection

    delrange = selection.Range

    delrange.Start = delrange.End

    delrange.End = doc.Range().End

    delrange.Delete()

    return None

# WARNING: Decompyle incomplete





def Xoa_heading(word, heading_idx):

    doc = word.ActiveDocument

    myrange = doc.Range()

    find = myrange.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Style = doc.Styles(f'''Heading {heading_idx}''')

    find.Text = ''

    find.Replacement.Text = ''

    find.Forward = True

    find.Wrap = 1

    find.Format = True

    find.MatchCase = False

    find.MatchWholeWord = False

    find.MatchWildcards = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def xoa_more_tool(root, word):

    pass

# WARNING: Decompyle incomplete





def xoa_dong_trang_Mix(doc, word):

    thay_the_replace(word, '[^13]{2,}', '^p')

    for paragraph in doc.Paragraphs:

        if not paragraph.Range.Text.strip() == '':

            continue

        paragraph.Range.Delete()

    return None

# WARNING: Decompyle incomplete





def xoa_dong_trang(word):

    doc = word.ActiveDocument

    xoa_dong_trang_Mix(doc, word)

    return None

# WARNING: Decompyle incomplete





def xoa_dong_text(word, text_del):

    doc = word.ActiveDocument

    for paragraph in doc.Paragraphs:

        if not text_del in paragraph.Range.Text:

            continue

        paragraph.Range.Delete()

    return None

# WARNING: Decompyle incomplete





def xoa_dong_startswith_text(word, text_del):

    doc = word.ActiveDocument

    for paragraph in doc.Paragraphs:

        text = paragraph.Range.Text.strip()

        if not text.lower().startswith(text_del.lower()):

            continue

        paragraph.Range.Delete()

    return None

# WARNING: Decompyle incomplete





def xoa_dong_startswith_text_any(word, prefixes):

    pass

# WARNING: Decompyle incomplete





def xoa_shift_enter(word):

    doc = word.ActiveDocument

    thay_the_replace(word, '^l', '^p')

    thay_the_replace(word, '^13', '^p')

    thay_the_replace(word, '^m', '^p')

    return None

# WARNING: Decompyle incomplete





def xoa_tac_gia(word):

    pass

# WARNING: Decompyle incomplete





def nhap_keywords_toplevel():

    pass

# WARNING: Decompyle incomplete





def xoa_key_nhap(root, word):

    pass

# WARNING: Decompyle incomplete





def xoa_headers_footers(word):

    doc = word.ActiveDocument

    for section in doc.Sections:

        for header in section.Headers:

            if not header.Exists:

                continue

            header.Range.Delete()

        for footer in section.Footers:

            if not footer.Exists:

                continue

            footer.Range.Delete()

    return None

# WARNING: Decompyle incomplete





def Xoa_chu_thich_cau(word):

    doc = word.ActiveDocument

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '(Câu [0-9]{1,3}[.:])([\\[\\(])(*)([\\]\\)])'

    find.Replacement.Text = '\\1'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Execute(Replace = 2)

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '(Câu [0-9]{1,3}[.:])([^32^9]{1,})([\\[\\(])(*)([\\]\\)])'

    find.Replacement.Text = '\\1 '

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def Xoa_chu_thich_vidu(word):

    doc = word.ActiveDocument

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '(Ví dụ [0-9]{1,3}[.:])([\\[\\(])(*)([\\]\\)])'

    find.Replacement.Text = '\\1'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Execute(Replace = 2)

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '(Ví dụ [0-9]{1,3}[.:])([^32^9]{1,})([\\[\\(])(*)([\\]\\)])'

    find.Replacement.Text = '\\1 '

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def Xoa_chu_thich_pan(word):

    doc = word.ActiveDocument

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '([^13^9][a-d]\\))([\\[\\(])(*)([\\]\\)])'

    find.Replacement.Text = '\\1'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Execute(Replace = 2)

    find.Text = '([^13^9][a-d]\\))([ ^9]{1,})([\\[\\(])(*)([\\]\\)])'

    find.Replacement.Text = '\\1 '

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def xoa_chu_thich_all(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        rng = word.Selection.Range

    else:

        rng = doc.Content

    for i in range(0, 2):

        rng.Select()

        Xoa_chu_thich_cau(word)

        rng.Select()

        Xoa_chu_thich_vidu(word)

        rng.Select()

        Xoa_chu_thich_pan(word)

    return None

# WARNING: Decompyle incomplete





def color_tool(root, word):

    pass

# WARNING: Decompyle incomplete





def Bold_to_blue(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        find = word.Selection.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Font.Bold = True

        find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

        find.Text = ''

        find.Replacement.Text = ''

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = False

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = True

        find.Execute(Replace = 2)

        return None

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Font.Bold = True

    find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

    find.Text = ''

    find.Replacement.Text = ''

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def Select_mau_blue(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange.Font.Color = win32api.RGB(0, 0, 255)

        myrange.Font.Bold = True

        return None

    messagebox.showinfo('Thông báo', 'Chưa chọn vùng')

    return None

# WARNING: Decompyle incomplete





def Select_mau_green(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange.Font.Color = win32api.RGB(0, 128, 0)

        myrange.Font.Bold = True

        return None

    messagebox.showinfo('Thông báo', 'Chưa chọn vùng')

    return None

# WARNING: Decompyle incomplete





def Select_mau_purple(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange = word.Selection.Range

        myrange.Font.Color = win32api.RGB(148, 0, 211)

        myrange.Font.Bold = True

        return None

    messagebox.showinfo('Thông báo', 'Chưa chọn vùng')

    return None

# WARNING: Decompyle incomplete





def Mau_black_All(word):

    doc = word.ActiveDocument

    myrange = doc.Range()

    myrange.Font.Color = win32api.RGB(0, 0, 0)

    return None

# WARNING: Decompyle incomplete





def dam_xanh_cau_En(word):

    doc = word.ActiveDocument

    Convert_Auto_To_Text(word)

    if word.Selection.Type != 1:

        find = word.Selection.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Replacement.Font.Bold = True

        find.Replacement.Font.Italic = False

        find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

        find.Text = '(Question [0-9]{1,3}[:.])'

        find.Replacement.Text = '\\1'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = True

        find.Execute(Replace = 2)

        return None

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = True

    find.Replacement.Font.Italic = False

    find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

    find.Text = '(Question [0-9]{1,3}[:.])'

    find.Replacement.Text = '\\1'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def dam_xanh_cau_for_new(word, text):

    doc = word.ActiveDocument

    Convert_Auto_To_Text(word)

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange = mo_rong_vung_chon(doc, myrange)

    else:

        myrange = doc.Range()

    text_find = '(^13' + text + ' [0-9]{1,}[.:])'

    find = myrange.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = True

    find.Replacement.Font.Underline = False

    find.Replacement.Font.Italic = False

    find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

    find.Text = text_find

    find.Replacement.Text = '\\1'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def dam_xanh_cau_font12_for_new(word, text):

    doc = word.ActiveDocument

    Convert_Auto_To_Text(word)

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange = mo_rong_vung_chon(doc, myrange)

    else:

        myrange = doc.Range()

    text_find = '(^13' + text + ' [0-9]{1,}[.:])'

    find = myrange.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = True

    find.Replacement.Font.Underline = False

    find.Replacement.Font.Italic = False

    find.Replacement.Font.Size = 12

    find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

    find.Text = text_find

    find.Replacement.Text = '\\1'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def dam_xanh_cau_font12(word):

    dam_xanh_cau_font12_for_new(word, 'Câu')

    dam_xanh_cau_font12_for_new(word, 'Question')

    dam_xanh_cau_font12_for_new(word, 'Bài')





def dam_xanh_cau(word):

    dam_xanh_cau_for_new(word, 'Câu')

    dam_xanh_cau_for_new(word, 'Question')

    dam_xanh_cau_for_new(word, 'Bài')





def dam_xanh_ABCD_cham(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange = mo_rong_vung_chon(doc, myrange)

        find = myrange.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Replacement.Font.Bold = True

        find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

        find.Text = '([^13^9])([ABCD].)'

        find.Replacement.Text = '\\1\\2'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = True

        find.Execute(Replace = 2)

        return None

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = True

    find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

    find.Text = '([^13^9])([ABCD].)'

    find.Replacement.Text = '\\1\\2'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def dam_xanh_abcd_ngoac_tool(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange = mo_rong_vung_chon(doc, myrange)

        find = myrange.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Replacement.Font.Bold = True

        find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

        find.Text = '([^13^9])([abcd])([\\)])'

        find.Replacement.Text = '\\1\\2\\3'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = True

        find.Execute(Replace = 2)

        return None

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = True

    find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

    find.Text = '([^13^9])([abcd])([\\)])'

    find.Replacement.Text = '\\1\\2\\3'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def gachchan_mau_do_tool(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange = mo_rong_vung_chon(doc, myrange)

    else:

        myrange = doc.Range()

    for i in range(1, 300):

        myrange_cau = myrange.Duplicate

        find = myrange_cau.Find

        find.ClearFormatting()

        find.Text = '([^13^9][ABCDabcd][.\\)])'

        find.MatchWildcards = True

        if find.Execute():

            myrange = doc.Range(myrange_cau.End, myrange.End)

            if myrange_cau.Tables.Count > 0:

                table = myrange_cau.Tables(1)

                table_range = table.Range

                table_end = table_range.End

                myrange.SetRange(Start = table_end, End = myrange.End)

            for Under_line in (1, 3, 6):

                myrange_find = myrange_cau.Duplicate

                find = myrange_find.Find

                find.ClearFormatting()

                find.Font.Underline = Under_line

                find.Replacement.ClearFormatting()

                find.Replacement.Font.Bold = True

                find.Replacement.Font.Underline = 1

                find.Replacement.Font.Color = win32api.RGB(255, 0, 0)

                find.Text = '([ABCDabcd])'

                find.Replacement.Text = '\\1'

                find.Forward = True

                find.MatchCase = True

                find.MatchWholeWord = False

                find.MatchWildcards = True

                find.MatchSoundsLike = False

                find.MatchAllWordForms = False

                find.Wrap = 0

                find.Format = True

                find.Execute(Replace = 1)

            continue

        return None

    return None

# WARNING: Decompyle incomplete





def ABCD_ngoac_qua_cham(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange = mo_rong_vung_chon(doc, myrange)

        find = myrange.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Replacement.Font.Bold = True

        find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

        find.Text = '([^13^9])([ABCD])([\\)])'

        find.Replacement.Text = '\\1\\2.'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = True

        find.Execute(Replace = 2)

        return None

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = True

    find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

    find.Text = '([^13^9])([A_D])([\\)])'

    find.Replacement.Text = '\\1\\2.'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def maudo_to_gachchan(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange = mo_rong_vung_chon(doc, myrange)

    else:

        myrange = doc.Range()

    for i in range(1, 501):

        myrange_cau = myrange.Duplicate

        find = myrange_cau.Find

        find.ClearFormatting()

        find.Text = '([^13^9][ABCDabcd][.\\)])'

        find.MatchWildcards = True

        if find.Execute():

            myrange = doc.Range(myrange_cau.End, myrange.End)

            rng = myrange_cau.Duplicate

            if rng.Tables.Count > 0:

                table = rng.Tables(1)

                table_range = table.Range

                table_end = table_range.End

                myrange.SetRange(Start = table_end, End = myrange.End)

            find = rng.Find

            find.ClearFormatting()

            find.Replacement.ClearFormatting()

            find.Text = '([ABCDabcd])'

            find.Forward = True

            find.MatchCase = True

            find.MatchWholeWord = False

            find.MatchWildcards = True

            find.MatchSoundsLike = False

            find.MatchAllWordForms = False

            find.Wrap = 0

            find.Format = True

            find.Replacement.Text = '\\1'

            find.Replacement.Font.Underline = 1

            find.Replacement.Font.Color = win32api.RGB(255, 0, 0)

            find.Replacement.Font.Bold = True

            if not find.Execute():

                continue

            if not rng.Font.Color == win32api.RGB(255, 0, 0) and rng.Font.Color == win32api.RGB(238, 0, 0) and rng.Font.Color == win32api.RGB(192, 0, 0):

                continue

            find.Execute(Replace = 1)

            continue

        range(1, 501)

        return None

    return None

# WARNING: Decompyle incomplete





def Highlight_to_mau_do_tool(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange = mo_rong_vung_chon(doc, myrange)

    else:

        myrange = doc.Range()

    for i in range(1, 100):

        myrange_cau = myrange.Duplicate

        find = myrange_cau.Find

        find.ClearFormatting()

        find.Text = '([^13^9][ABCDabcd][.\\)])'

        find.MatchWildcards = True

        if find.Execute():

            myrange = doc.Range(myrange_cau.End, myrange.End)

            if myrange_cau.Tables.Count > 0:

                table = myrange_cau.Tables(1)

                table_range = table.Range

                table_end = table_range.End

                myrange.SetRange(Start = table_end, End = myrange.End)

            find = myrange_cau.Find

            find.ClearFormatting()

            find.Highlight = True

            find.Replacement.ClearFormatting()

            find.Replacement.Font.Bold = True

            find.Replacement.Font.Underline = 1

            find.Replacement.Font.Color = win32api.RGB(255, 0, 0)

            find.Text = '([ABCDabcd])'

            find.Replacement.Text = '\\1'

            find.Forward = True

            find.MatchCase = True

            find.MatchWholeWord = False

            find.MatchWildcards = True

            find.MatchSoundsLike = False

            find.MatchAllWordForms = False

            find.Wrap = 0

            find.Format = True

            find.Execute(Replace = 1)

            continue

        range(1, 100)

        return None

    return None

# WARNING: Decompyle incomplete





def dam_xanh_vidu(word):

    doc = word.ActiveDocument

    Convert_Auto_To_Text(word)

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        find = myrange.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Replacement.Font.Bold = True

        find.Replacement.Font.Italic = False

        find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

        find.Text = '(Ví dụ [0-9]{1,3}[:.])'

        find.Replacement.Text = '\\1'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = True

        find.Execute(Replace = 2)

        return None

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = True

    find.Replacement.Font.Italic = False

    find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

    find.Text = '(Ví dụ [0-9]{1,3}[:.])'

    find.Replacement.Text = '\\1'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def Dam_cau_pa(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        rng = word.Selection.Range

    else:

        rng = doc.Content

    add_blank_line_after_table(word)

    rng.Select()

    dam_xanh_cau(word)

    rng.Select()

    dam_xanh_ABCD_cham(word)

    rng.Select()

    dam_xanh_abcd_ngoac_tool(word)

    rng.Select()

    dam_xanh_vidu(word)

    xoa_dong_trang(word)

    return None

# WARNING: Decompyle incomplete





def Gach_chan_red_dap_an(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        rng = word.Selection.Range

    else:

        rng = doc.Content

    add_blank_line_after_table(word)

    rng.Select()

    Highlight_to_mau_do_tool(word)

    rng.Select()

    maudo_to_gachchan(word)

    rng.Select()

    gachchan_mau_do_tool(word)

    xoa_dong_trang(word)

    return None

# WARNING: Decompyle incomplete





def dam_red_Dapso(word):

    doc = word.ActiveDocument

    Convert_Auto_To_Text(word)

    if word.Selection.Type != 1:

        rng = word.Selection.Range

        rng = mo_rong_vung_chon(doc, myrange)

    else:

        rng = doc.Content

    add_blank_line_after_table(word)

    find = rng.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = True

    find.Replacement.Font.Color = win32api.RGB(255, 0, 0)

    find.Text = '(ĐS:)(*)(^13)'

    find.Replacement.Text = '\\1\\2\\3'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Format = True

    find.Execute(Replace = 2)

    xoa_dong_trang(word)

    return None

# WARNING: Decompyle incomplete





def Bold_color_key(word):

    pass

# WARNING: Decompyle incomplete





def Bold_purple_chi_muc(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange = mo_rong_vung_chon(doc, myrange)

    else:

        myrange = doc.Content

    for i in range(1, 100):

        myrange_find = myrange.Duplicate

        find = myrange_find.Find

        find.ClearFormatting()

        find.Text = '^13[0-9].'

        find.MatchWildcards = True

        if find.Execute():

            current_line = myrange_find.Paragraphs(2).Range

            current_line.Font.Color = win32api.RGB(148, 0, 211)

            current_line.Font.Bold = True

            myrange.SetRange(Start = myrange_find.End, End = myrange.End)

            myrange_find = myrange.Duplicate

            find = myrange_find.Find

            continue

        range(1, 100)

        return None

    return None

# WARNING: Decompyle incomplete





def fix_chuthich_cau(word):

    doc = word.ActiveDocument

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '(Câu [0-9]{1,2}.)([\\[\\(])(*)([\\]\\)])'

    find.Replacement.Text = '\\1 \\2\\3\\4'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Format = False

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def fix_chuthich_vidu(word):

    doc = word.ActiveDocument

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '(Ví dụ [0-9]{1,2}[:.])([\\[\\(])(*)([\\]\\)])'

    find.Replacement.Text = '\\1 \\2\\3\\4'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Format = False

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def dam_mau_chu_thich_vidu(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

    else:

        myrange = doc.Range()

    for i in range(1, 100):

        myrange_cau = myrange.Duplicate

        find = myrange_cau.Find

        find.ClearFormatting()

        find.Text = '(Ví dụ [0-9]{1,2}[:.] )([\\[\\(])(*)([\\]\\)])'

        find.MatchWildcards = True

        if find.Execute():

            myrange_find = myrange_cau.Duplicate

            myrange = doc.Range(myrange_cau.End, myrange.End)

            find = myrange_find.Find

            find.ClearFormatting()

            find.Replacement.ClearFormatting()

            find.Replacement.Font.Bold = True

            find.Replacement.Font.Color = win32api.RGB(253, 11, 252)

            find.Text = '(\\[)(*)(\\])'

            find.Replacement.Text = '\\1\\2\\3'

            find.Forward = True

            find.MatchCase = True

            find.MatchWholeWord = False

            find.MatchWildcards = True

            find.MatchSoundsLike = False

            find.MatchAllWordForms = False

            find.Wrap = 0

            find.Format = True

            find.Execute(Replace = 1)

            myrange_find = myrange_cau.Duplicate

            find = myrange_find.Find

            find.ClearFormatting()

            find.Replacement.ClearFormatting()

            find.Replacement.Font.Bold = True

            find.Replacement.Font.Color = win32api.RGB(0, 128, 0)

            find.Text = '([\\(])(*)([\\)])'

            find.Replacement.Text = '\\1\\2\\3'

            find.Forward = True

            find.MatchCase = True

            find.MatchWholeWord = False

            find.MatchWildcards = True

            find.MatchSoundsLike = False

            find.MatchAllWordForms = False

            find.Wrap = 0

            find.Format = True

            find.Execute(Replace = 1)

            continue

        range(1, 100)

        return None

    return None

# WARNING: Decompyle incomplete





def dam_mau_chu_thich(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

    else:

        myrange = doc.Range()

    All_range = myrange.Duplicate

    myrange.Select()

    fix_chuthich_cau(word)

    myrange.Select()

    fix_chuthich_vidu(word)

    for i in range(1, 100):

        myrange_cau = myrange.Duplicate

        find = myrange_cau.Find

        find.ClearFormatting()

        find.Text = '(Câu [0-9]{1,2}. )([\\[\\(])(*)([\\]\\)])'

        find.MatchWildcards = True

        if find.Execute():

            myrange_find = myrange_cau.Duplicate

            myrange = doc.Range(myrange_cau.End, myrange.End)

            find = myrange_find.Find

            find.ClearFormatting()

            find.Replacement.ClearFormatting()

            find.Replacement.Font.Bold = True

            find.Replacement.Font.Color = win32api.RGB(253, 11, 252)

            find.Text = '(\\[)(*)(\\])'

            find.Replacement.Text = '\\1\\2\\3'

            find.Forward = True

            find.MatchCase = True

            find.MatchWholeWord = False

            find.MatchWildcards = True

            find.MatchSoundsLike = False

            find.MatchAllWordForms = False

            find.Wrap = 0

            find.Format = True

            find.Execute(Replace = 1)

            myrange_find = myrange_cau.Duplicate

            find = myrange_find.Find

            find.ClearFormatting()

            find.Replacement.ClearFormatting()

            find.Replacement.Font.Bold = True

            find.Replacement.Font.Color = win32api.RGB(0, 128, 0)

            find.Text = '([\\(])(*)([\\)])'

            find.Replacement.Text = '\\1\\2\\3'

            find.Forward = True

            find.MatchCase = True

            find.MatchWholeWord = False

            find.MatchWildcards = True

            find.MatchSoundsLike = False

            find.MatchAllWordForms = False

            find.Wrap = 0

            find.Format = True

            find.Execute(Replace = 1)

            continue

        range(1, 100)

    All_range.Select()

    dam_mau_chu_thich_vidu(word)

    return None

# WARNING: Decompyle incomplete





def BoderVang(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        selection = word.Selection

        selection.Shading.Texture = 0

        selection.Shading.ForegroundPatternColor = -16777216

        selection.Shading.BackgroundPatternColor = win32api.RGB(255, 242, 204)

        border_top = selection.Borders(-1)

        border_top.LineStyle = word.Options.DefaultBorderLineStyle

        border_top.LineWidth = word.Options.DefaultBorderLineWidth

        border_top.Color = word.Options.DefaultBorderColor

        border_left = selection.Borders(-2)

        border_left.LineStyle = word.Options.DefaultBorderLineStyle

        border_left.LineWidth = word.Options.DefaultBorderLineWidth

        border_left.Color = word.Options.DefaultBorderColor

        border_bottom = selection.Borders(-3)

        border_bottom.LineStyle = word.Options.DefaultBorderLineStyle

        border_bottom.LineWidth = word.Options.DefaultBorderLineWidth

        border_bottom.Color = word.Options.DefaultBorderColor

        border_right = selection.Borders(-4)

        border_right.LineStyle = word.Options.DefaultBorderLineStyle

        border_right.LineWidth = word.Options.DefaultBorderLineWidth

        border_right.Color = word.Options.DefaultBorderColor

        paragraph_format = word.Selection.ParagraphFormat

        paragraph_format.LeftIndent = InchesToPoints(0.06)

        paragraph_format.RightIndent = InchesToPoints(0.06)

        paragraph_format.FirstLineIndent = InchesToPoints(0)

        return None

    messagebox.showinfo('Thông báo', 'Chưa chọn vùng')

    return None

# WARNING: Decompyle incomplete





def chinhlai_boder_all(word):

    doc = word.ActiveDocument

    for paragraph in doc.Paragraphs:

        if paragraph.Range.Shading.BackgroundPatternColor != -16777216:

            paragraph.Range.Select()

            paragraph_format = word.Selection.ParagraphFormat

            paragraph_format.LeftIndent = InchesToPoints(0.06)

            paragraph_format.RightIndent = InchesToPoints(0.06)

            paragraph_format.FirstLineIndent = InchesToPoints(0)

    continue

    return None

# WARNING: Decompyle incomplete





def xoa_BoderAndShading(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        selection = word.Selection

        selection.Shading.Texture = 0

        selection.Shading.ForegroundPatternColor = -16777216

        selection.Shading.BackgroundPatternColor = -16777216

        border_top = selection.Borders(-1)

        border_top.LineStyle = 0

        border_left = selection.Borders(-2)

        border_left.LineStyle = 0

        border_bottom = selection.Borders(-3)

        border_bottom.LineStyle = 0

        border_right = selection.Borders(-4)

        border_right.LineStyle = 0

        paragraph_format = word.Selection.ParagraphFormat

        paragraph_format.LeftIndent = InchesToPoints(0)

        paragraph_format.RightIndent = InchesToPoints(0)

        return None

    messagebox.showinfo('Thông báo', 'Chưa chọn vùng')

    return None

# WARNING: Decompyle incomplete





def dam_xanh_loi_giai(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        find = word.Selection.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Replacement.Font.Bold = True

        find.Replacement.Font.Italic = False

        find.Replacement.Font.Color = win32api.RGB(0, 32, 96)

        find.Replacement.ParagraphFormat.Alignment = 1

        find.Text = '(Lời giải)'

        find.Replacement.Text = '\\1'

        find.Forward = True

        find.MatchCase = False

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = True

        find.Execute(Replace = 2)

        return None

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = True

    find.Replacement.Font.Italic = False

    find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

    find.Replacement.ParagraphFormat.Alignment = 1

    find.Text = '(Lời giải)'

    find.Replacement.Text = '\\1'

    find.Forward = True

    find.MatchCase = False

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def chuanhoa_loigiai(word):

    doc = word.ActiveDocument

    pattern = re.compile('\\s*(giải|lời giải|hướng dẫn|hướng dẫn giải)[ \\t]*[.:]*[ \\t]*', re.IGNORECASE)

    paragraphs = list(doc.Paragraphs)

    doc_length = len(paragraphs)

    for i in reversed(range(doc_length)):

        para = paragraphs[i]

        text = para.Range.Text.strip()

        match = pattern.fullmatch(text)

        if not match:

            continue

        matched_text = match.group()

        clean_text = 'Lời giải'

        start = para.Range.Start

        end = para.Range.End - 1

        fix_range = doc.Range(Start = start, End = end)

        fix_range.Text = clean_text

        fix_range.Font.Bold = True

        fix_range.Font.Italic = False

        fix_range.ParagraphFormat.Alignment = 1

        fix_range.Font.Color = win32api.RGB(0, 0, 255)

    return None

# WARNING: Decompyle incomplete





def add_cau_before_phan(word):

    doc = word.ActiveDocument

    thay_the_replace(word, '(^13PHẦN)', '^13Câu 00.@\\1')

    thay_the_replace(word, '(^13Phần)', '^13Câu 00.@\\1')

    return None

# WARNING: Decompyle incomplete





def add_cau_before_Vidu_Bai(word):

    doc = word.ActiveDocument

    thay_the_replace(word, '(^13Ví dụ [0-9]{1,2})', '^13Câu 00.@\\1')

    thay_the_replace(word, '(^13Bài [0-9]{1,2})', '^13Câu 00.@\\1')

    return None

# WARNING: Decompyle incomplete





def add_cau_before_heading(word, heading_idx):

    doc = word.ActiveDocument

    myrange = doc.Range()

    for i in range(300):

        myrange_find = myrange.Duplicate

        find = myrange_find.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Style = doc.Styles(f'''Heading {heading_idx}''')

        find.Text = ''

        find.Replacement.Text = ''

        find.Forward = True

        find.Wrap = 0

        find.Format = True

        find.MatchCase = False

        find.MatchWholeWord = False

        find.MatchWildcards = False

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        if find.Execute():

            word.Selection.SetRange(myrange_find.Start, myrange_find.Start)

            word.Selection.TypeText('Câu 00.@')

            word.Selection.TypeParagraph()

            myrange.SetRange(myrange_find.End, myrange.End)

            continue

        range(300)

        return None

    return None

# WARNING: Decompyle incomplete





def xoa_loi_giai_vidu(word):

    doc = word.ActiveDocument

    thay_the_replace_stop(word, '(Lời giải)(*)(HG\\@)', '\\3')

    return None

# WARNING: Decompyle incomplete





def xoa_loi_giai_cau_vidu_keep_header(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange.Select()

        myrange_find = myrange.Duplicate

        find = myrange_find.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = '(Lời giải)(*)(^13Câu [0-9]{1,2})'

        find.Replacement.Text = '\\3'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        find.Execute(Replace = 2)

        return None

    myrange = doc.Content

    Convert_Auto_To_Text(word)

    add_cau_before_phan(word)

    add_cau_before_header(word)

    add_cau_before_Vidu_Bai(word)

    them_cau_acong_cuoi(word)

    myrange.Select()

    xoa_loi_giai_vidu(word)

    myrange_find = myrange.Duplicate

    find = myrange_find.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '(Lời giải)(*)(^13Câu [0-9]{1,2}[.:])'

    find.Replacement.Text = '\\3'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = False

    find.Execute(Replace = 2)

    xoa_cau_00(word)

    xoa_dong_trang(word)

    return None

# WARNING: Decompyle incomplete





def xoa_loi_giai_new(word):

    doc = word.ActiveDocument

    

    def la_cau(text):

        text = text.strip()

        return re.match('^(Câu|Bài|Ví dụ)\\s*\\d+', text)



    

    def la_HET(text):

        text = text.strip()

        return re.search('HẾT', text)



    

    def la_phan(text):

        text = text.strip()

        return re.match('^PHẦN\\s*[IVX\\d]+', text)



    paras = doc.Paragraphs

    i = 1

    if i <= paras.Count:

        text = paras(i).Range.Text.strip()

        if text.startswith('Lời giải'):

            start = paras(i).Range.Start

            j = i + 1

            if j <= paras.Count:

                t = paras(j).Range.Text.strip()

                if la_cau(t) and la_HET(t) or la_phan(t):

                    pass

                else:

                    j += 1

                    if j <= paras.Count:

                        continue

            if j <= paras.Count:

                end = paras(j).Range.Start

            else:

                end = doc.Content.End

            doc.Range(start, end).Delete()

            paras = doc.Paragraphs

            i = 1

        else:

            i += 1

        if i <= paras.Count:

            continue

        return None

    return None

# WARNING: Decompyle incomplete





def xoa_loi_giai_cau_vidu(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange.Select()

        xoa_loi_giai_vidu(word)

        myrange_find = myrange.Duplicate

        find = myrange_find.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = '(Lời giải)(*)(^13Câu [0-9]{1,2})'

        find.Replacement.Text = '\\3'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        find.Execute(Replace = 2)

        return None

    myrange = doc.Content

    Convert_Auto_To_Text(word)

    add_cau_before_phan(word)

    them_cau_acong_cuoi(word)

    myrange.Select()

    xoa_loi_giai_vidu(word)

    myrange_find = myrange.Duplicate

    find = myrange_find.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '(Lời giải)(*)(^13Câu [0-9]{1,2}[.:])'

    find.Replacement.Text = '\\3'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = False

    find.Execute(Replace = 2)

    xoa_cau_00(word)

    xoa_dong_trang(word)

    return None

# WARNING: Decompyle incomplete





def xoa_loi_giai_tool(root, word):

    pass

# WARNING: Decompyle incomplete





def fix_cau(word):

    doc = word.ActiveDocument

    Convert_Auto_To_Text(word)

    pattern = re.compile('\\s*(câu|bài|question)\\s*(\\d+)[ \\t]*[.:]*[ \\t]*', re.IGNORECASE)

    for para in doc.Paragraphs:

        text = para.Range.Text

        match = pattern.match(text)

        if not match:

            continue

        tu_khoa = match.group(1).capitalize()

        so = match.group(2)

        matched_text = match.group()

        clean_text = f'''{tu_khoa} {so}. '''

        start = para.Range.Start

        end = start + len(matched_text)

        fix_range = doc.Range(Start = start, End = end)

        fix_range.Text = clean_text

        fix_range.Font.Bold = True

        fix_range.Font.Color = win32api.RGB(0, 0, 255)

    return None

# WARNING: Decompyle incomplete





def fix_cau_for_tool(word):

    doc = word.ActiveDocument

    Convert_Auto_To_Text(word)

    pattern = re.compile('\\s*(câu|bài|question)\\s*(\\d+)[ \\t]*[.:]*[ \\t]*', re.IGNORECASE)

    for para in doc.Paragraphs:

        text = para.Range.Text

        match = pattern.match(text)

        if not match:

            continue

        tu_khoa = match.group(1).capitalize()

        so = match.group(2)

        matched_text = match.group()

        clean_text = f'''{tu_khoa} {so}. '''

        start = para.Range.Start

        end = start + len(matched_text)

        fix_range = doc.Range(Start = start, End = end)

        fix_range.Text = clean_text

        fix_range.Font.Bold = True

        fix_range.Font.Size = 12

        fix_range.Font.Name = 'Times New Roman'

        fix_range.Font.Color = win32api.RGB(0, 0, 255)

    return None

# WARNING: Decompyle incomplete





def STT_cau_for_when_need(word, myrange, text, i, symbol):

    if text != '':

        text_find = '(^13' + text + ' )([0-9]{1,})([.:])'

    else:

        text_find = '(^13)([0-9]{1,})([.:])'

    myrange_new = myrange.Duplicate

    for s in range(0, 500):

        myrange_find = myrange_new.Duplicate

        find = myrange_find.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = text_find

        find.Replacement.Text = f'''\\1{i}{symbol}'''

        find.Replacement.Font.Bold = True

        find.Replacement.Font.Underline = 0

        find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

        find.MatchWildcards = True

        find.Format = True

        find.Wrap = 0

        if find.Execute(Replace = 1):

            myrange_new.SetRange(Start = myrange_find.End, End = myrange_new.End)

            if myrange_find.Tables.Count > 0:

                table = myrange_find.Tables(1)

                table_range = table.Range

                table_end = table_range.End

                myrange_new.SetRange(Start = table_end, End = myrange_new.End)

            i += 1

            continue

        range(0, 500)

        return None





def STT_cau_for_new(word, myrange, cau, i, symbol):

    doc = word.ActiveDocument

    STT = i

    if cau != '':

        pattern = re.compile(f'''\\s*{cau}\\s*\\d+[ \\t]*[.:]*[ \\t]*''', re.IGNORECASE)

        paragraphs = list(myrange.Paragraphs)

        for idx, para in enumerate(paragraphs):

            text = para.Range.Text

            if text:

                match = pattern.match(text)

                if match:

                    matched_text = match.group()

                    tu_khoa = cau.capitalize() if cau else ''

                    clean_text = f'''{tu_khoa + ' ' if tu_khoa else ''}{STT}{symbol} '''

                    start = para.Range.Start

                    end = start + len(matched_text)

                    format_range = doc.Range(Start = start, End = end)

                    format_range.Text = clean_text

                    format_range.Font.Bold = True

                    format_range.Font.Color = win32api.RGB(0, 0, 255)

                    STT += 1

        continue

        return None

    pattern = re.compile('\\s*\\d+[ \\t]*[.:]+[ \\t]*')

    paragraphs = list(myrange.Paragraphs)

    for idx, para in enumerate(paragraphs):

        text = para.Range.Text

        if text:

            match = pattern.match(text)

            if match:

                matched_text = match.group()

                clean_text = f'''{STT}{symbol} '''

                start = para.Range.Start

                end = start + len(matched_text)

                format_range = doc.Range(Start = start, End = end)

                format_range.Text = clean_text

                format_range.Font.Bold = True

                format_range.Font.Italic = False

                format_range.Font.Color = win32api.RGB(0, 0, 255)

                STT += 1

    continue

    return None

# WARNING: Decompyle incomplete





def STT_cau_for_tuy_chon_new(myrange, cau, i, symbol):

    pass

# WARNING: Decompyle incomplete





def STT_tuy_chon_khac(root, word):

    pass

# WARNING: Decompyle incomplete





def STT_cau_new(root, word):

    pass

# WARNING: Decompyle incomplete





def STT_cau_new_for_multi(word):

    doc = word.ActiveDocument

    Convert_Auto_To_Text(word)

    myrange = doc.Range()

    STT_cau_for_new(word, myrange, 'Câu', 1, '.')

    STT_cau_for_new(word, myrange, 'Bài', 1, ':')

    STT_cau_for_new(word, myrange, 'Question', 1, '.')

    STT_cau_for_new(word, myrange, 'Ví dụ', 1, ':')

    return None

# WARNING: Decompyle incomplete





def STT_Cau_Auto(word, cau, symbol1, symbol2):

    doc = word.ActiveDocument

    if cau != '':

        pattern = re.compile(f'''\\s*{cau}\\s*\\d+[ \\t]*[.:]*[ \\t]*''', re.IGNORECASE)

        for para in doc.Paragraphs:

            text = para.Range.Text

            match = pattern.match(text)

            if not match:

                continue

            matched_text = match.group()

            clean_text = '# '

            start = para.Range.Start

            end = start + len(matched_text)

            format_range = doc.Range(Start = start, End = end)

            format_range.Text = clean_text

    else:

        pattern = re.compile('\\s*\\d+\\s*[.:]+[ \\t]*')

        for para in doc.Paragraphs:

            text = para.Range.Text

            match = pattern.match(text)

            if not match:

                continue

            matched_text = match.group()

            clean_text = '# '

            start = para.Range.Start

            end = start + len(matched_text)

            format_range = doc.Range(Start = start, End = end)

            format_range.Text = clean_text

    danhsach = doc.Content

    if danhsach.Find.Execute(FindText = '#', Forward = True):

        danhsach.Select()

        list_template = word.ListGalleries(win32com.client.constants.wdNumberGallery).ListTemplates(1)

        list_level = list_template.ListLevels(1)

        if cau != '':

            if symbol2 != '':

                list_level.NumberFormat = f'''{cau} %1{symbol2}'''

            else:

                list_level.NumberFormat = f'''{cau} %1'''

        elif symbol2 != '':

            list_level.NumberFormat = f'''%1{symbol2}'''

        else:

            list_level.NumberFormat = '%1'

        list_level.TrailingCharacter = win32com.client.constants.wdTrailingSpace

        list_level.NumberStyle = win32com.client.constants.wdListNumberStyleArabic

        list_level.NumberPosition = CentimetersToPoints(0)

        list_level.Alignment = win32com.client.constants.wdListLevelAlignLeft

        list_level.TextPosition = CentimetersToPoints(0)

        list_level.TabPosition = win32com.client.constants.wdUndefined

        list_level.ResetOnHigher = 0

        list_level.StartAt = 1

        list_level.LinkedStyle = ''

        list_level.Font.Bold = True

        list_level.Font.Color = win32com.client.constants.wdColorBlue

        word.Selection.Range.ListFormat.ApplyListTemplateWithLevel(ListTemplate = list_template, ContinuePreviousList = True, ApplyTo = win32com.client.constants.wdListApplyToWholeList, DefaultListBehavior = win32com.client.constants.wdWord10ListBehavior)

        word.Selection.Delete(Unit = win32com.client.constants.wdCharacter, Count = 1)

        if danhsach.Find.Execute(FindText = '#', Forward = True):

            continue

    word.Selection.HomeKey(Unit = win32com.client.constants.wdStory)

    return None

# WARNING: Decompyle incomplete





def chuyen_doi_STT(root, word):

    pass

# WARNING: Decompyle incomplete





def STT_2025(word):

    pass

# WARNING: Decompyle incomplete





def tim_phan_tiep(doc, start_pos):

    rng = doc.Range(Start = start_pos, End = doc.Content.End)

    find = rng.Find

    find.ClearFormatting()

    find.Text = 'PHẦN [I1-4]'

    find.MatchWildcards = True

    find.Forward = True

    find.Wrap = 0

    find.MatchCase = False

    if find.Execute():

        return rng.Start





def STT_2025_new(word):

    doc = word.ActiveDocument

    vbf.Convert_Auto_To_Text(word)

    start = 0

    end = tim_phan_tiep(doc, start)

# WARNING: Decompyle incomplete





def vbf_tab_btp_2cot(word):

    doc = word.ActiveDocument

    selection = word.Selection

    selection.ParagraphFormat.TabStops.ClearAll()

    doc.DefaultTabStop = InchesToPoints(0.04)

    wdAlignTabLeft = 0

    wdTabLeaderSpaces = 0

    tab_stops = [

        0.1,

        0.98,

        1.86,

        2.74]

    for pos in tab_stops:

        selection.ParagraphFormat.TabStops.Add(Position = InchesToPoints(pos), Alignment = wdAlignTabLeft, Leader = wdTabLeaderSpaces)

    return None

# WARNING: Decompyle incomplete





def vbf_tab_btp(word):

    doc = word.ActiveDocument

    columns = doc.PageSetup.TextColumns

    a = columns.Count

    if a == 1:

        selection = word.Selection

        selection.ParagraphFormat.TabStops.ClearAll()

        doc.DefaultTabStop = InchesToPoints(0.2)

        wdAlignTabLeft = 0

        wdTabLeaderSpaces = 0

        tab_stops = [

            0.21,

            2.07,

            3.93,

            5.79]

        for pos in tab_stops:

            selection.ParagraphFormat.TabStops.Add(Position = InchesToPoints(pos), Alignment = wdAlignTabLeft, Leader = wdTabLeaderSpaces)

    if a == 2:

        vbf_tab_btp_2cot(word)

        return None

    return None

# WARNING: Decompyle incomplete





def font_time_12(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        rng = word.Selection.Range

    else:

        rng = doc.Content

    rng.Font.Name = 'Times New Roman'

    rng.Font.Size = 12

    rng.ParagraphFormat.LineSpacing = LinesToPoints(1.15)

    rng.ParagraphFormat.SpaceBefore = 0

    rng.ParagraphFormat.SpaceAfter = 0

    rng.ParagraphFormat.SpaceBeforeAuto = False

    rng.ParagraphFormat.SpaceAfterAuto = False

    rng.ParagraphFormat.LineUnitBefore = 0

    rng.ParagraphFormat.LineUnitAfter = 0

    return None

# WARNING: Decompyle incomplete





def bo_gach_ABCD_cham(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange = mo_rong_vung_chon(doc, myrange)

        myrange_find = myrange.Duplicate

        find = myrange_find.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Replacement.Font.Bold = True

        find.Replacement.Font.Underline = 0

        find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

        find.Text = '([^13^9])([ABCD].[ ]{1,})'

        find.Replacement.Text = '\\1\\2'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = True

        find.Execute(Replace = 2)

        find = myrange_find.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Replacement.Font.Bold = True

        find.Replacement.Font.Underline = 0

        find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

        find.Text = '([^13^9])([ABCD].)'

        find.Replacement.Text = '\\1\\2'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = True

        find.Execute(Replace = 2)

        return None

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = True

    find.Replacement.Font.Underline = 0

    find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

    find.Text = '([^13^9])([ABCD].[ ]{1,})'

    find.Replacement.Text = '\\1\\2'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = True

    find.Replacement.Font.Underline = 0

    find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

    find.Text = '([^13^9])([ABCD].)'

    find.Replacement.Text = '\\1\\2'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def bo_gach_abcd_ngoac_tool(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange = mo_rong_vung_chon(doc, myrange)

        find = myrange.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Replacement.Font.Bold = True

        find.Replacement.Font.Underline = 0

        find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

        find.Text = '([^13^9])([abcd])([\\)])'

        find.Replacement.Text = '\\1\\2\\3'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = True

        find.Execute(Replace = 2)

        return None

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = True

    find.Replacement.Font.Underline = 0

    find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

    find.Text = '([^13^9])([abcd])([\\)])'

    find.Replacement.Text = '\\1\\2\\3'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def bo_Unline_Hightlight_Red(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        rng = word.Selection.Range

        rng.Select()

        rng.Shading.Texture = 0

        rng.Shading.ForegroundPatternColor = -16777216

        rng.Shading.BackgroundPatternColor = -16777216

        rng.HighlightColorIndex = 0

        rng.Select()

        bo_gach_ABCD_cham(word)

        rng.Select()

        bo_gach_abcd_ngoac_tool(word)

        return None

    rng = doc.Content

    rng.Select()

    add_blank_line_after_table(word)

    rng.Shading.Texture = 0

    rng.Shading.ForegroundPatternColor = -16777216

    rng.Shading.BackgroundPatternColor = -16777216

    rng.HighlightColorIndex = 0

    rng.Select()

    bo_gach_ABCD_cham(word)

    rng.Select()

    bo_gach_abcd_ngoac_tool(word)

    xoa_dong_trang(word)

    return None

# WARNING: Decompyle incomplete





def bo_mau_tung_kitu(word):

    doc = word.ActiveDocument

    selection = word.Selection

    if word.Selection.Type != 1:

        for rng in selection.Characters:

            char = rng

            char.Font.Shading.BackgroundPatternColor = 16777215

        return None

    messagebox.showinfo('Thông báo', 'Chưa có Vùng được chọn')

    return None

# WARNING: Decompyle incomplete





def fix_chon_ABCD(word):

    doc = word.ActiveDocument

    thay_the_replace(word, '(Chọn)([ ^9]{1,})', '\\1 ')

    thay_the_replace(word, '(Chọn)([ABCD])', '\\1 \\2')

    thay_the_replace(word, '(Chọn [ABCD])(.)', '\\1')

    thay_the_replace(word, '([!^13])(Chọn [ABCD])', '\\1^13\\2')

    thay_the_replace(word, '(Chọn [ABCD])([!^13])', '\\1^13\\2')

    return None

# WARNING: Decompyle incomplete





def Highlight_Chon_ABCD(word):

    doc = word.ActiveDocument

    myrange = doc.Content

    for i in range(50):

        find_range = myrange.Duplicate

        find = find_range.Find

        find.ClearFormatting()

        find.Text = 'Chọn [A-D]'

        find.MatchWildcards = True

        if not find.Execute():

            continue

        myrange.SetRange(Start = find_range.End + 1, End = doc.Content.End)

        find_range.HighlightColorIndex = 4

    return None

# WARNING: Decompyle incomplete





def Select_mau_canh_giua(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange.Font.Color = win32api.RGB(0, 32, 96)

        myrange.Font.Bold = True

        myrange.ParagraphFormat.Alignment = 1

        return None

    find = word.Selection.Find

    find.ClearFormatting()

    find.ParagraphFormat.Alignment = 1

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Color = win32api.RGB(0, 32, 96)

    find.Text = ''

    find.Replacement.Text = ''

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def vung_canh_giua(doc, word):

    find = word.Selection.Find

    find.ClearFormatting()

    find.Font.Color = win32api.RGB(0, 32, 96)

    find.Replacement.ClearFormatting()

    find.Replacement.ParagraphFormat.Alignment = 1

    find.Text = ''

    find.Replacement.Text = ''

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)





def Select_mau_canh_phai(word):

    doc = word.ActiveDocument

    if word.Selection.Type != win32com.client.constants.wdSelectionIP:

        myrange = word.Selection.Range

        myrange.Font.Color = win32api.RGB(0, 0, 30)

        myrange.ParagraphFormat.Alignment = 2

        return None

    find = word.Selection.Find

    find.ClearFormatting()

    find.ParagraphFormat.Alignment = 2

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Color = win32api.RGB(0, 0, 30)

    find.Text = ''

    find.Replacement.Text = ''

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def vung_canh_phai(doc, word):

    find = word.Selection.Find

    find.ClearFormatting()

    find.Font.Color = win32api.RGB(0, 0, 30)

    find.Replacement.ClearFormatting()

    find.Replacement.ParagraphFormat.Alignment = 2

    find.Text = ''

    find.Replacement.Text = ''

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = False

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)





def check_pic_float(doc, word):

    dem_so = 0

    shapes_count = doc.Shapes.Count

    for i in range(shapes_count, 0, -1):

        shape = doc.Shapes(i)

        if not shape.Type != 9:

            continue

        dem_so += 1

    return dem_so

# WARNING: Decompyle incomplete





def pic_inline(doc, word):

    shapes_count = doc.Shapes.Count

    for i in range(shapes_count, 0, -1):

        shape = doc.Shapes(i)

        if not shape.Type != 9:

            continue

        shape.ConvertToInlineShape()

    return None

# WARNING: Decompyle incomplete





def canh_pict_giua_OLD(doc, word):

    for inline_shape in doc.InlineShapes:

        paragraph = inline_shape.Range.Paragraphs(1)

        paragraph_text = paragraph.Range.Text.strip()

        if not len(paragraph_text) == 1:

            continue

        inline_shape.Select()

        word.Selection.ParagraphFormat.Alignment = 1





def canh_pict_giua(doc, word):

    for inline_shape in doc.InlineShapes:

        if not inline_shape.Type != 1:

            continue

        paragraph = inline_shape.Range.Paragraphs(1)

        paragraph_text = paragraph.Range.Text.strip()

        if not len(paragraph_text) == 1:

            continue

        paragraph.Alignment = 1





def pic_inline_center(word):

    doc = word.ActiveDocument

    pic_inline(doc, word)

    canh_pict_giua(doc, word)

    return None

# WARNING: Decompyle incomplete





def pic_inline_center_Select(word):

    doc = word.ActiveDocument

    if word.Selection.Type != win32com.client.constants.wdSelectionIP:

        myrange = word.Selection.Range

    else:

        myrange = doc.Range()

    for inline_shape in myrange.InlineShapes:

        if not inline_shape.Type in (3, 4):

            continue

        paragraph = inline_shape.Range.Paragraphs(1)

        paragraph_text = paragraph.Range.Text.strip()

        if not len(paragraph_text) == 1:

            continue

        paragraph.Alignment = 1

    return None

# WARNING: Decompyle incomplete





def pic_inline_center_tool(root, word):

    pass

# WARNING: Decompyle incomplete





def resize_images_to_column_width_Mix(doc, word):

    for iShp in doc.InlineShapes:

        if not iShp.Width > InchesToPoints(3.5):

            continue

        iShp.LockAspectRatio = True

        iShp.Width = InchesToPoints(3.5)





def resize_images_to_column_width(word):

    doc = word.ActiveDocument

    resize_images_to_column_width_Mix(doc, word)

    return None

# WARNING: Decompyle incomplete





def delete_ink_drawings(word):

    doc = word.ActiveDocument

    for i in range(doc.Shapes.Count, 0, -1):

        shp = doc.Shapes(i)

        if not shp.Type in (22, 23):

            continue

        shp.Delete()

    for i in range(doc.InlineShapes.Count, 0, -1):

        shp = doc.InlineShapes(i)

        if not shp.Type in (22, 23):

            continue

        shp.Delete()

    return None

# WARNING: Decompyle incomplete





def lech_chu_position(word):

    doc = word.ActiveDocument

    selection = word.Selection

    if word.Selection.Type != 1:

        rng = word.Selection.Range

        rng.Font.Position = 0

        return None

    messagebox.showinfo('Thông báo', 'Chưa có Vùng được chọn')

    return None

# WARNING: Decompyle incomplete





def lech_pic(word):

    doc = word.ActiveDocument

    selection = word.Selection

# WARNING: Decompyle incomplete





def lech_pic_not(word):

    doc = word.ActiveDocument

    selection = word.Selection

    if selection.Type == 1:

        messagebox.showinfo('Thông báo', 'Chưa có Vùng được chọn')

        return None

    rng = selection.Range

    rng.Font.Position = 0

    for inline_shape in rng.InlineShapes:

        if not inline_shape.Type in (3, 4):

            continue

        height = inline_shape.Height

        font = rng.Font

        font.Name = '+Body'

        font.Size = 11

        font.Bold = False

        font.Italic = False

        font.Underline = 0

        font.Color = 0

        font.Position = -height / 2 + 3.5 if height >= 14.4 else 3

    return None

# WARNING: Decompyle incomplete





def macro_exists(word, macro_name):

    vb_project = word.Application.VBE.ActiveVBProject

    for component in vb_project.VBComponents:

        code_module = component.CodeModule

        for line_num in range(1, code_module.CountOfLines + 1):

            line = code_module.Lines(line_num, 1)

            if not macro_name in line:

                continue

            range(1, code_module.CountOfLines + 1)

            vb_project.VBComponents

            return True

    return False





def lech_pos_pic_math2(word):

    doc = word.ActiveDocument

    selection = word.Selection

    if word.Selection.Type != 1:

        rng = word.Selection.Range

        rng.Font.Position = 0

        rng.Select()

        word.Application.Run('MTCommand_TeXToggle')

        rng.Select()

        thay_the_replace_stop_false(word, '\\[', '$')

        rng.Select()

        thay_the_replace_stop_false(word, '\\]', '$')

        rng.Select()

        thay_the_replace_stop(word, '(\\$)(*)(\\$)', '\\1{\\2}\\3')

        rng.Select()

        word.Application.Run('MTCommand_TeXToggle')

        return None

    messagebox.showinfo('Thông báo', 'Chưa có Vùng được chọn')

    return None

# WARNING: Decompyle incomplete





def lech_pos_pic_math(word):

    doc = word.ActiveDocument

    selection = word.Selection

    if word.Selection.Type != 1:

        rng = word.Selection.Range

        rng.Font.Position = 0

        rng.Select()

        mathtype_to_latex(word, mode = 'select', thongbao = 'no')

        rng.Select()

        latex_to_mathtype(word, mode = 'select', thongbao = 'no')

        return None

    messagebox.showinfo('Thông báo', 'Chưa có Vùng được chọn')

    return None

# WARNING: Decompyle incomplete





def lowered_chemdraw_in_word(word):

    doc = word.ActiveDocument

    selection = word.Selection

    if selection.Type == 1:

        messagebox.showinfo('Thông báo', 'Chưa có Vùng được chọn')

        return None

    rng = selection.Range

    rng.Font.Position = 0

    for shape in rng.InlineShapes:

        if not shape.Type == 1:

            continue

        progid = shape.OLEFormat.ProgID

        if progid and 'chemdraw' in progid.lower():

            shape.Range.Font.Position = -7

    continue

    return None

# WARNING: Decompyle incomplete





def Ba_dong_at(word):

    doc = word.ActiveDocument

    word.Selection.HomeKey(Unit = 5)

    daucham = '……………………………………………………………………………………'

    word.Selection.Font.Bold = False

    word.Selection.Font.Italic = False

    word.Selection.Font.Name = 'Times New Roman'

    word.Selection.Font.Size = 8

    word.Selection.Font.Color = win32api.RGB(0, 128, 0)

    word.Selection.ParagraphFormat.SpaceBefore = 8

    word.Selection.ParagraphFormat.SpaceAfter = 0

    word.Selection.ParagraphFormat.LineSpacing = LinesToPoints(1.15)

    word.Selection.ParagraphFormat.Alignment = 3

    word.Selection.TypeText(f'''{daucham}\r\n{daucham}\r\n{daucham}\r\n''')

    return None

# WARNING: Decompyle incomplete





def Ba_dong_all_new(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        result = messagebox.askokcancel('Xác nhận', 'Bạn đang đánh dấu 1 vùng văn bản, nên công việc chỉ thực hiện trên vùng đó, Bạn có muốn tiếp tục với đoạn đã chọn không?')

        if not result:

            return None

        Ba_dong_all(word, myrange)

    else:

        them_cau_acong_cuoi(word)

        myrange = doc.Range()

        thay_the_replace(word, '(^13)(PHẦN I)', '\\1Câu 00.@^13\\2')

        thay_the_replace(word, '(^13)(Phần)', '\\1PHẦN')

        myrange_find = myrange.Duplicate

        find = myrange_find.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = '(PHẦN [I1234])'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        if find.Execute():

            for i in range(10):

                myrange.SetRange(Start = myrange_find.Start + 1, End = myrange.End)

                myrange_find = myrange.Duplicate

                find = myrange_find.Find

                find.ClearFormatting()

                find.Replacement.ClearFormatting()

                find.Text = '(PHẦN [I1234])'

                find.Forward = True

                find.MatchCase = True

                find.MatchWholeWord = False

                find.MatchWildcards = True

                find.MatchSoundsLike = False

                find.MatchAllWordForms = False

                find.Wrap = 0

                find.Format = False

                myrange_cau = myrange_find.Duplicate

                if find.Execute():

                    myrange_cau.SetRange(Start = myrange.Start, End = myrange_find.Start)

                    Ba_dong_all(word, myrange_cau)

                    continue

                Ba_dong_all(word, myrange_cau)

                range(10)

        else:

            Ba_dong_all(word, myrange)

        xoa_cau_00(word)

    dam_xanh_cau_font12(word)

    xoa_dong_trang(word)

    return None

# WARNING: Decompyle incomplete





def Ba_dong_all(word, myrange):

    doc = word.ActiveDocument

    word.Selection.HomeKey(Unit = 6)

    myrange_find = myrange.Duplicate

    find = myrange_find.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '(^13Câu [0-9]{1,3}.)'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Format = False

    if find.Execute():

        myrange.SetRange(Start = myrange_find.End, End = myrange.End)

        daucham = '^13……………………………………………………………………………………'

        myrange_find = myrange.Duplicate

        find = myrange_find.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Replacement.Font.Bold = False

        find.Replacement.Font.Name = 'Times New Roman'

        find.Replacement.Font.Size = 8

        find.Replacement.ParagraphFormat.SpaceBefore = 8

        find.Replacement.ParagraphFormat.SpaceAfter = 0

        find.Replacement.ParagraphFormat.LineSpacing = LinesToPoints(1.15)

        find.Replacement.ParagraphFormat.Alignment = 3

        find.Replacement.Font.Color = win32api.RGB(0, 128, 0)

        find.Text = '(^13Câu [0-9]{1,3})'

        find.Replacement.Text = f'''{daucham}{daucham}{daucham}\\1'''

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = True

        find.Execute(Replace = 2)

        return None

    return None

# WARNING: Decompyle incomplete





def fix_format_dong_cham(word):

    doc = word.ActiveDocument

    daucham = '[…]{1,}'

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = False

    find.Replacement.Font.Name = 'Times New Roman'

    find.Replacement.Font.Size = 8

    find.Replacement.ParagraphFormat.SpaceBefore = 8

    find.Replacement.ParagraphFormat.SpaceAfter = 0

    find.Replacement.ParagraphFormat.LineSpacing = LinesToPoints(1.15)

    find.Replacement.ParagraphFormat.Alignment = 3

    find.Replacement.Font.Color = win32api.RGB(0, 128, 0)

    find.Text = f'''({daucham})'''

    find.Replacement.Text = '\\1'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def Ba_dong_at_dai(word):

    doc = word.ActiveDocument

    word.Selection.HomeKey(Unit = 5)

    daucham = '……………………………………………………………………………………………………………………………………………………………………………'

    word.Selection.Font.Bold = False

    word.Selection.Font.Italic = False

    word.Selection.Font.Name = 'Times New Roman'

    word.Selection.Font.Size = 8

    word.Selection.Font.Color = win32api.RGB(0, 128, 0)

    word.Selection.ParagraphFormat.SpaceBefore = 8

    word.Selection.ParagraphFormat.SpaceAfter = 0

    word.Selection.ParagraphFormat.LineSpacing = LinesToPoints(1.15)

    word.Selection.ParagraphFormat.Alignment = 3

    word.Selection.TypeText(f'''{daucham}\r\n{daucham}\r\n{daucham}\r\n''')

    return None

# WARNING: Decompyle incomplete





def Ba_dong_all_dai_new(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        result = messagebox.askokcancel('Xác nhận', 'Bạn đang đánh dấu 1 vùng văn bản, nên công việc chỉ thực hiện trên vùng đó, Bạn có muốn tiếp tục với đoạn đã chọn không?')

        if not result:

            return None

        Ba_dong_all_dai(word, myrange)

    else:

        them_cau_acong_cuoi(word)

        myrange = doc.Range()

        thay_the_replace(word, '(^13)(PHẦN I)', '\\1Câu 00.@^13\\2')

        myrange_find = myrange.Duplicate

        find = myrange_find.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = '(PHẦN I)'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        if find.Execute():

            for i in range(10):

                myrange.SetRange(Start = myrange_find.Start + 1, End = myrange.End)

                myrange_find = myrange.Duplicate

                find = myrange_find.Find

                find.ClearFormatting()

                find.Replacement.ClearFormatting()

                find.Text = '(PHẦN I)'

                find.Forward = True

                find.MatchCase = True

                find.MatchWholeWord = False

                find.MatchWildcards = True

                find.MatchSoundsLike = False

                find.MatchAllWordForms = False

                find.Wrap = 0

                find.Format = False

                myrange_cau = myrange_find.Duplicate

                if find.Execute():

                    myrange_cau.SetRange(Start = myrange.Start, End = myrange_find.Start)

                    Ba_dong_all_dai(word, myrange_cau)

                    continue

                Ba_dong_all_dai(word, myrange_cau)

                range(10)

        else:

            Ba_dong_all_dai(word, myrange)

        xoa_cau_00(word)

    dam_xanh_cau_font12(word)

    xoa_dong_trang(word)

    return None

# WARNING: Decompyle incomplete





def Ba_dong_all_dai(word, myrange):

    doc = word.ActiveDocument

    them_cau_acong_cuoi(word)

    word.Selection.HomeKey(Unit = 6)

    myrange_find = myrange.Duplicate

    find = myrange_find.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Text = '(^13Câu [0-9]{1,3}.)'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 0

    find.Format = False

    if find.Execute():

        myrange.SetRange(Start = myrange_find.End, End = myrange.End)

        daucham = '^13……………………………………………………………………………………………………………………………………………………………………………'

        myrange_find = myrange.Duplicate

        find = myrange_find.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Replacement.Font.Bold = False

        find.Replacement.Font.Name = 'Times New Roman'

        find.Replacement.Font.Size = 8

        find.Replacement.ParagraphFormat.SpaceBefore = 8

        find.Replacement.ParagraphFormat.SpaceAfter = 0

        find.Replacement.ParagraphFormat.LineSpacing = LinesToPoints(1.15)

        find.Replacement.Font.Color = win32api.RGB(0, 128, 0)

        find.Replacement.ParagraphFormat.Alignment = 3

        find.Text = '(^13Câu [0-9]{1,3})'

        find.Replacement.Text = f'''{daucham}{daucham}{daucham}\\1'''

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = True

        find.Execute(Replace = 2)

        return None

    return None

# WARNING: Decompyle incomplete





def ba_dong_cham_Vidu(word):

    doc = word.ActiveDocument

    myrange = doc.Content

    find = myrange.Find

    myrange = doc.Range()

    daucham = '^13……………………………………………………………………………………'

    find = myrange.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = False

    find.Replacement.Font.Name = 'Times New Roman'

    find.Replacement.Font.Size = 8

    find.Replacement.ParagraphFormat.SpaceBefore = 8

    find.Replacement.ParagraphFormat.SpaceAfter = 0

    find.Replacement.ParagraphFormat.LineSpacing = LinesToPoints(1.15)

    find.Replacement.Font.Color = win32api.RGB(0, 128, 0)

    find.Text = '(^13[\\(]Dòng chấm \\@)'

    find.Replacement.Text = f'''{daucham}{daucham}{daucham}\\1'''

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def ba_dong_dai_Vidu(word):

    doc = word.ActiveDocument

    myrange = doc.Content

    find = myrange.Find

    myrange = doc.Range()

    daucham = '^13……………………………………………………………………………………………………………………………………………………………………………'

    find = myrange.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = False

    find.Replacement.Font.Name = 'Times New Roman'

    find.Replacement.Font.Size = 8

    find.Replacement.ParagraphFormat.SpaceBefore = 8

    find.Replacement.ParagraphFormat.SpaceAfter = 0

    find.Replacement.ParagraphFormat.LineSpacing = LinesToPoints(1.15)

    find.Replacement.Font.Color = win32api.RGB(0, 128, 0)

    find.Text = '(^13[\\(]Dòng chấm \\@)'

    find.Replacement.Text = f'''{daucham}{daucham}{daucham}\\1'''

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def xoa_dong_cham(word):

    doc = word.ActiveDocument

    thay_the_replace(word, '([…]{2,})', '')

    thay_the_replace(word, '([.]{10,})', '')

    thay_the_replace(word, '^13.', '^13')

    thay_the_replace(word, '^13 ', '^13')

    xoa_dong_trang(word)

    return None

# WARNING: Decompyle incomplete





def fix_dong_cham_rng(rng):

    pass

# WARNING: Decompyle incomplete





def fix_dong_cham_gian_format_All(word):

    doc = word.ActiveDocument

    if word.Selection.Type != win32com.client.constants.wdSelectionIP:

        myrange = word.Selection.Range

        myrange = mo_rong_vung_chon(doc, myrange)

    else:

        myrange = doc.Range()

    for i in range(0, 500):

        rng = myrange.Duplicate

        find = rng.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = '[.…]{5,}'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        if find.Execute():

            fix_dong_cham_rng(rng)

            myrange.SetRange(Start = rng.End, End = myrange.End)

            continue

        range(0, 500)

        return None

    return None

# WARNING: Decompyle incomplete





def dong_cham_tool(root, word):

    pass

# WARNING: Decompyle incomplete





def them_dap_so_cham(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        myrange = mo_rong_vung_chon(doc, myrange)

        insert_range = myrange.Duplicate

        insert_range.Collapse(0)

        find = myrange.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Replacement.Font.Bold = True

        word.Selection.Font.Name = 'Times New Roman'

        find.Replacement.Font.Size = 12

        word.Selection.ParagraphFormat.SpaceBefore = 0

        word.Selection.ParagraphFormat.SpaceAfter = 0

        word.Selection.ParagraphFormat.LineSpacing = LinesToPoints(1.15)

        find.Text = '(^13Câu)'

        find.Replacement.Text = '^13Trả lời: …………\\1'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        find.Execute(Replace = 2)

        myrange.Select()

        return None

    messagebox.showinfo('Thông báo', 'Chưa có Vùng được chọn Câu TLN')

    return None

# WARNING: Decompyle incomplete





def save_as_to_name(root, word):

    pass

# WARNING: Decompyle incomplete





def save_convert_sendpath_docx(word, entry_input):

    doc = word.ActiveDocument

    doc.Save()

    file_path = doc.FullName

    file_path = chuan_hoa_duong_dan(file_path)

    (file_name, file_extension) = os.path.splitext(file_path)

    if file_extension.lower() != '.docx':

        new_file_name = file_name + '.docx'

        doc.SaveAs(FileName = new_file_name, FileFormat = 12)

        doc.Convert()

        doc.Save()

    doc = word.ActiveDocument

    file_path = doc.FullName

    file_path = chuan_hoa_duong_dan(file_path)

    entry_input.delete(0, 'end')

    entry_input.insert(0, file_path)

    doc.Close(SaveChanges = True)

    if word.Documents.Count == 0:

        word.Quit()

        return None

    return None

# WARNING: Decompyle incomplete





def Chu_ba_dong_cham_acong(word):

    doc = word.ActiveDocument

    word.Selection.Font.Bold = False

    word.Selection.Font.Italic = False

    word.Selection.Font.Name = 'Times New Roman'

    word.Selection.Font.Size = 12

    word.Selection.Font.Color = win32api.RGB(0, 128, 0)

    word.Selection.ParagraphFormat.SpaceBefore = 8

    word.Selection.ParagraphFormat.SpaceAfter = 0

    word.Selection.TypeText('(Dòng chấm @:)')

    word.Selection.TypeParagraph()

    return None

# WARNING: Decompyle incomplete





def End_giai_acong(word):

    doc = word.ActiveDocument

    word.Selection.Font.Bold = True

    word.Selection.Font.Italic = False

    word.Selection.Font.Name = 'Times New Roman'

    word.Selection.Font.Size = 12

    word.Selection.Font.Color = win32api.RGB(0, 128, 0)

    word.Selection.ParagraphFormat.SpaceBefore = 8

    word.Selection.ParagraphFormat.SpaceAfter = 0

    word.Selection.TypeText('HG@:')

    word.Selection.TypeParagraph()

    return None

# WARNING: Decompyle incomplete





def fix_cauSGK(word):

    if word.Selection.Type != 1:

        thay_the_replace_stop(word, '(^13)(Câu )([0-9]{1,3}.)([0-9]{1,3})([.:])', '\\1\\3\\4\\5')

        thay_the_replace_stop(word, '(^13)([0-9]{1,3}.)([0-9]{1,3})([.:])', '\\1Câu 1. (SGK \\2\\3 )')

        return None

    thay_the_replace(word, '(^13)(Câu )([0-9]{1,3}.)([0-9]{1,3})([.:])', '\\1\\3\\4\\5')

    thay_the_replace(word, '(^13)([0-9]{1,3}.)([0-9]{1,3})([.:])', '\\1Câu 1. (SGK \\2\\3 )')





def Ltap_VD_HD_to_Vidu(word):

    if word.Selection.Type != 1:

        thay_the_replace_stop(word, '(VD[0-9])([.:])', 'Ví dụ 1. (\\1)')

        thay_the_replace_stop(word, '(LT[0-9])([.:])', 'Ví dụ 1. (\\1)')

        thay_the_replace_stop(word, '(Luyện tập [0-9])([.:])', 'Ví dụ 1. (\\1)')

        thay_the_replace_stop(word, '(HĐ[0-9])([.:])', 'Ví dụ 1. (\\1)')

        thay_the_replace_stop(word, '(HĐ [0-9])([.:])', 'Ví dụ 1. (\\1)')

        thay_the_replace_stop(word, '(Vận dụng [0-9])([.:])', 'Ví dụ 1. (\\1)')

        thay_the_replace_stop(word, '(Vận dụng)([.:])', 'Ví dụ 1. (\\1)')

        return None

    thay_the_replace(word, '(LT[0-9])([.:])', 'Ví dụ 1. (\\1)')

    thay_the_replace(word, '(Luyện tập [0-9])([.:])', 'Ví dụ 1. (\\1)')

    thay_the_replace(word, '(HĐ[0-9])([.:])', 'Ví dụ 1. (\\1)')

    thay_the_replace(word, '(HĐ [0-9])([.:])', 'Ví dụ 1. (\\1)')

    thay_the_replace(word, '(Vận dụng)([.:])', 'Ví dụ 1. (\\1)')

    thay_the_replace(word, '(Vận dụng [0-9])([.:])', 'Ví dụ 1. (\\1)')





def Chu_y_Nxet_dam_xanh(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        find = word.Selection.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Replacement.Font.Bold = True

        find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

        find.Text = '(Chú ý)([.:])'

        find.Replacement.Text = '\\1\\2^13'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = True

        find.Execute(Replace = 2)

        find.Text = '(Nhận xét)([:.])'

        find.Replacement.Text = '\\1\\2^13'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = True

        find.Execute(Replace = 2)

        return None

    find = word.Selection.Find

    find.ClearFormatting()

    find.Replacement.ClearFormatting()

    find.Replacement.Font.Bold = True

    find.Replacement.Font.Color = win32api.RGB(0, 0, 255)

    find.Text = '(Chú ý)([.:])'

    find.Replacement.Text = '\\1\\2^13'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    find.Text = '(Nhận xét)([:.])'

    find.Replacement.Text = '\\1\\2^13'

    find.Forward = True

    find.MatchCase = True

    find.MatchWholeWord = False

    find.MatchWildcards = True

    find.MatchSoundsLike = False

    find.MatchAllWordForms = False

    find.Wrap = 1

    find.Format = True

    find.Execute(Replace = 2)

    return None

# WARNING: Decompyle incomplete





def xu_li_SGK(root, word):

    pass

# WARNING: Decompyle incomplete





def Page_Left_Right00(word):

    doc = word.ActiveDocument

    selection = word.Selection

    selection.ParagraphFormat.LeftIndent = CentimetersToPoints(0)

    selection.ParagraphFormat.RightIndent = CentimetersToPoints(0)

    return None

# WARNING: Decompyle incomplete





def add_blank_line_at_Home(word):

    doc = word.ActiveDocument

    selection = word.Selection

    selection.HomeKey(Unit = 6)

    if selection.Tables.Count == 0:

        selection.TypeParagraph()

        return None

    if selection.Tables.Count == 1:

        selection.SplitTable()

        return None

    return None

# WARNING: Decompyle incomplete





def xoa_blank_line_at_Home(word):

    doc = word.ActiveDocument

    if doc.Paragraphs.Count > 0:

        first_paragraph = doc.Paragraphs(1)

        if first_paragraph.Range.Text.strip() == '':

            first_paragraph.Range.Delete()

            return None

        return None

    return None

# WARNING: Decompyle incomplete





def add_blank_line_after_table(word):

    doc = word.ActiveDocument

    myrange = doc.Range()

    myrange_find = myrange.Duplicate

    find = myrange_find.Find

    find.ClearFormatting()

    find.Text = 'PHẦN'

    find.MatchCase = False

    if find.Execute():

        myrange.SetRange(Start = myrange_find.Start, End = myrange.End)

    else:

        find.Text = 'Câu [0-9]{1,2}'

        find.MatchWildcards = True

        if find.Execute():

            myrange.SetRange(Start = myrange_find.Start, End = myrange.End)

        else:

            return None

    for i in range(1, myrange.Tables.Count + 1):

        table = myrange.Tables(i)

        table_range = table.Range

        table_end = table_range.End

        doc.Range(Start = table_end, End = table_end).InsertAfter('\r\n')

    return None

# WARNING: Decompyle incomplete





def check_sentences_in_tables(word, messages):

    doc = word.ActiveDocument

    pattern = re.compile('^Câu \\d+[.:]')

    a = doc.Range().Tables.Count

    for i in range(1, a + 1):

        table = doc.Range().Tables(i)

        table_range = table.Range

        para = table_range.Paragraphs

        for p in range(1, para.Count + 1):

            text = para(p).Range.Text.strip()

            if not pattern.match(text):

                continue

            messages.append(f'''Có Câu nằm trong bảng số {i}, nếu nó là 1 câu của đề thi thì phải đưa nó ra ngoài bảng đó nhé!\nCó thể select vùng đó và dùng \'Convert Table to text (F1)\' trong Tool 1.''')

            range(1, para.Count + 1)

            range(1, a + 1)

            return None

    return None

# WARNING: Decompyle incomplete





def dem_so_cau(word):

    doc = word.ActiveDocument

    

    def indices_phan(doc):

        paragraphs = doc.Paragraphs

        pattern = re.compile('^phần [i1234]', re.IGNORECASE)

        indices = []

        for i in range(1, paragraphs.Count + 1):

            para = paragraphs(i)

            text = para.Range.Text.strip()

            if not pattern.match(text):

                continue

            indices.append(i)

        indices.append(paragraphs.Count + 1)

        return indices



    doc.Range().ListFormat.ConvertNumbersToText()

    danh_sach = indices_phan(doc)

    messages = []

    paragraphs = doc.Paragraphs

    pattern = re.compile('^Câu [0-9]{1,}[.:]')

    so_cau_tong = 0

    for i in range(1, paragraphs.Count + 1):

        para = paragraphs(i)

        text = para.Range.Text.strip()

        if not pattern.match(text):

            continue

        so_cau_tong += 1

    messages.append(f'''Số câu tất cả: {so_cau_tong} câu''')

    if len(danh_sach) > 1:

        for k in range(len(danh_sach) - 1):

            so_cau = 0

            for p in range(danh_sach[k], danh_sach[k + 1]):

                para = paragraphs(p)

                text = para.Range.Text.strip()

                if not pattern.match(text):

                    continue

                so_cau += 1

            messages.append(f'''Phần {k + 1} có {so_cau} câu''')

    thong_bao = '\n'.join(messages)

    messagebox.showinfo('Thông báo', thong_bao)

    return None

# WARNING: Decompyle incomplete





def autofit_one_table(word, table):

    table.AutoFitBehavior(1)

    table.Rows.Alignment = 1





def autofit_window_one_table(word, table):

    table.AutoFitBehavior(2)





def autofit_table_tool(word):

    doc = word.ActiveDocument

    if word.Selection.Type == 1:

        messagebox.showinfo('Thông báo', 'Chưa có vùng được chọn')

        return None

    selection = word.Selection

    myrange = word.Selection.Range

    table_count = myrange.Tables.Count

    if table_count > 0:

        for i in range(1, table_count + 1):

            table = myrange.Tables(i)

            autofit_one_table(word, table)

        return None

    return None

# WARNING: Decompyle incomplete





def autofit_window_table_tool(word):

    doc = word.ActiveDocument

    if word.Selection.Type == 1:

        messagebox.showinfo('Thông báo', 'Chưa có vùng được chọn')

        return None

    selection = word.Selection

    myrange = word.Selection.Range

    table_count = myrange.Tables.Count

    if table_count > 0:

        for i in range(1, table_count + 1):

            table = myrange.Tables(i)

            autofit_window_one_table(word, table)

        return None

    return None

# WARNING: Decompyle incomplete





def autofit_content_window_table_tool(word):

    doc = word.ActiveDocument

    if word.Selection.Type == 1:

        messagebox.showinfo('Thông báo', 'Chưa có vùng được chọn')

        return None

    selection = word.Selection

    myrange = word.Selection.Range

    table_count = myrange.Tables.Count

    if table_count > 0:

        for i in range(1, table_count + 1):

            table = myrange.Tables(i)

            autofit_one_table(word, table)

            autofit_window_one_table(word, table)

        return None

    return None

# WARNING: Decompyle incomplete





def gui_autoFit_table(root, word):

    pass

# WARNING: Decompyle incomplete





def autofit_table(word):

    doc = word.ActiveDocument

    rng = doc.Range()

    if rng.Tables.Count > 0:

        for i in range(1, rng.Tables.Count + 1):

            table = rng.Tables(i)

            autofit_one_table(word, table)

        return None

    return None

# WARNING: Decompyle incomplete





def autofit_table_no_first(word):

    doc = word.ActiveDocument

    rng = doc.Range()

    if rng.Tables.Count > 1:

        for i in range(2, rng.Tables.Count + 1):

            table = rng.Tables(i)

            autofit_one_table(word, table)

        return None

    return None

# WARNING: Decompyle incomplete





def Strong_table1(word):

    doc = word.ActiveDocument

    table = doc.Tables(1)

    for row in table.Rows:

        for cell in row.Cells:

            shapes_count = cell.Range.ShapeRange.Count if cell.Range.ShapeRange.Count > 0 else 0

            for i in range(shapes_count, 0, -1):

                shape = cell.Range.ShapeRange(i)

                shape.Delete()

            inline_shapes_count = cell.Range.InlineShapes.Count

            for i in range(inline_shapes_count, 0, -1):

                inline_shape = cell.Range.InlineShapes(i)

                inline_shape.Delete()

    table.ConvertToText(Separator = ' ', NestedTables = False)

    return None

# WARNING: Decompyle incomplete





def boder_TableGrid_select(word, table):

    table.Borders.Enable = True





def fix_equation_select(word):

    doc = word.ActiveDocument

    selection = word.Selection

    selection.Find.ClearFormatting()

    for math_obj in selection.OMaths:

        math_range = math_obj.Range

        math_range.Font.Name = 'Cambria Math'

        math_range.Font.Bold = False

        math_range.Font.Color = win32.constants.wdColorAutomatic

        math_obj.ConvertToMathText()

    return None

# WARNING: Decompyle incomplete





def copy_add_dotm(addin_name):

    appdata = os.getenv('APPDATA')

    if not appdata:

        raise EnvironmentError('Không thể tìm thấy biến môi trường %APPDATA%.')

    startup_path = os.path.join(appdata, 'Microsoft', 'Word', 'STARTUP')

    if not os.path.exists(startup_path):

        os.makedirs(startup_path)

    addin_name_dotm = f'''{addin_name}.dotm'''

    source_file_1 = os.path.join(os.getcwd(), 'file_mau_khac', addin_name_dotm)

    source_file_2 = os.path.join(os.getcwd(), addin_name_dotm)

    if os.path.isfile(source_file_1):

        source_file = source_file_1

    elif os.path.isfile(source_file_2):

        source_file = source_file_2

    else:

        messagebox.showwarning('Thông báo', f'''Không tìm thấy {addin_name_dotm}, vui lòng chọn thủ công.''')

        source_file = filedialog.askopenfilename(title = 'Chọn file Add-in (.dotm)', filetypes = [

            ('Word Add-in', '*.dotm')])

        if not source_file:

            raise Exception('Người dùng chưa chọn file.')

    destination_file = os.path.join(startup_path, addin_name_dotm)

    shutil.copy2(source_file, destination_file)

    messagebox.showinfo('Thông báo', f'''Đã sao chép file {addin_name_dotm} thành công!''')

    return None

# WARNING: Decompyle incomplete





def open_startup_folder():

    appdata = os.getenv('APPDATA')

    if not appdata:

        raise EnvironmentError('Không thể tìm thấy biến môi trường %APPDATA%.')

    startup_path = os.path.join(appdata, 'Microsoft', 'Word', 'STARTUP')

    if not os.path.exists(startup_path):

        os.makedirs(startup_path)

    os.startfile(startup_path)

    return None

# WARNING: Decompyle incomplete





def insert_symbol(word):

    doc = word.ActiveDocument

    selection = word.Selection

    selection.InsertSymbol(Font = 'Webdings', CharacterNumber = -3997, Unicode = True)

    return None

# WARNING: Decompyle incomplete





def chuanhoa_dap_so_py32(word):

    doc = word.ActiveDocument

    pattern = re.compile('\\s*(đáp án|đa|đáp số|đs|trả lời|kết quả|kq)[ \\t]*:[ \\t]*', re.IGNORECASE)

    for para in doc.Paragraphs:

        text = para.Range.Text

        match = pattern.match(text)

        if not match:

            continue

        start = para.Range.Start

        matched_len = len(match.group(0))

        replace_range = doc.Range(Start = start, End = start + matched_len)

        replace_range.Text = 'ĐS:'

        para.Range.Font.Bold = True

        para.Range.Font.Underline = False

        para.Range.Font.Color = win32api.RGB(255, 0, 0)

        para.Range.ParagraphFormat.Alignment = 0

    return None

# WARNING: Decompyle incomplete





def xu_li_nhieu_viec(root, word):

    pass

# WARNING: Decompyle incomplete





def check_pages_count(doc, limit = (10,)):

    '''Trả về True nếu số trang <= limit'''

    num_pages = doc.ComputeStatistics(2)

    return num_pages <= limit





def process_shape(word, shape):

    '''Chuyển 1 shape MathType sang TeX nếu đúng loại'''

    if shape.Type == 1:

        shape.Range.Select()

        word.Run('MTCommand_TeXToggle')

        return None

    return None

# WARNING: Decompyle incomplete





def fix_he_math(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

    else:

        myrange = doc.Content

    for i in range(500):

        find_range = myrange.Duplicate

        find = find_range.Find

        find.ClearFormatting()

        find.Text = '\\$*\\$'

        find.MatchWildcards = True

        if find.Execute():

            myrange.SetRange(Start = find_range.End, End = myrange.End)

            find = find_range.Find

            find.ClearFormatting()

            find.Text = '^13'

            find.Replacement.Text = ' '

            find.MatchWildcards = True

            find.Execute(Replace = 2)

            continue

        range(500)

        return None

    return None

# WARNING: Decompyle incomplete





def fix_lim_latex(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

    else:

        myrange = doc.Content

    find_range = myrange.Duplicate

    find = find_range.Find

    find.ClearFormatting()

    find.Text = '\\$*\\$'

    find.MatchWildcards = True

    if not find.Execute():

        return None

    latex = find_range.Text

    core = latex[1:-1].strip()

    if not re.search('\\\\lim(?![A-Za-z])', core) and re.search('\\\\displaystyle', core):

        new_latex = '\\displaystyle ' + core

        find_range.Text = f'''${new_latex}$'''

    myrange.SetRange(Start = find_range.End, End = myrange.End)

    continue

# WARNING: Decompyle incomplete





def fix_phan_tram(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

    else:

        myrange = doc.Content

    find_range = myrange.Duplicate

    find = find_range.Find

    find.ClearFormatting()

    find.Text = '\\$*\\$'

    find.MatchWildcards = True

    if not find.Execute():

        return None

    latex = find_range.Text

    latex = re.sub('\\\\*%', '\\%', latex)

    find_range.Text = latex

    myrange.SetRange(Start = find_range.End, End = myrange.End)

    continue

# WARNING: Decompyle incomplete





def fix_brace_latex(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

    else:

        myrange = doc.Content

    find_range = myrange.Duplicate

    find = find_range.Find

    find.ClearFormatting()

    find.Text = '\\$*\\$'

    find.MatchWildcards = True

    if not find.Execute():

        return None

    latex = find_range.Text

    new_latex = latex[1:-1].strip()

    if not new_latex.startswith('{') or new_latex.endswith('}'):

        new_latex = '{' + new_latex + '}'

        find_range.Text = f'''${new_latex}$'''

    myrange.SetRange(Start = find_range.End, End = myrange.End)

    continue

# WARNING: Decompyle incomplete





def toggle_shapes(word, shapes):

    '''Chuyển tất cả InlineShapes trong iterable'''

    for shape in shapes:

        process_shape(word, shape)





def mathtype_to_latex(word, mode, thongbao = ('whole', 'yes')):

    MAX_PAGES = 5

    doc = word.ActiveDocument

    word.Visible = True

    if not mode == 'whole' and check_pages_count(doc, limit = MAX_PAGES):

        ok = messagebox.askokcancel('Xác nhận tiếp tục', f'''Tài liệu quá {MAX_PAGES} trang.\n\nKhông khuyến nghị chuyển toàn bộ.\nBạn có muốn tiếp tục không?''')

        if not ok:

            return None

    word.ScreenUpdating = False

    if mode == 'whole':

        toggle_shapes(word, doc.InlineShapes)

        thay_the_replace(word, '(\\\\\\[)(*)(\\\\\\])', '$\\2$')

        thay_the_replace(word, '(\\<)([! ])', '< \\2')

        thay_the_replace(word, '[ ]{1,}\\$', ' $')

        thay_the_replace(word, '\\$[ ]{1,}', '$ ')

        fix_phan_tram(word)

        fix_he_math(word)

        fix_lim_latex(word)

    else:

        sel = word.Selection

        rng = sel.Range

        if sel.Type != 1:

            toggle_shapes(word, sel.Range.InlineShapes)

            rng.Select()

            thay_the_replace_stop(word, '(\\\\\\[)(*)(\\\\\\])', '$\\2$')

            rng.Select()

            thay_the_replace_stop(word, '(\\<)([! ])', '< \\2')

            rng.Select()

            thay_the_replace_stop(word, '[ ]{1,}\\$', ' $')

            rng.Select()

            thay_the_replace_stop(word, '\\$[ ]{1,}', '$ ')

            rng.Select()

            fix_phan_tram(word)

            rng.Select()

            fix_he_math(word)

            rng.Select()

            fix_lim_latex(word)

        else:

            messagebox.showinfo('Thông báo', 'Chưa chọn vùng')

            return None

    word.ScreenUpdating = True

    if thongbao == 'yes':

        messagebox.showinfo('Thông báo', 'Xong')

        return None

    return None

# WARNING: Decompyle incomplete





def gui_mathtype2Latex(root, word):

    pass

# WARNING: Decompyle incomplete





def Latext2Mathtype_range(word, rng):

    patterns = [

        '\\$(*)\\$']

    for pattern in patterns:

        search_range = rng.Duplicate

        find = search_range.Find

        find.ClearFormatting()

        find.MatchWildcards = True

        find.Wrap = 0

        find.Text = pattern

        if not find.Execute():

            continue

        search_range.Select()

        word.Run('MTCommand_TeXToggle')

        search_range.SetRange(Start = search_range.End + 1, End = rng.End)

        find = search_range.Find

        find.ClearFormatting()

        find.MatchWildcards = True

        find.Wrap = 0

        if find.Execute():

            continue

    continue

    return None

# WARNING: Decompyle incomplete





def latex_to_mathtype(word, mode, thongbao = ('whole', 'yes')):

    '''Convert LaTeX text ($...$) sang MathType object'''

    doc = word.ActiveDocument

    word.Visible = True

    word.ScreenUpdating = False

    if mode == 'whole':

        rng = doc.Content

        thay_the_replace(word, '(\\\\\\[)(*)(\\\\\\])', '$\\2$')

        fix_brace_latex(word)

        Latext2Mathtype_range(word, rng)

    else:

        sel = word.Selection

        rng = sel.Range

        if sel.Type != 1:

            rng.Select()

            thay_the_replace_stop(word, '(\\\\\\[)(*)(\\\\\\])', '$\\2$')

            rng.Select()

            fix_brace_latex(word)

            Latext2Mathtype_range(word, rng)

        else:

            messagebox.showinfo('Thông báo', 'Chưa chọn vùng')

    word.ScreenUpdating = True

    if thongbao == 'yes':

        messagebox.showinfo('Thông báo', 'Xong')

        return None

    return None

# WARNING: Decompyle incomplete





def gui_latex2Mathtype(root, word):

    pass

# WARNING: Decompyle incomplete





def xu_ly_bang_dung_sai(word):

    doc = word.ActiveDocument

    selection = word.Selection

    myrange = word.Selection.Range

    table_count = myrange.Tables.Count

    if table_count > 0:

        for table in selection.Tables:

            first_row = []

            for c in range(1, table.Rows(1).Cells.Count + 1):

                text = table.Rows(1).Cells(c).Range.Text.strip().replace('\r\x07', '')

                first_row.append(text)

            if len(first_row) >= 3:

                col1 = first_row[0].lower()

                col2 = first_row[1].lower()

                col3 = first_row[2].lower()

                for r in range(1, table.Rows.Count + 1):

                    row = table.Rows(r)

                    cell_1 = row.Cells(1)

                    cell_2 = row.Cells(2)

                    cell_3 = row.Cells(3)

                    text_1 = cell_1.Range.Text.strip().replace('\r\x07', '')

                    text_2 = cell_2.Range.Text.strip().replace('\r\x07', '')

                    text_3 = cell_3.Range.Text.strip().replace('\r\x07', '')

                    if not 'Đ' in text_2:

                        continue

                    rng = cell_1.Range

                    text = rng.Text[0]

                    char_range = rng.Duplicate

                    char_range.End = char_range.Start + 1

                    char_range.Font.Underline = True

                if 'mệnh' in col1 and 'đúng' in col2 and 'sai' in col3:

                    table.Rows(1).Delete()

                table.Columns(2).Delete()

                table.Columns(2).Delete()

                table.ConvertToText(Separator = ' ', NestedTables = False)

        continue

        return None

    messagebox.showinfo('Thông báo', 'Chưa có vùng chọn, hoặc vùng chọn không có Bảng')

    return None

# WARNING: Decompyle incomplete





def convert_DS_of_btp(word):

    doc = word.ActiveDocument

    if word.Selection.Type == 1:

        messagebox.showinfo('Thông báo', 'Chưa có vùng được chọn')

        return None

    thay_the_replace_stop(word, '[abcd\\- \\)]', '')

    return None

# WARNING: Decompyle incomplete





def gui_table_dung_sai(root, word):

    pass

# WARNING: Decompyle incomplete





def them_cham_ABCD(word):

    

    def thay_the_replace_stop_only(word, text, text_replace):

        doc = word.ActiveDocument

        if word.Selection.Type != 1:

            rng = word.Selection.Range

            find = rng.Find

            find.ClearFormatting()

            find.Replacement.ClearFormatting()

            find.Text = text

            find.Replacement.Text = text_replace

            find.Forward = True

            find.MatchCase = True

            find.MatchWholeWord = False

            find.MatchWildcards = True

            find.MatchSoundsLike = False

            find.MatchAllWordForms = False

            find.Wrap = 0

            find.Format = False

            find.Execute(Replace = 1)

            return None

        return None

    # WARNING: Decompyle incomplete



    

    def thay_the_replace_stop_no_format_only(word, text, text_replace):

        doc = word.ActiveDocument

        if word.Selection.Type != 1:

            rng = word.Selection.Range

            find = rng.Find

            find.ClearFormatting()

            find.Replacement.ClearFormatting()

            find.Text = text

            find.Replacement.Text = text_replace

            find.Replacement.Font.Bold = False

            find.Replacement.Font.Underline = False

            find.Replacement.Font.Color = win32api.RGB(0, 0, 0)

            find.Forward = True

            find.MatchCase = True

            find.MatchWholeWord = False

            find.MatchWildcards = True

            find.MatchSoundsLike = False

            find.MatchAllWordForms = False

            find.Wrap = 0

            find.Format = True

            find.Execute(Replace = 1)

            return None

        return None

    # WARNING: Decompyle incomplete



    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        doc_length = doc.Content.End

        start_old = myrange.Start

        end_old = myrange.End

        new_start = start_old - 1 if start_old > 0 else start_old

        new_end = end_old + 1 if end_old + 1 <= doc_length else end_old

        myrange.SetRange(Start = new_start, End = new_end)

        for i in range(300):

            myrange_find = myrange.Duplicate

            find = myrange_find.Find

            find.ClearFormatting()

            find.Replacement.ClearFormatting()

            find.Text = '([^13^9][A-D].)(*)([^t^13])'

            find.MatchWildcards = True

            if find.Execute():

                myrange.SetRange(Start = myrange_find.End - 2, End = myrange.End)

                myrange_find.SetRange(Start = myrange_find.End - 2, End = myrange_find.End)

                myrange_find.Select()

                thay_the_replace_stop_only(word, '([^t^13])', '.\\1')

                myrange_find.SetRange(Start = myrange_find.Start, End = myrange_find.End + 1)

                myrange_find.Select()

                thay_the_replace_stop_only(word, '([.]{2,})', '.')

                myrange_find.Select()

                thay_the_replace_stop_no_format_only(word, '(.)([^t^13])', '\\1\\2')

                continue

            range(300)

            return None

        return None

    messagebox.showinfo('Thông báo', 'Bạn chưa chọn vùng xử lí')

    return None

# WARNING: Decompyle incomplete





def qua_abcd_ngoac(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        my_select_range = word.Selection.Range

        my_select_range = mo_rong_vung_chon(doc, my_select_range)

        myrange = my_select_range.Duplicate

        find = myrange.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = '([^13^9])([Aa])([.\\)])'

        find.Replacement.Text = '\\1a)'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        find.Execute(Replace = 2)

        myrange = my_select_range.Duplicate

        find = myrange.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = '([^13^9])([Bb])([.\\)])'

        find.Replacement.Text = '\\1b)'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        find.Execute(Replace = 2)

        myrange = my_select_range.Duplicate

        find = myrange.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = '([^13^9])([Cc])([.\\)])'

        find.Replacement.Text = '\\1c)'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        find.Execute(Replace = 2)

        myrange = my_select_range.Duplicate

        find = myrange.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = '([^13^9])([Dd])([.\\)])'

        find.Replacement.Text = '\\1d)'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        find.Execute(Replace = 2)

        return None

    messagebox.showinfo('Thông báo', 'Bạn chưa chọn vùng xử lí')

    return None

# WARNING: Decompyle incomplete





def qua_ABCD_cham(word):

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        my_select_range = word.Selection.Range

        my_select_range = mo_rong_vung_chon(doc, my_select_range)

        myrange = my_select_range.Duplicate

        find = myrange.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = '([^13^9])([Aa])([.\\)])'

        find.Replacement.Text = '\\1A.'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        find.Execute(Replace = 2)

        myrange = my_select_range.Duplicate

        find = myrange.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = '([^13^9])([Bb])([.\\)])'

        find.Replacement.Text = '\\1B.'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        find.Execute(Replace = 2)

        myrange = my_select_range.Duplicate

        find = myrange.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = '([^13^9])([Cc])([.\\)])'

        find.Replacement.Text = '\\1C.'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        find.Execute(Replace = 2)

        myrange = my_select_range.Duplicate

        find = myrange.Find

        find.ClearFormatting()

        find.Replacement.ClearFormatting()

        find.Text = '([^13^9])([Dd])([.\\)])'

        find.Replacement.Text = '\\1D.'

        find.Forward = True

        find.MatchCase = True

        find.MatchWholeWord = False

        find.MatchWildcards = True

        find.MatchSoundsLike = False

        find.MatchAllWordForms = False

        find.Wrap = 0

        find.Format = False

        find.Execute(Replace = 2)

        return None

    messagebox.showinfo('Thông báo', 'Bạn chưa chọn vùng xử lí')

    return None

# WARNING: Decompyle incomplete





def fix_dap_an_ABCD(word):

    

    def thay_ky_tu_gia_tao_latin(word):

        '''Chuẩn hóa các ký tự Greek/Cyrillic trông giống chữ Latin trong Word.'''

        doc = word.ActiveDocument

        char_map = {

            'Α': 'A',

            'Β': 'B',

            'А': 'A',

            'В': 'B',

            'С': 'C',

            'а': 'a',

            'в': 'b',

            'с': 'c' }

        for old_char, new_char in char_map.items():

            find = word.Selection.Find

            find.ClearFormatting()

            find.Replacement.ClearFormatting()

            find.Text = old_char

            find.Replacement.Text = new_char

            find.Forward = True

            find.MatchCase = True

            find.Wrap = 0

            find.Execute(Replace = 2)

        return None

    # WARNING: Decompyle incomplete



    doc = word.ActiveDocument

    if word.Selection.Type == 1:

        messagebox.showinfo('Thông báo', 'Chưa có vùng được chọn')

        return None

    thay_ky_tu_gia_tao_latin(word)

    thay_the_replace_stop(word, '([^t^13])([ ]{1,})([A-Da-d][.\\)])', '\\1\\3')

    thay_the_replace_stop(word, '([ABCDabcd][.\\)])([^t ]{1,})', '\\1 ')

    thay_the_replace_stop(word, '([ ]{3,})([A-Da-d][.\\)])', '^t\\2')

    return None

# WARNING: Decompyle incomplete





def gui_remove_collapsible(root, word):

    pass

# WARNING: Decompyle incomplete





def gui_fix_Dap_an(root, word):

    pass

# WARNING: Decompyle incomplete





def gui_replace(root, word):

    pass

# WARNING: Decompyle incomplete



