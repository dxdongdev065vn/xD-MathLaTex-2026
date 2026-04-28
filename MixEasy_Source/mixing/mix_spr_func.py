from spire.doc import Section, Paragraph, Table, TextRange, Shape, FileFormat, HorizontalAlignment, Regex, CharacterFormat

from spire.doc import UnderlineStyle

from spire.doc import Document as SpireDocument

from spire.doc import Color

import os

import re

import random

import tkinter as tk

from tkinter import ttk, Label, filedialog, Entry, Button, Frame, Listbox, Scrollbar, messagebox

from tkinter import simpledialog

from openpyxl import Workbook

from openpyxl.styles import Font

from openpyxl.styles import Alignment

from openpyxl.styles import Border, Side

from openpyxl.styles import PatternFill

from openpyxl.styles import GradientFill

from openpyxl.utils import get_column_letter

from openpyxl import load_workbook



def find_phan(doc, text_start, text_end):

    indices_S = []

    indices_E = []

    section = doc.Sections[0]

    for j in range(section.Body.ChildObjects.Count):

        child = section.Body.ChildObjects[j]

        if not isinstance(child, Paragraph):

            continue

        text = child.Text.strip()

        if text.startswith(text_start):

            indices_S.append(j)

            continue

        if not text.startswith(text_end):

            continue

        indices_E.append(j)

    return (indices_S, indices_E)





def tim_cau_trong_phan(doc, text_start, text_end):

    (indices_S, indices_E) = find_phan(doc, text_start, text_end)

    all_indices_cau = []

    if len(indices_S) != len(indices_E):

        messagebox.showinfo('Thông báo', f'''Số kí hiệu {text_start} và {text_end} không bằng nhau, hãy chỉnh sửa phù hợp''')

        return None

    if len(indices_S) != 0:

        section = doc.Sections[0]

        pattern1 = '^Câu [0-9]{1,}'

        for start, end in zip(indices_S, indices_E):

            indices_cau = []

            for j in range(start, end):

                child = section.Body.ChildObjects[j]

                if not isinstance(child, Paragraph):

                    continue

                if not re.match(pattern1, child.Text):

                    continue

                indices_cau.append(j)

            indices_cau.append(end)

            all_indices_cau.append(indices_cau)

    return all_indices_cau





def tim_phuong_an_trong_cau_phan_I(doc, from_a, to_b):

    section = doc.Sections[0]

    to_end = to_b

    for j in range(from_a, to_b):

        child = section.Body.ChildObjects[j]

        if not isinstance(child, Paragraph):

            continue

        text = child.Text.strip()

        if not text == 'Lời giải':

            continue

        to_end = j

        range(from_a, to_b)

    indices_phuong_an = []

    labels = [

        'A.',

        'B.',

        'C.',

        'D.']

    for j in range(from_a, to_end):

        child = section.Body.ChildObjects[j]

        if not isinstance(child, Paragraph):

            continue

        text = child.Text.strip()

        if not text.startswith(tuple(labels)):

            continue

        indices_phuong_an.append(j)

    indices_phuong_an.append(to_end)

    return indices_phuong_an





def tim_phuong_an_trong_cau_phan_II(doc, from_a, to_b):

    section = doc.Sections[0]

    to_end = to_b

    for j in range(from_a, to_b):

        child = section.Body.ChildObjects[j]

        if not isinstance(child, Paragraph):

            continue

        text = child.Text.strip()

        if not text == 'Lời giải':

            continue

        to_end = j

        range(from_a, to_b)

    indices_phuong_an = []

    labels = [

        'a)',

        'b)',

        'c)',

        'd)']

    for j in range(from_a, to_end):

        child = section.Body.ChildObjects[j]

        if not isinstance(child, Paragraph):

            continue

        text = child.Text.strip()

        if not text.startswith(tuple(labels)):

            continue

        indices_phuong_an.append(j)

    indices_phuong_an.append(to_end)

    return indices_phuong_an





def find_phan_EN(doc):

    text_start = '<S@>'

    text_end = '<E@>'

    (indices_S, indices_E) = find_phan(doc, text_start, text_end)

    return (indices_S, indices_E)





def tim_cau_trong_phan_EN(doc):

    (indices_S, indices_E) = find_phan_EN(doc)

    all_indices_cau = []

    if len(indices_S) != len(indices_E):

        messagebox.showinfo('Thông báo', 'Số kí hiệu <S@> và <E@> không bằng nhau, hãy chỉnh sửa phù hợp')

        return None

    if len(indices_S) != 0:

        section = doc.Sections[0]

        pattern1 = '^Question [0-9]{1,}'

        pattern2 = '^Câu [0-9]{1,}'

        for start, end in zip(indices_S, indices_E):

            indices_cau = []

            for j in range(start, end):

                child = section.Body.ChildObjects[j]

                if not isinstance(child, Paragraph):

                    continue

                if not re.match(pattern1, child.Text) and re.match(pattern2, child.Text):

                    continue

                indices_cau.append(j)

        indices_cau.append(end)

        all_indices_cau.append(indices_cau)

        continue

    return all_indices_cau



