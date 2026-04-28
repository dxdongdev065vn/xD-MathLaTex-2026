from spire.doc import Section, Paragraph, Table, TextRange, Shape, FileFormat, HorizontalAlignment, Regex

from spire.doc import UnderlineStyle

from spire.doc import Document as SpireDocument

from spire.doc import Color

from spire.doc import DocumentObjectType

from spire.doc import OfficeMath

from docx import Document

from docx.text.paragraph import Paragraph

from docx.shared import RGBColor

from docx.shared import Inches, Pt

from docx.oxml import OxmlElement

from docx.oxml.ns import qn

from docx.oxml import parse_xml

from lxml import etree



ElementTree

from xml.etree.ElementTree import QName

import xml.etree.ElementTree, etree

import os

import re

import random

import zipfile

import bisect

import tkinter as tk

from tkinter import ttk, Label, filedialog, Entry, Button, Frame, Listbox, Scrollbar, messagebox

from tkinter import simpledialog

from openpyxl import Workbook

from openpyxl.styles import Font

from openpyxl.styles import Alignment

from openpyxl.styles import Border, Side

from openpyxl.styles import PatternFill

from openpyxl.styles import GradientFill

from openpyxl.utils import get_column_letter

from openpyxl import load_workbook

from formatting import chuan_hoa_docx as chdocx

from mixing import mix_docx as mdocx

from mixing import mix_docx_func as mdocx_f

from tools import tool_by_docx as docxtool



def open_doc_off_spire(doc_path):

    doc = SpireDocument()

    doc.LoadFromFile(doc_path)

    return doc





def save_and_close_spire(doc, output_file):

    doc.SaveToFile(output_file)

    doc.Close()

    return None

# WARNING: Decompyle incomplete





def update_so_trang_spr(doc_path):

    doc = SpireDocument()

    doc.LoadFromFile(doc_path)

    num_pages = doc.GetPageCount()

    new_text = f'''{num_pages:02}'''

    doc.Replace('<sotrang>', new_text, True, False)

    save_and_close_spire(doc, doc_path)





def delete_only_bookmark_tags(doc):

    excluded_names = {

        'MDH',

        'num_page'}

    bookmark_start_ids_to_keep = set()

    for elem in doc.element.body.iter():

        if not elem.tag == qn('w:bookmarkStart'):

            continue

        name = elem.get(qn('w:name'))

        bookmark_id = elem.get(qn('w:id'))

        if not name in excluded_names:

            continue

        bookmark_start_ids_to_keep.add(bookmark_id)

# WARNING: Decompyle incomplete





def delete_bookmarks(doc):

    excluded_names = {

        'MDH',

        'num_page'}

    bookmark_start_ids_to_keep = set()

    for elem in doc.element.body.iter():

        if not elem.tag == qn('w:bookmarkStart'):

            continue

        name = elem.get(qn('w:name'))

        bookmark_id = elem.get(qn('w:id'))

        if not name in excluded_names:

            continue

        bookmark_start_ids_to_keep.add(bookmark_id)

# WARNING: Decompyle incomplete





def find_nhom_all(doc, text_start, text_end):

    (indices_S, indices_E) = mdocx_f.find_phan(doc, text_start, text_end)

    return (indices_S, indices_E)





def Tao_group_tu_indices(doc, indices):

    groups = mdocx.Tao_group_tu_indices(doc, indices)

    return groups





def tron_group_tu_indices(doc, indices):

    mdocx.tron_group_tu_indices(doc, indices)





def tim_nhom_tu_a_den_b_tron_nhom(doc, from_a, to_b):

    text_start = '<SNHOM@>'

    body = doc.element.body

    paragraphs = list(body.iterchildren())

    indices_S = []

    for j in range(from_a, min(to_b, len(paragraphs))):

        para = paragraphs[j]

        if not para.tag.endswith('p'):

            continue

        paragraph = Paragraph(para, doc)

        text = paragraph.text

        if not text.startswith(text_start):

            continue

        indices_S.append(j)

    return indices_S





def tim_cau_tu_a_den_b(doc, from_a, to_b):

    pattern1 = '^Question [0-9]{1,}[.:]'

    pattern2 = '^Câu [0-9]{1,}[.:]'

    body = doc.element.body

    paragraphs = list(body.iterchildren())

    indices_cau = []

    for j in range(from_a, min(to_b, len(paragraphs))):

        para = paragraphs[j]

        if not para.tag.endswith('p'):

            continue

        paragraph = Paragraph(para, doc)

        text = paragraph.text

        if not re.match(pattern1, text) and re.match(pattern2, text):

            continue

        indices_cau.append(j)

    indices_cau.append(to_b)

    return indices_cau





def tron_phuong_an_trong_vung(doc, snhom, enhom):

    (indices_S, indices_E) = find_nhom_all(doc, snhom, enhom)

    if len(indices_S) != len(indices_E):

        messagebox.showinfo('Thông báo', f'''Số kí hiệu {snhom} và {enhom} không bằng nhau, hãy chỉnh sửa phù hợp''')

        return None

    if len(indices_S) != 0:

        for start, end in zip(indices_S, indices_E):

            indices_cau = tim_cau_tu_a_den_b(doc, start, end)

            for k in range(len(indices_cau) - 1):

                phuong_an = mdocx_f.tim_phuong_an_trong_cau_phan_I(doc, indices_cau[k], indices_cau[k + 1])

                tron_group_tu_indices(doc, phuong_an)

        return None





def tron_phuong_an(doc):

    tron_phuong_an_trong_vung(doc, '<SNHOM@><D', '<ENHOM@><D')

    tron_phuong_an_trong_vung(doc, '<SNHOM@><CD>', '<ENHOM@><CD>')





def tron_cau_trong_vung(doc, text_start, text_end):

    (indices_S, indices_E) = find_nhom_all(doc, text_start, text_end)

    if len(indices_S) != len(indices_E):

        messagebox.showinfo('Thông báo', f'''Số kí hiệu {text_start} và {text_end} không bằng nhau, hãy chỉnh sửa phù hợp''')

        return None

    if len(indices_S) != 0:

        for start, end in zip(indices_S, indices_E):

            indices_cau = tim_cau_tu_a_den_b(doc, start, end)

            tron_group_tu_indices(doc, indices_cau)

        return None





def tron_cau(doc):

    tron_cau_trong_vung(doc, '<SNHOM@><C', '<ENHOM@><C')

    tron_cau_trong_vung(doc, '<SNHOM@><DC>', '<ENHOM@><DC>')





def tron_nhom(doc):

    (indices_S, indices_E) = mdocx_f.find_phan_EN(doc)

    if len(indices_S) != len(indices_E):

        messagebox.showinfo('Thông báo', 'Số kí hiệu S@ và E@ không bằng nhau, hãy chỉnh sửa phù hợp')

        return None

    if len(indices_S) != 0:

        for start, end in zip(indices_S, indices_E):

            indices_nhom = tim_nhom_tu_a_den_b_tron_nhom(doc, start, end)

            tron_group_tu_indices(doc, indices_nhom)

        return None





def dem_so_cau(doc):

    cau_p1 = mdocx_f.tim_cau_trong_phan_EN(doc)

    so_cau_p1 = 0

    if len(cau_p1) > 0:

        for i in range(len(cau_p1)):

            so_cau_p1 += len(cau_p1[i]) - 1

    return so_cau_p1





def lay_dap_an_vao_excel_TN(doc, ws, column = (1,)):

    dong = 0

    labels = [

        'A.',

        'B.',

        'C.',

        'D.']

    labels2 = [

        'A',

        'B',

        'C',

        'D']

    body = doc.element.body

    paragraphs = list(body.iterchildren())

    all_indices_cau = mdocx_f.tim_cau_trong_phan_EN(doc)

    if len(all_indices_cau) > 0:

        for idx in range(len(all_indices_cau)):

            cau = all_indices_cau[idx]

            for k in range(len(cau) - 1):

                phuong_an = mdocx_f.tim_phuong_an_trong_cau_phan_I(doc, cau[k], cau[k + 1])

                encounter_count = 0

                for h in range(phuong_an[0], phuong_an[-1]):

                    para = paragraphs[h]

                    if not para.tag.endswith('p'):

                        continue

                    paragraph = Paragraph(para, doc)

                    text = paragraph.text.strip()

                    if not text.startswith(tuple(labels)):

                        continue

                    current_label = labels2[encounter_count % len(labels2)]

                    encounter_count += 1

                    if not paragraph.runs:

                        continue

                    first_run = paragraph.runs[0]

                    if not first_run.font.underline:

                        first_run.font.underline

                        if first_run.font.color:

                            first_run.font.color

                    is_correct = first_run.font.color.rgb == 'FF0000'

                    if not is_correct:

                        continue

                    ws.cell(row = dong + 4, column = column, value = current_label)

                    range(phuong_an[0], phuong_an[-1])

            dong += 1

        continue

        return None





def tron_and_xuat_de_NN(entry_input, entry_lantron, entry_newname, check6_1_var, check7_1_var, check8_1_var):

    newname_first = entry_lantron.get().strip()

    input_file_get = entry_input.get()

    input_file = os.path.normpath(input_file_get)

    if not os.path.exists(input_file):

        messagebox.showinfo('Thông báo', 'Hãy tắt hết các file word, và chọn lại file gốc để trộn.')

        return None

    page_kind = 'A4'

    bo_loi_giai = 'YES'

    langue = 'english'

    if check6_1_var.get():

        page_kind = '2C'

    if check7_1_var.get():

        bo_loi_giai = 'NO'

    if check8_1_var.get():

        langue = 'japan'

    new_names = entry_newname.get().strip().split(',')

    if not new_names:

        messagebox.showwarning('Warning', 'Please enter new name(s).')

        return None

    i = 0

# WARNING: Decompyle incomplete



