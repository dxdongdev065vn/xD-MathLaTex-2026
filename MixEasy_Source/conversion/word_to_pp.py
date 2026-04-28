import os

import time

import datetime

import win32com.client as win32com

from win32com.client import client as win32

import win32api

import re

import sys

import shutil

import pythoncom

import tempfile

import wmi

import docx

import random

from docx import Document

from docx.shared import Pt

from docx.enum.text import WD_PARAGRAPH_ALIGNMENT

from docx.shared import RGBColor

from openpyxl import Workbook

from openpyxl.styles import Font

from openpyxl.styles import Alignment

from openpyxl.styles import Border, Side

from openpyxl.styles import PatternFill

from openpyxl.styles import GradientFill

from openpyxl.utils import get_column_letter

from openpyxl import load_workbook

import tkinter as tk

from tkinter import ttk, Label, filedialog, Entry, Button, Frame, Listbox, Scrollbar, messagebox

from tkinter import simpledialog

import webbrowser

from core import functions as vbf



def xoa_shape(slide, name):

    for i in range(1, 7):

        slide.Shapes(name).Delete()

    return None

# WARNING: Decompyle incomplete





def delete_result_box():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    for slide in presentation.Slides:

        for i in range(slide.Shapes.Count, 0, -1):

            shape = slide.Shapes(i)

            shapes_to_delete = [

                'ResultBoxA',

                'ResultBoxB',

                'ResultBoxC',

                'ResultBoxD']

            if shape.Name in shapes_to_delete:

                shape.Delete()

                continue

            if not shape.Name == 'Dap_so_TLN':

                continue

            shape.Visible = 0





def remove_shapes_without_text(slide):

    shapes = slide.Shapes

    for i in range(shapes.Count, 0, -1):

        shape = shapes(i)

        if not shape.HasTextFrame:

            continue

        text_frame = shape.TextFrame

        if not text_frame.HasText:

            continue

        text = text_frame.TextRange.Text.strip()

        if text:

            continue

        shape.Delete()





def make_box_score(slide):

    shape = slide.Shapes('score')

    shape.Delete()

    textbox = slide.Shapes.AddTextbox(Orientation = 1, Left = 45, Top = 481.75, Width = 55, Height = 43)

    textbox.Name = 'score'

    text_range = textbox.TextFrame.TextRange

    text_range.Text = '0'

    textbox.Fill.ForeColor.RGB = win32api.RGB(255, 67, 67)

    text_range.Font.Color.RGB = win32api.RGB(255, 255, 255)

    text_range.Font.Size = 30

    text_range.Font.Bold = True

    textbox.TextFrame.TextRange.ParagraphFormat.Alignment = 2

    textbox.TextFrame.VerticalAnchor = 3

    return None

# WARNING: Decompyle incomplete





def make_box_time(slide):

    shape = slide.Shapes('CountdownTimer')

    shape.Delete()

    textbox = slide.Shapes.AddTextbox(Orientation = 1, Left = 183, Top = 481.75, Width = 55, Height = 43)

    textbox.Name = 'CountdownTimer'

    text_range = textbox.TextFrame.TextRange

    text_range.Text = '30'

    text_range.Font.Color.RGB = win32api.RGB(255, 255, 255)

    text_range.Font.Size = 30

    text_range.Font.Bold = True

    textbox.TextFrame.TextRange.ParagraphFormat.Alignment = 2

    textbox.TextFrame.VerticalAnchor = 3

    return None

# WARNING: Decompyle incomplete





def make_box_del(slide):

    shape = slide.Shapes('Delete_DA')

    shape.Delete()

    textbox = slide.Shapes.AddTextbox(Orientation = 1, Left = 886, Top = 487, Width = 70, Height = 50)

    textbox.Name = 'Delete_DA'

    text_range = textbox.TextFrame.TextRange

    text_range.Text = 'Xóa ĐA'

    textbox.Fill.ForeColor.RGB = win32api.RGB(238, 18, 49)

    text_range.Font.Color.RGB = win32api.RGB(255, 255, 255)

    text_range.Font.Size = 18

    action = textbox.ActionSettings(1)

    action.Action = 7

    action.Run = 'DeleteResultBox'

    return None

# WARNING: Decompyle incomplete





def make_show_TLN(slide):

    shape = slide.Shapes('Show_DS')

    shape.Delete()

    textbox = slide.Shapes.AddTextbox(Orientation = 1, Left = 600, Top = 487, Width = 105, Height = 50)

    textbox.Name = 'Show_DS'

    text_range = textbox.TextFrame.TextRange

    text_range.Text = 'Hiện đáp số'

    textbox.Fill.ForeColor.RGB = win32api.RGB(216, 110, 204)

    text_range.Font.Color.RGB = win32api.RGB(255, 255, 255)

    action = textbox.ActionSettings(1)

    action.Action = 7

    action.Run = 'Show_DS_TLN'

    return None

# WARNING: Decompyle incomplete





def make_hide_TLN(slide):

    shape = slide.Shapes('Hide_DS')

    shape.Delete()

    textbox = slide.Shapes.AddTextbox(Orientation = 1, Left = 750, Top = 487, Width = 105, Height = 50)

    textbox.Name = 'Hide_DS'

    text_range = textbox.TextFrame.TextRange

    text_range.Text = 'Ẩn đáp số'

    textbox.Fill.ForeColor.RGB = win32api.RGB(216, 110, 204)

    text_range.Font.Color.RGB = win32api.RGB(255, 255, 255)

    action = textbox.ActionSettings(1)

    action.Action = 7

    action.Run = 'Hide_DS_TLN'

    return None

# WARNING: Decompyle incomplete





def max_Dap_an(slide):

    shape_names = [

        'Dap_an_a',

        'Dap_an_b',

        'Dap_an_c',

        'Dap_an_d']

    max_length = 0

    shapes = slide.Shapes

    for shape_name in shape_names:

        shape = shapes(shape_name)

        length = shape.Width

        if length > max_length:

            max_length = length

    continue

    return max_length

# WARNING: Decompyle incomplete





def vitri_left_top(slide, shape_name):

    shapes = slide.Shapes

    shape = shapes(shape_name)

    left = shape.Left

    top = shape.Top

    return (left, top)





def fix_len(slide, shape_name, new_length):

    shapes = slide.Shapes

    shape = shapes(shape_name)

    shape.Width = new_length

    return None

# WARNING: Decompyle incomplete





def move_shape_to(slide, shape_name, new_left, new_top):

    shape = None

    for shp in slide.Shapes:

        if not shp.Name == shape_name:

            continue

        shape = shp

        slide.Shapes

    if shape:

        shape.Left = new_left

        shape.Top = new_top

        return None





def fix_width_shape(shape):

    if shape.HasTextFrame:

        text_range = shape.TextFrame.TextRange

        text_range.Font.Name = 'Times New Roman'

        text_range.Font.Size = 28

        shape.Width = 900

        shape.TextFrame.AutoSize = 1

        required_width = text_range.BoundWidth

        if required_width < 900:

            shape.Width = required_width + 20

            return None

        return None





def fix_width_all_shape(powerpoint, slide):

    for i in range(1, slide.Shapes.Count + 1):

        shape = slide.Shapes(i)

        fix_width_shape(shape)





def Fix_width_dap_an_for(powerpoint, slide):

    max_len_da = max_Dap_an(slide)

    if max_len_da != 0:

        fix_len(slide, 'Dap_an_a', max_len_da)

        fix_len(slide, 'Dap_an_b', max_len_da)

        fix_len(slide, 'Dap_an_c', max_len_da)

        fix_len(slide, 'Dap_an_d', max_len_da)

        return None

    return None

# WARNING: Decompyle incomplete





def Fix_result_for(powerpoint, slide):

    names_del = {

        'score',

        'Hide_DS',

        'Show_DS',

        'Delete_DA',

        'ResultBoxA',

        'ResultBoxB',

        'ResultBoxC',

        'ResultBoxD',

        'CountdownTimer'}

    for name in names_del:

        xoa_shape(slide, name)

    for i in range(1, slide.Shapes.Count + 1):

        shape = slide.Shapes(i)

        if not shape.HasTextFrame:

            continue

        text_range = shape.TextFrame.TextRange

        text = text_range.Text

        first_two_chars = text[:2]

        if first_two_chars in frozenset({'A.', 'B.', 'C.', 'D.', 'a)', 'b)', 'c)', 'd)'}):

            first_char = text_range.Characters(1)

            if not first_char.Font.Underline:

                first_char.Font.Underline

            is_correct = first_char.Font.Color.RGB == win32api.RGB(255, 0, 0)

            if first_two_chars == 'A.' or first_two_chars == 'a)':

                shape.Name = 'Dap_an_a'

                macro_Result = 'ShowResultCorrectA' if is_correct else 'ShowResult_IncorrectA'

                action = shape.ActionSettings(1)

                action.Action = 7

                action.Run = macro_Result

                shape.Fill.ForeColor.RGB = win32api.RGB(78, 167, 46)

                text_range.Font.Color.RGB = win32api.RGB(255, 255, 255)

                first_char = text_range.Characters(1)

                first_char.Font.Underline = 0

                first_char.Font.Color.RGB = win32api.RGB(0, 0, 255)

                text_range.Characters(2).Font.Underline = 0

                text_range.Characters(2).Font.Color.RGB = win32api.RGB(0, 0, 255)

                text_range.Characters(3).Font.Underline = 0

            if first_two_chars == 'B.' or first_two_chars == 'b)':

                shape.Name = 'Dap_an_b'

                macro_Result = 'ShowResultCorrectB' if is_correct else 'ShowResult_IncorrectB'

                action = shape.ActionSettings(1)

                action.Action = 7

                action.Run = macro_Result

                shape.Fill.ForeColor.RGB = win32api.RGB(78, 167, 46)

                text_range.Font.Color.RGB = win32api.RGB(255, 255, 255)

                first_char = text_range.Characters(1)

                first_char.Font.Underline = 0

                first_char.Font.Color.RGB = win32api.RGB(0, 0, 255)

                text_range.Characters(2).Font.Underline = 0

                text_range.Characters(2).Font.Color.RGB = win32api.RGB(0, 0, 255)

                text_range.Characters(3).Font.Underline = 0

            if first_two_chars == 'C.' or first_two_chars == 'c)':

                shape.Name = 'Dap_an_c'

                macro_Result = 'ShowResultCorrectC' if is_correct else 'ShowResult_IncorrectC'

                action = shape.ActionSettings(1)

                action.Action = 7

                action.Run = macro_Result

                shape.Fill.ForeColor.RGB = win32api.RGB(78, 167, 46)

                text_range.Font.Color.RGB = win32api.RGB(255, 255, 255)

                first_char = text_range.Characters(1)

                first_char.Font.Underline = 0

                first_char.Font.Color.RGB = win32api.RGB(0, 0, 255)

                text_range.Characters(2).Font.Underline = 0

                text_range.Characters(2).Font.Color.RGB = win32api.RGB(0, 0, 255)

                text_range.Characters(3).Font.Underline = 0

            if first_two_chars == 'D.' or first_two_chars == 'd)':

                shape.Name = 'Dap_an_d'

                macro_Result = 'ShowResultCorrectD' if is_correct else 'ShowResult_IncorrectD'

                action = shape.ActionSettings(1)

                action.Action = 7

                action.Run = macro_Result

                shape.Fill.ForeColor.RGB = win32api.RGB(78, 167, 46)

                text_range.Font.Color.RGB = win32api.RGB(255, 255, 255)

                first_char = text_range.Characters(1)

                first_char.Font.Underline = 0

                first_char.Font.Color.RGB = win32api.RGB(0, 0, 255)

                text_range.Characters(2).Font.Underline = 0

                text_range.Characters(2).Font.Color.RGB = win32api.RGB(0, 0, 255)

                text_range.Characters(3).Font.Underline = 0

            shape = slide.Shapes('Show_DS')

            shape.Delete()

            shape = slide.Shapes('Hide_DS')

            shape.Delete()

            make_box_del(slide)

            continue

        if not first_two_chars == 'ĐS':

            continue

        shape.Name = 'Dap_so_TLN'

        shape.Fill.ForeColor.RGB = win32api.RGB(78, 167, 46)

        text_range.Font.Color.RGB = win32api.RGB(255, 255, 255)

        move_shape_to(slide, 'Dap_so_TLN', 386, 480)

        shape.Visible = 0

        shape = slide.Shapes('Delete_DA')

        shape.Delete()

        make_show_TLN(slide)

        make_hide_TLN(slide)

    make_box_score(slide)

    make_box_time(slide)

    return None

# WARNING: Decompyle incomplete





def Fix_gach_chan_for(powerpoint, slide):

    for i in range(1, slide.Shapes.Count + 1):

        shape = slide.Shapes(i)

        if not shape.HasTextFrame:

            continue

        text_range = shape.TextFrame.TextRange

        text = text_range.Text

        first_three_chars = text[:3]

        if not first_three_chars in frozenset({'A. ', 'B. ', 'C. ', 'D. ', 'a) ', 'b) ', 'c) ', 'd) '}):

            continue

        for k in range(1, 4):

            text_range.Characters(k).Font.Underline = 0





def Fix_canh_dap_an_for(powerpoint, slide):

    max_len_da = max_Dap_an(slide)

    if max_len_da != 0:

        fix_len(slide, 'Dap_an_a', max_len_da)

        fix_len(slide, 'Dap_an_b', max_len_da)

        fix_len(slide, 'Dap_an_c', max_len_da)

        fix_len(slide, 'Dap_an_d', max_len_da)

        vitri = 455

        d = slide.Shapes('Dap_an_d').Height

        vitri = vitri - d

        move_shape_to(slide, 'Dap_an_d', 10, vitri)

        d = slide.Shapes('Dap_an_c').Height

        vitri = vitri - d - 10

        move_shape_to(slide, 'Dap_an_c', 10, vitri)

        d = slide.Shapes('Dap_an_b').Height

        vitri = vitri - d - 10

        move_shape_to(slide, 'Dap_an_b', 10, vitri)

        d = slide.Shapes('Dap_an_a').Height

        vitri = vitri - d - 10

        move_shape_to(slide, 'Dap_an_a', 10, vitri)

        return None

    return None

# WARNING: Decompyle incomplete





def Fix_AllShape_in_silde_for(powerpoint, slide):

    fix_width_all_shape(powerpoint, slide)

    Fix_result_for(powerpoint, slide)

    Fix_canh_dap_an_for(powerpoint, slide)

    remove_shapes_without_text(slide)





def word_to_powerpoint_for(word, powerpoint, myrange):

    doc = word.ActiveDocument

    presentation = powerpoint.ActivePresentation

    slide = presentation.Slides.Add(presentation.Slides.Count + 1, 1)

    for k in range(0, slide.Shapes.Count):

        slide.Shapes(1).Delete()

    top_position = 0

    last_table_range_end = 0

    for para in myrange.Paragraphs:

        text_range = para.Range

        if not text_range.Text.strip() and text_range.InlineShapes.Count > 0:

            continue

        top_position += 10

        if text_range.Start < last_table_range_end:

            continue

        if text_range.Information(12):

            table = text_range.Tables(1)

            table_range = table.Range

            table_range.Copy()

            time.sleep(0.3)

            pic = slide.Shapes.Paste()

            pic.Left = 10

            pic.Top = top_position

            top_position += pic.Height + 10

            last_table_range_end = table_range.End

            continue

        if text_range.InlineShapes.Count > 0:

            text_range.Copy()

            time.sleep(0.3)

            pic = slide.Shapes.Paste()

            time.sleep(0.3)

            pic.Left = 700

            pic.Top = top_position + 10

            continue

        if not text_range.Text.strip():

            continue

        text_range.Copy()

        time.sleep(0.3)

        text_box = slide.Shapes.AddTextbox(Orientation = 1, Left = 10, Top = top_position, Width = 900, Height = 50)

        text_box.TextFrame.TextRange.Paste()

        time.sleep(0.3)

        fix_width_shape(text_box)

        top_position += text_box.Height

    Fix_result_for(powerpoint, slide)

    Fix_width_dap_an_for(powerpoint, slide)

    remove_shapes_without_text(slide)

    return None

# WARNING: Decompyle incomplete





def word_to_powerpoint_myrange_cau(word, powerpoint, myrange_cau):

    myrange_find = myrange_cau.Duplicate

    find = myrange_find.Find

    find.ClearFormatting()

    find.Text = '(^13Lời giải)'

    find.MatchWildcards = True

    if find.Execute():

        myrange_find.SetRange(Start = myrange_cau.Start, End = myrange_find.Start)

        word_to_powerpoint_for(word, powerpoint, myrange_find)

        myrange_cau.SetRange(Start = myrange_find.End + 1, End = myrange_cau.End)

        myrange_cau.Select()

    word_to_powerpoint_for(word, powerpoint, myrange_cau)





def find_cau_or_title_paragraphs(doc):

    pattern = re.compile('^Câu\\s\\d+', re.UNICODE)

    indices = []

    for i in range(1, doc.Paragraphs.Count + 1):

        para_text = doc.Paragraphs(i).Range.Text.strip()

        if not para_text:

            continue

        para = doc.Paragraphs(i)

        style_name = para.Range.Style.NameLocal

        if not pattern.match(para_text) and style_name.startswith('Heading'):

            continue

        indices.append(i)

    if indices[0] != 1:

        indices.insert(0, 1)

    indices.append(doc.Paragraphs.Count + 1)

    return indices





def w2p_all(word):

    powerpoint = vbf.khoi_tao_powerpoint_2()

    doc = word.ActiveDocument

    vbf.Convert_Auto_To_Text(word)

    indices = find_cau_or_title_paragraphs(doc)

    for i in range(len(indices) - 1):

        myrange = doc.Range(doc.Paragraphs(indices[i]).Range.Start, doc.Paragraphs(indices[i + 1] - 1).Range.End)

        word_to_powerpoint_myrange_cau(word, powerpoint, myrange)

    return None

# WARNING: Decompyle incomplete





def w2p_all_OLD(word):

    powerpoint = vbf.khoi_tao_powerpoint_2()

    doc = word.ActiveDocument

    vbf.Convert_Auto_To_Text(word)

    myrange = doc.Range()

    for i in range(1, 200):

        myrange_cau = myrange.Duplicate

        find = myrange_cau.Find

        find.ClearFormatting()

        find.Text = '(^13Câu [0-9])'

        find.MatchWildcards = True

        if find.Execute():

            myrange_cau.SetRange(Start = myrange.Start, End = myrange_cau.End - 6)

            word_to_powerpoint_myrange_cau(word, powerpoint, myrange_cau)

            myrange = doc.Range(myrange_cau.End + 1, doc.Range().End)

            continue

        word_to_powerpoint_myrange_cau(word, powerpoint, myrange_cau)

        range(1, 200)

        return None

    return None

# WARNING: Decompyle incomplete





def w2p_select(word):

    powerpoint = vbf.khoi_tao_powerpoint_2()

    doc = word.ActiveDocument

    if word.Selection.Type != 1:

        myrange = word.Selection.Range

        word_to_powerpoint_for(word, powerpoint, myrange)

        return None

    messagebox.showinfo('Thông báo', 'Chưa có câu được chọn')

    return None

# WARNING: Decompyle incomplete





def Fix_AllShape_in_silde():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    slide = powerpoint.ActiveWindow.View.Slide

    Fix_AllShape_in_silde_for(powerpoint, slide)





def Fix_size_Font(slide, size_font):

    names = {

        'score',

        'Hide_DS',

        'Show_DS',

        'Delete_DA',

        'ResultBoxA',

        'ResultBoxB',

        'ResultBoxC',

        'ResultBoxD',

        'CountdownTimer'}

    for i in range(1, slide.Shapes.Count + 1):

        shape = slide.Shapes(i)

        if not shape.HasTextFrame:

            continue

        if not shape.Name not in names:

            continue

        text_range = shape.TextFrame.TextRange

        text_range.Font.Size = size_font





def Fix_size_Font_one_silde():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    slide = powerpoint.ActiveWindow.View.Slide

    size_font = simpledialog.askstring('Nhập size font', 'Bạn muốn Font bao nhiêu')

    Fix_size_Font(slide, size_font)





def Fix_size_Font_All_silde():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    size_font = simpledialog.askstring('Nhập size font', 'Bạn muốn Font bao nhiêu')

    names = {

        'score',

        'Hide_DS',

        'Show_DS',

        'Delete_DA',

        'ResultBoxA',

        'ResultBoxB',

        'ResultBoxC',

        'ResultBoxD',

        'CountdownTimer'}

    for slide in presentation.Slides:

        Fix_size_Font(slide, size_font)





def Fix_name_Font(slide, name_font):

    names = {

        'score',

        'Hide_DS',

        'Show_DS',

        'Delete_DA',

        'ResultBoxA',

        'ResultBoxB',

        'ResultBoxC',

        'ResultBoxD',

        'CountdownTimer'}

    for i in range(1, slide.Shapes.Count + 1):

        shape = slide.Shapes(i)

        if not shape.HasTextFrame:

            continue

        if not shape.Name not in names:

            continue

        text_range = shape.TextFrame.TextRange

        text_range.Font.Name = name_font





def Fix_name_Font_one_silde():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    slide = powerpoint.ActiveWindow.View.Slide

    name_font = simpledialog.askstring('Nhập tên font', 'copy tên font dán vào đây:')

    Fix_name_Font(slide, name_font)





def Fix_name_Font_All_silde():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    name_font = simpledialog.askstring('Nhập tên font', 'copy tên font dán vào đây:')

    for slide in presentation.Slides:

        Fix_name_Font(slide, name_font)





def Canh_dap_an_one_slide_4_dong():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    slide = powerpoint.ActiveWindow.View.Slide

    Fix_canh_dap_an_for(powerpoint, slide)





def Canh_dap_an_one_slide_2_dong():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    slide = powerpoint.ActiveWindow.View.Slide

    max_len_da = max_Dap_an(slide)

    if max_len_da != 0:

        fix_len(slide, 'Dap_an_a', 460)

        fix_len(slide, 'Dap_an_b', 460)

        fix_len(slide, 'Dap_an_c', 460)

        fix_len(slide, 'Dap_an_d', 460)

        vitri = 455

        a = slide.Shapes('Dap_an_a').Height

        b = slide.Shapes('Dap_an_b').Height

        c = slide.Shapes('Dap_an_c').Height

        d = slide.Shapes('Dap_an_d').Height

        vitri = vitri - max(c, d)

        move_shape_to(slide, 'Dap_an_c', 10, vitri)

        move_shape_to(slide, 'Dap_an_d', 490, vitri)

        vitri = vitri - max(a, b) - 10

        move_shape_to(slide, 'Dap_an_a', 10, vitri)

        move_shape_to(slide, 'Dap_an_b', 490, vitri)

        return None

    return None

# WARNING: Decompyle incomplete





def Canh_dap_an_one_slide_1_dong():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    slide = powerpoint.ActiveWindow.View.Slide

    fix_len(slide, 'Dap_an_a', 220)

    fix_len(slide, 'Dap_an_b', 220)

    fix_len(slide, 'Dap_an_c', 220)

    fix_len(slide, 'Dap_an_d', 220)

    vitri = 450

    a = slide.Shapes('Dap_an_a').Height

    b = slide.Shapes('Dap_an_b').Height

    c = slide.Shapes('Dap_an_c').Height

    d = slide.Shapes('Dap_an_d').Height

    vitri = vitri - max(a, b, c, d)

    move_shape_to(slide, 'Dap_an_a', 10, vitri)

    move_shape_to(slide, 'Dap_an_b', 250, vitri)

    move_shape_to(slide, 'Dap_an_c', 490, vitri)

    move_shape_to(slide, 'Dap_an_d', 730, vitri)

    return None

# WARNING: Decompyle incomplete





def Canh_dap_an_All_slide():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    for slide in presentation.Slides:

        Fix_canh_dap_an_for(powerpoint, slide)





def Fix_result_at():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    slide = powerpoint.ActiveWindow.View.Slide

    Fix_result_for(powerpoint, slide)





def slide_show_start_1():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    slide_show = presentation.SlideShowSettings

    slide_show.StartingSlide = 1

    slide_show.EndingSlide = presentation.Slides.Count

    slide_show.AdvanceMode = 2

    slide_show.RangeType = 1

    slide_show.LoopUntilStopped = False

    slide_show.Run()





def slide_show_at_now():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    view = powerpoint.ActiveWindow.View

    current_slide_index = view.Slide.SlideIndex

    slide_show = presentation.SlideShowSettings

    slide_show.StartingSlide = 1

    slide_show.EndingSlide = presentation.Slides.Count

    slide_show.AdvanceMode = 2

    slide_show.RangeType = 1

    slide_show.LoopUntilStopped = False

    slide_show.Run()

    time.sleep(1)

    if presentation.SlideShowWindow:

        presentation.SlideShowWindow.View.GotoSlide(current_slide_index)

        return None





def go_next_slide():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    slide_show = presentation.SlideShowWindow

    slide_show.View.Next()





def go_Previous_slide():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    slide_show = presentation.SlideShowWindow

    slide_show.View.Previous()





def exit_slide_show():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    slide_show = presentation.SlideShowWindow

    slide_show.View.Exit()

    return None

# WARNING: Decompyle incomplete





def open_pen():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    slide_show = presentation.SlideShowWindow

    slide_show.View.PointerType = 2





def open_Eraser():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    slide_show = presentation.SlideShowWindow

    slide_show.View.PointerType = 5





def return_mouse():

    powerpoint = vbf.khoi_tao_powerpoint_2()

    presentation = powerpoint.ActivePresentation

    slide_show = presentation.SlideShowWindow

    slide_show.View.PointerType = 1





def xem_thong_tin_Allshape(slide_index):

    powerpoint = win32com.client.Dispatch('PowerPoint.Application')

    presentation = powerpoint.ActivePresentation

    slide = presentation.Slides(slide_index)

    print(f'''Slide {slide.SlideIndex}:''')

    for shape in slide.Shapes:

        shape_name = shape.Name

        shape_width = shape.Width

        shape_height = shape.Height

        print(f''' - Shape: {shape_name}''')

        print(f'''   + Width: {shape_width} points''')

        print(f'''   + Height: {shape_height} points''')





def xem_vi_tri_one_shape(slide_index, shape_name):

    powerpoint = win32com.client.Dispatch('PowerPoint.Application')

    slide = powerpoint.ActivePresentation.Slides(slide_index)

    shape = None

    for shp in slide.Shapes:

        if not shp.Name == shape_name:

            continue

        shape = shp

        slide.Shapes

    if shape:

        left = shape.Left

        top = shape.Top

        width = shape.Width

        height = shape.Height

        print(f'''Shape \'{shape_name}\' có vị trí: Left = {left}, Top = {top}, Width = {width}, Height = {height}.''')

        return (left, top, width, height)

    None(f'''Shape \'{shape_name}\' không tồn tại trên slide {slide_index}.''')



