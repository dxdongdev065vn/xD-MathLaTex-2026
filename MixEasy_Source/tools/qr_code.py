import os

import qrcode

import openpyxl

from openpyxl import load_workbook

from openpyxl import Workbook

from openpyxl.styles import Font

from openpyxl.styles import Alignment

from openpyxl.styles import Border, Side

from openpyxl.styles import PatternFill

from openpyxl.styles import GradientFill

from openpyxl.utils import get_column_letter

import tkinter as tk

from tkinter import ttk, Label, filedialog, Entry, Button, Frame, Listbox, Scrollbar, messagebox

from tkinter import simpledialog

import webbrowser



def Qr_tool(root):

    pass

# WARNING: Decompyle incomplete





def extract_data_from_excel(file_path):

    workbook = load_workbook(file_path)

    sheet = workbook.active

    so_de = 0

# WARNING: Decompyle incomplete





def TNmaker_qr_code(text, output_file, box_size, border = (10, 4)):

    qr = qrcode.QRCode(version = 1, error_correction = qrcode.constants.ERROR_CORRECT_L, box_size = box_size, border = border)

    qr.add_data(text)

    qr.make(fit = True)

    img = qr.make_image(fill_color = 'black', back_color = 'white')

    img.save(output_file)





def TNmaker_qr_code_View(text, output_file, box_size, border = (10, 4)):

    qr = qrcode.QRCode(version = 1, error_correction = qrcode.constants.ERROR_CORRECT_L, box_size = box_size, border = border)

    qr.add_data(text)

    qr.make(fit = True)

    img = qr.make_image(fill_color = 'black', back_color = 'white')

    img.show()





def creat_Qr_TNmaker():

    pass

# WARNING: Decompyle incomplete





def create_qr_code(link):

    '''Tạo QR code từ link và lưu dưới dạng qrcode.png trong thư mục hiện tại.'''

    if not link:

        messagebox.showwarning('Cảnh báo', 'Vui lòng nhập một đường link!')

        return None

    qr = qrcode.QRCode(version = 1, error_correction = qrcode.constants.ERROR_CORRECT_L, box_size = 10, border = 4)

    qr.add_data(link)

    qr.make(fit = True)

    img = qr.make_image(fill = 'black', back_color = 'white')

    output_file = os.path.join(os.getcwd(), 'qrcode.png')

    img.save(output_file)

    messagebox.showinfo('Thành công', f'''QR code đã được lưu tại:\n{output_file}''')





def create_qr_code_View(link):

    '''Tạo QR code từ link và lưu dưới dạng qrcode.png trong thư mục hiện tại.'''

    if not link:

        messagebox.showwarning('Cảnh báo', 'Vui lòng nhập một đường link!')

        return None

    qr = qrcode.QRCode(version = 1, error_correction = qrcode.constants.ERROR_CORRECT_L, box_size = 10, border = 4)

    qr.add_data(link)

    qr.make(fit = True)

    img = qr.make_image(fill = 'black', back_color = 'white')

    img.show()





def generate_Qr_link():

    '''Mở cửa sổ nhập link và tạo QR code.'''

    pass

# WARNING: Decompyle incomplete



